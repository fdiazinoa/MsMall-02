
# Backend: FastAPI API para MSMALL Audit
import asyncio
import base64
import csv
import html
import io
import logging
import socket
import time
import threading
import re
import secrets
import unicodedata
import weakref
from datetime import date, datetime, timedelta
from typing import List, Optional, Dict, Any, Tuple, Set
from uuid import uuid4

from fastapi import FastAPI, HTTPException, Header, UploadFile, File, Depends, Query, Request, status, Body
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd
from thefuzz import process, fuzz
import paramiko
import json
import xmltodict
from ftplib import FTP
import stat
import urllib.error
import urllib.request
from worker_importacion import (
    api_provider_name,
    fetch_bundaberg_sales,
    fetch_studio_g_sales,
    process_webservice_import,
    run_worker_async,
)
from analytics_service import AnalyticsService
from routers import recipes, comparisons, admin_tools, big_data
from routers.token_auth import (
    AuthContext as TokenAuthContext,
    TOKEN_TYPE_EXPORTER,
    ACTIVE as TOKEN_ACTIVE,
    DISABLED as TOKEN_DISABLED,
    REVOKED as TOKEN_REVOKED,
    CreateServiceAccountRequest as TokenCreateServiceAccountRequest,
    CreateTokenRequest as TokenCreateTokenRequest,
    PatchServiceAccountStatusRequest as TokenPatchServiceAccountStatusRequest,
    PatchTokenStatusRequest as TokenPatchTokenStatusRequest,
    RevokeMallRequest as TokenRevokeMallRequest,
    RevokeLocalRequest as TokenRevokeLocalRequest,
    RevokeRequest as TokenRevokeRequest,
    RevokeServiceAccountTokensRequest as TokenRevokeServiceAccountTokensRequest,
    UpsertExporterWebserviceConfigRequest as TokenUpsertExporterWebserviceConfigRequest,
    _hash_token as token_auth_hash_token,
    _parse_scopes as token_auth_parse_scopes,
    build_default_service as build_token_auth_service,
    create_router as create_token_auth_router,
    require_token_auth,
    request_explicit_never_expires as token_auth_request_explicit_never_expires,
    sanitize_exporter_webservice_config_row as sanitize_token_exporter_webservice_config_row,
    sanitize_service_account_row as sanitize_token_service_account_row,
    sanitize_token_row as sanitize_token_auth_row,
    utcnow as token_auth_utcnow,
    validate_exporter_payload_mapping,
)
from services.sensitive_ops_service import SensitiveOpsService, sanitize_error_text as sanitize_sensitive_ops_error
from services.local_custom_fields_service import LocalCustomFieldsService
from services.big_data_sprint2_service import BigDataSprint2Service
from services.connection_monitor_service import (
    ConnectionMonitorService,
    RetryPolicyBlocked,
)
from services.date_parsing_service import normalize_sale_date
from services.dashboard_analytics_service import DashboardAnalyticsService
from services.load_log_service import build_load_log_payload, insert_load_log_row
from services.missing_days_email_service import (
    DEFAULT_CONSOLIDATED_BODY_TEMPLATE,
    DEFAULT_CONSOLIDATED_SUBJECT_TEMPLATE,
    DEFAULT_MISSING_DAYS_BODY_TEMPLATE,
    DEFAULT_MISSING_DAYS_SUBJECT_TEMPLATE,
    MISSING_DAYS_CONSOLIDATED_NOTIFICATION_TYPE,
    MISSING_DAYS_NOTIFICATION_TYPE,
    MISSING_DAYS_NOTIFICATION_TYPES,
    build_missing_days_email_html,
    send_missing_days_emails_for_mall,
)
from supabase import create_client, Client
import os
from dotenv import load_dotenv

load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL")
# Prefer Service Role Key for backend operations to bypass RLS
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
SUPABASE_KEY = SUPABASE_SERVICE_ROLE_KEY or os.getenv("SUPABASE_KEY")
supabase: Optional[Client] = None


# Setup Logger first so we can see errors
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("msmall-api")

def _parse_bool_env(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}

def _is_api_scheduler_enabled() -> bool:
    return _parse_bool_env("ENABLE_API_SCHEDULER", default=False)
_CORS_LOCK_V1_ORIGINS = ["https://msmall.vercel.app"]
# Preview URLs are generated per deployment, while production keeps the stable
# origin above. Restrict direct API access to this Vercel project instead of
# opening CORS to arbitrary `*.vercel.app` origins.
_CORS_LOCK_V1_ORIGIN_REGEX = r"https://msmall-[a-z0-9-]+-felix-diaz-s-projects\.vercel\.app"

if SUPABASE_URL and SUPABASE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        logger.info(f"Supabase Client initialized (Service Role: {'Yes' if os.getenv('SUPABASE_SERVICE_ROLE_KEY') else 'No'})")
    except Exception as e:
        logger.error(f"CRITICAL: Failed to initialize Supabase client: {e}")
        # Dont crash, just continue without supabase

SYSTEM_ADMIN_EMAILS = {
    e.strip().lower()
    for e in (os.getenv("SYSTEM_ADMIN_EMAILS", "fdiaz@mercasend.net")).split(",")
    if e and e.strip()
}
RESEND_API_KEY_ENV = "RESEND_API_KEY"
RESEND_DOMAIN = "mercasend.net"
RESEND_FROM_EMAIL = "notificaciones@mercasend.net"
RESEND_FROM_NAME = "MercaSend Notificaciones"
RESEND_USER_AGENT = "MSMALL-API/1.0 (mercasend.net)"
RESEND_SENDER_EMAIL_KEY = "RESEND_FROM_EMAIL"
RESEND_SENDER_NAME_KEY = "RESEND_FROM_NAME"
COPILOT_ENABLED_KEY = "COPILOT_ENABLED"
COPILOT_PROVIDER_KEY = "COPILOT_PROVIDER"
COPILOT_OPENAI_API_KEY_KEY = "COPILOT_OPENAI_API_KEY"
COPILOT_GEMINI_API_KEY_KEY = "COPILOT_GEMINI_API_KEY"
COPILOT_OPENAI_MODEL_KEY = "COPILOT_OPENAI_MODEL"
COPILOT_GEMINI_MODEL_KEY = "COPILOT_GEMINI_MODEL"
COPILOT_DEFAULT_MODELS = {
    "openai": "gpt-4o-mini",
    "gemini": "gemini-1.5-flash",
}
ADMIN_ROLES = {"admin", "superadmin", "super_admin", "administrador"}
IT_ROLES = {"it", "tic"}

RBAC_ACTIONS = {"view", "create", "update", "delete"}
FACTORY_ROLE_PERMISSIONS: Dict[str, Dict[str, Dict[str, bool]]] = {
    "admin": {module: {action: True for action in RBAC_ACTIONS} for module in (
        "dashboard", "sales_reports", "stores", "imports", "monitor", "financial",
        "cube", "comparisons", "malls", "users", "roles",
    )},
    "it": {
        "dashboard": {"view": True}, "sales_reports": {"view": True, "create": True},
        "stores": {"view": True, "create": True, "update": True},
        "imports": {"view": True, "create": True, "update": True},
        "monitor": {"view": True, "create": True, "update": True},
        "financial": {"view": True}, "cube": {"view": True}, "comparisons": {"view": True},
    },
    "auditor": {
        "dashboard": {"view": True}, "sales_reports": {"view": True},
        "monitor": {"view": True}, "financial": {"view": True}, "cube": {"view": True},
        "comparisons": {"view": True},
    },
    "visualizador": {
        "dashboard": {"view": True}, "sales_reports": {"view": True},
        "financial": {"view": True}, "cube": {"view": True}, "comparisons": {"view": True},
    },
}

def _sensitive_ops_service() -> SensitiveOpsService:
    return SensitiveOpsService(supabase, logger)


def _local_custom_fields_service() -> LocalCustomFieldsService:
    return LocalCustomFieldsService(supabase, logger)

def _connection_monitor_service() -> ConnectionMonitorService:
    return ConnectionMonitorService(supabase, logger)


def _analytics_service() -> AnalyticsService:
    return AnalyticsService(supabase_client=supabase, logger=logger)


async def _run_local_risk_analysis_async(local_id: Optional[str], trigger: str) -> Optional[Dict[str, Any]]:
    if not supabase or not local_id:
        return None
    try:
        return await asyncio.to_thread(
            lambda: _analytics_service().run_and_persist_local_analysis(local_id, trigger=trigger)
        )
    except Exception as exc:
        logger.error("Error ejecutando analisis IA para local %s: %s", local_id, exc)
        return None


# --- LIGHTWEIGHT IN-MEMORY CACHE (TTL) ---
_CACHE: Dict[str, Dict[str, Any]] = {}
_CACHE_LOCK = threading.Lock()
_CACHE_MISS = object()
_DASHBOARD_LOAD_LOCKS: weakref.WeakValueDictionary[str, asyncio.Lock] = weakref.WeakValueDictionary()
_DASHBOARD_LOAD_LOCKS_GUARD = threading.Lock()
_INFLIGHT_MANUAL_EXEC: set = set()
_INFLIGHT_MANUAL_EXEC_LOCK = threading.Lock()
_COPILOT_DOWNLOADS: Dict[str, Dict[str, Any]] = {}
_COPILOT_DOWNLOADS_LOCK = threading.Lock()
_COPILOT_DOWNLOAD_TTL_SECONDS = 15 * 60
_COPILOT_EMAIL_DRAFTS: Dict[str, Dict[str, Any]] = {}
_COPILOT_EMAIL_DRAFTS_LOCK = threading.Lock()

def _env_int(name: str, default: int, min_value: int = 1, max_value: int = 3600) -> int:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        value = int(raw)
        if value < min_value:
            return min_value
        if value > max_value:
            return max_value
        return value
    except (TypeError, ValueError):
        return default

_CACHE_MAX_ITEMS = _env_int("CACHE_MAX_ITEMS", 300, min_value=50, max_value=5000)
STUDIO_G_PREVIEW_HISTORY_DAYS = _env_int(
    "STUDIO_G_PREVIEW_HISTORY_DAYS",
    120,
    min_value=7,
    max_value=730,
)

# Endpoint-specific TTLs (seconds), configurable via environment variables.
TTL_DASHBOARD = _env_int("CACHE_TTL_DASHBOARD", 90, min_value=5, max_value=1800)
TTL_RANKING = _env_int("CACHE_TTL_RANKING", 60, min_value=5, max_value=1800)
TTL_HEATMAP = _env_int("CACHE_TTL_HEATMAP", 120, min_value=5, max_value=1800)

def _cache_get(key: str):
    now = time.time()
    with _CACHE_LOCK:
        item = _CACHE.get(key)
        if not item:
            return _CACHE_MISS
        if item["expires_at"] <= now:
            _CACHE.pop(key, None)
            return _CACHE_MISS
        return item["value"]

def _cache_set(key: str, value: Any, ttl: int):
    now = time.time()
    with _CACHE_LOCK:
        # Opportunistic cleanup to keep memory bounded.
        if len(_CACHE) >= _CACHE_MAX_ITEMS:
            expired = [k for k, v in _CACHE.items() if v["expires_at"] <= now]
            for k in expired:
                _CACHE.pop(k, None)
            # If still full, remove oldest expiration first.
            if len(_CACHE) >= _CACHE_MAX_ITEMS:
                oldest_key = min(_CACHE.keys(), key=lambda k: _CACHE[k]["expires_at"])
                _CACHE.pop(oldest_key, None)
        _CACHE[key] = {"value": value, "expires_at": now + max(1, ttl)}

def _cache_delete_prefix(prefix: str) -> int:
    with _CACHE_LOCK:
        matching_keys = [key for key in _CACHE if key.startswith(prefix)]
        for key in matching_keys:
            _CACHE.pop(key, None)
        return len(matching_keys)


def _dashboard_mode() -> str:
    return DashboardAnalyticsService.normalize_mode(os.getenv("DASHBOARD_KPI_MODE", "legacy"))


def _dashboard_load_lock(cache_key: str) -> asyncio.Lock:
    with _DASHBOARD_LOAD_LOCKS_GUARD:
        lock = _DASHBOARD_LOAD_LOCKS.get(cache_key)
        if lock is None:
            lock = asyncio.Lock()
            _DASHBOARD_LOAD_LOCKS[cache_key] = lock
        return lock


def _invalidate_dashboard_cache(mall_id: Optional[str]) -> int:
    normalized_mall_id = str(mall_id or "").strip()
    if not normalized_mall_id:
        return 0
    mall_marker = f":{normalized_mall_id}:"
    with _CACHE_LOCK:
        matching_keys = [
            key
            for key in _CACHE
            if key.startswith("analytics:dashboard:") and mall_marker in key
        ]
        for key in matching_keys:
            _CACHE.pop(key, None)
        return len(matching_keys)


def flatten_json(y):
    out = {}

    def flatten(x, name=''):
        if isinstance(x, dict):
            for a in x:
                flatten(x[a], name + a + '.')
        elif isinstance(x, list):
             out[name[:-1]] = json.dumps(x)
        else:
            out[name[:-1]] = x

    flatten(y)
    return out

def diagnosticar_archivo(file_bytes):
    reporte = []
    
    # PASO 1: Decodificación
    try:
        content = file_bytes.decode('utf-8-sig') # Vital para quitar el BOM
        reporte.append("SUCCESS: Decodificación UTF-8-SIG correcta.")
    except:
        reporte.append("ERROR: Falló decodificación UTF-8-SIG.")
        return reporte

    # PASO 2: JSON Parsing
    try:
        data = json.loads(content)
        keys = list(data.keys()) if isinstance(data, dict) else ["<Lista>"]
        reporte.append(f"SUCCESS: JSON Válido. Claves raíz: {keys}")
    except Exception as e:
        reporte.append(f"ERROR: json.loads falló. {str(e)}")
        return reporte

    # PASO 3: Detección de Lista
    target_data = data
    if isinstance(data, dict):
        if "invoices" in data:
            target_data = data["invoices"]
            reporte.append("INFO: Se detectó clave 'invoices' y se entró en ella.")
        else:
             # Try smart search
            found = False
            for k, v in data.items():
                if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                    target_data = v
                    reporte.append(f"INFO: Se detectó lista en clave '{k}'.")
                    found = True
                    break
            if not found:
                 reporte.append("WARN: No se encontró lista 'invoices' ni otra lista candidata.")
    
    if not isinstance(target_data, list):
        reporte.append(f"ERROR: Los datos no son una lista. Son tipo: {type(target_data)}")
        # Try wrapping
        reporte.append("INFO: Intentando envolver en lista...")
        target_data = [target_data]

    # PASO 4: Pandas Normalize
    try:
        df = pd.json_normalize(target_data)
        cols = list(df.columns)
        reporte.append(f"SUCCESS: DataFrame creado con {len(df)} filas.")
        reporte.append(f"COLUMNAS DETECTADAS: {cols}")
        
        # Muestra una fila de ejemplo para ver si los datos estan anidados
        if not df.empty:
             reporte.append(f"EJEMPLO FILA 1: {df.iloc[0].to_dict()}")

    except Exception as e:
        reporte.append(f"ERROR: pd.json_normalize falló. {str(e)}")

    return reporte



app = FastAPI(title="MSMALL Sales Audit API", version="1.0.2")

@app.post("/api/v1/debug/diagnose-file")
async def diagnose_file_endpoint(file: UploadFile = File(...)):
    """
    Diagnostic endpoint to inspect JSON file structure and parsing status.
    """
    try:
        content = await file.read()
        report = diagnosticar_archivo(content)
        return {"filename": file.filename, "report": report}
    except Exception as e:
        return {"error": str(e)}

app.include_router(recipes.router)
app.include_router(comparisons.router)
app.include_router(admin_tools.router)
app.include_router(big_data.router)
app.include_router(create_token_auth_router())
_api_scheduler_task = None

async def scheduler_loop():
    await asyncio.sleep(10) # Initial delay
    while True:
        logger.info("[Scheduler] Iniciando ciclo de importación automática...")
        try:
            # Ahora el worker es nativamente async
            await run_worker_async()
        except Exception as e:
            logger.error(f"[Scheduler] Error en ciclo: {e}")
        logger.info("[Scheduler] Durmiendo 1 hora...")
        await asyncio.sleep(3600)

@app.on_event("startup")
async def startup_event():
    global _api_scheduler_task
    logger.info("MSMALL API Starting up... routes loaded.")
    api_scheduler_enabled = _is_api_scheduler_enabled()
    logger.info("API scheduler enabled: %s", str(api_scheduler_enabled).lower())
    if api_scheduler_enabled:
        _api_scheduler_task = asyncio.create_task(scheduler_loop())
    else:
        logger.info("API embedded scheduler disabled; worker is the scheduler authority.")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_CORS_LOCK_V1_ORIGINS,
    allow_origin_regex=_CORS_LOCK_V1_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger.info("CORS_LOCK_V1 origins=%s origin_regex=%s", _CORS_LOCK_V1_ORIGINS, _CORS_LOCK_V1_ORIGIN_REGEX)

# --- SECURITY & MULTI-TENANT MIDDLEWARE ---
security = HTTPBearer()

def _normalize_role(role: Optional[str]) -> str:
    return (role or "").strip().lower().replace("-", "_").replace(" ", "_")

def _is_system_admin_email(email: Optional[str]) -> bool:
    if not email:
        return False
    return email.strip().lower() in SYSTEM_ADMIN_EMAILS

def _extract_auth_user(auth_result: Any) -> Optional[Any]:
    if auth_result is None:
        return None
    user = getattr(auth_result, "user", None)
    if user is not None:
        return user
    if isinstance(auth_result, dict):
        if auth_result.get("user") is not None:
            return auth_result.get("user")
        data = auth_result.get("data")
        if isinstance(data, dict) and data.get("user") is not None:
            return data.get("user")
    return auth_result

def _user_field(user: Any, field: str, default: Any = None) -> Any:
    if user is None:
        return default
    if isinstance(user, dict):
        return user.get(field, default)
    return getattr(user, field, default)

def _resolve_effective_role(email: Optional[str], role_candidates: List[Optional[str]]) -> str:
    if _is_system_admin_email(email):
        return "admin"

    normalized = [_normalize_role(r) for r in role_candidates if r]
    if any(r in ADMIN_ROLES for r in normalized):
        return "admin"
    if any(r in IT_ROLES for r in normalized):
        return "it"
    if any(r == "auditor" for r in normalized):
        return "auditor"
    if any(r in {"visualizador", "viewer"} for r in normalized):
        return "visualizador"
    return "auditor"

def _canonical_admin_role(raw_role: Optional[str]) -> str:
    """
    Normalizes incoming admin-editable roles to the allowed canonical set.
    """
    normalized = _normalize_role(raw_role)
    if normalized in ADMIN_ROLES:
        return "admin"
    if normalized in IT_ROLES:
        return "it"
    if normalized == "auditor":
        return "auditor"
    if normalized in {"visualizador", "viewer"}:
        return "visualizador"
    return ""

def _parse_auth_users_result(auth_users_result: Any) -> List[Any]:
    if auth_users_result is None:
        return []
    if isinstance(auth_users_result, list):
        return auth_users_result

    users = getattr(auth_users_result, "users", None)
    if isinstance(users, list):
        return users

    data = getattr(auth_users_result, "data", None)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        maybe_users = data.get("users")
        if isinstance(maybe_users, list):
            return maybe_users

    if isinstance(auth_users_result, dict):
        maybe_users = auth_users_result.get("users")
        if isinstance(maybe_users, list):
            return maybe_users
        nested_data = auth_users_result.get("data")
        if isinstance(nested_data, list):
            return nested_data
        if isinstance(nested_data, dict) and isinstance(nested_data.get("users"), list):
            return nested_data.get("users")

    return []

def _list_all_auth_users(per_page: int = 100, max_pages: int = 50) -> List[Any]:
    all_users: List[Any] = []

    # Preferred path: paginated fetch to avoid silently missing older users.
    try:
        for page in range(1, max_pages + 1):
            chunk = _parse_auth_users_result(
                supabase.auth.admin.list_users(page=page, per_page=per_page)
            )
            if not chunk:
                break
            all_users.extend(chunk)
            if len(chunk) < per_page:
                break
    except TypeError:
        # Some client versions may not support pagination kwargs.
        pass
    except Exception as e:
        logger.warning(f"Error paginando usuarios de auth: {e}")

    if not all_users:
        try:
            all_users = _parse_auth_users_result(supabase.auth.admin.list_users())
        except Exception as e:
            logger.error(f"Error listando usuarios de auth: {e}")
            return []

    deduped: Dict[str, Any] = {}
    for u in all_users:
        uid = _user_field(u, "id")
        if uid:
            deduped[uid] = u
    return list(deduped.values())

def _find_auth_user_by_email(email: str) -> Optional[Any]:
    target = (email or "").strip().lower()
    if not target:
        return None
    for u in _list_all_auth_users():
        if (_user_field(u, "email") or "").strip().lower() == target:
            return u
    return None

async def get_current_user_id(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        # Verify token with Supabase
        user = supabase.auth.get_user(token)
        if not user or not user.user:
            raise HTTPException(status_code=401, detail="Invalid token")
        return user.user.id
    except Exception as e:
        logger.error(f"Auth error: {e}")
        raise HTTPException(status_code=401, detail="Authentication failed")

async def _get_access_context(user_id: str) -> Dict[str, Any]:
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")

    email = None
    metadata_role = None
    profile_role = None
    mall_roles: List[str] = []

    try:
        auth_result = supabase.auth.admin.get_user_by_id(user_id)
        auth_user = _extract_auth_user(auth_result)
        email = _user_field(auth_user, "email")
        user_metadata = _user_field(auth_user, "user_metadata", {}) or {}
        if isinstance(user_metadata, dict):
            metadata_role = user_metadata.get("rol") or user_metadata.get("role")
    except Exception as e:
        logger.warning(f"No se pudo cargar auth user para validar admin: {e}")

    try:
        prof = supabase.table("profiles").select("role").eq("id", user_id).maybe_single().execute()
        if prof and prof.data:
            profile_role = prof.data.get("role")
    except Exception as e:
        logger.warning(f"No se pudo cargar role de profiles para {user_id}: {e}")

    try:
        roles_res = supabase.table("usuarios_malls").select("rol").eq("usuario_id", user_id).execute()
        mall_roles = [r.get("rol") for r in (roles_res.data or []) if r.get("rol")]
    except Exception as e:
        logger.warning(f"No se pudo cargar roles de usuarios_malls para {user_id}: {e}")

    effective_role = _resolve_effective_role(email, [profile_role, metadata_role, *mall_roles])
    role_key = effective_role
    role_name = effective_role.title()
    permissions = FACTORY_ROLE_PERMISSIONS.get(effective_role, FACTORY_ROLE_PERMISSIONS["auditor"])
    has_assignment = False

    # The RBAC tables are intentionally optional during rollout so current users keep access
    # until the migration is applied. Once assigned, configurable permissions are authoritative.
    try:
        assignment = (
            supabase.table("profile_role_assignments")
            .select("role_id")
            .eq("user_id", user_id)
            .maybe_single()
            .execute()
        )
        if assignment and assignment.data and assignment.data.get("role_id"):
            role_id = assignment.data["role_id"]
            role_res = supabase.table("app_roles").select("id,key,nombre").eq("id", role_id).maybe_single().execute()
            if role_res and role_res.data:
                role_key = str(role_res.data.get("key") or effective_role)
                role_name = str(role_res.data.get("nombre") or role_key)
                rows = supabase.table("app_role_permissions").select(
                    "module_key,can_view,can_create,can_update,can_delete"
                ).eq("role_id", role_id).execute().data or []
                permissions = {
                    str(row["module_key"]): {
                        "view": bool(row.get("can_view")),
                        "create": bool(row.get("can_create")),
                        "update": bool(row.get("can_update")),
                        "delete": bool(row.get("can_delete")),
                    }
                    for row in rows
                }
                has_assignment = True
    except Exception as e:
        logger.warning("No se pudo cargar RBAC para %s: %s", user_id, e)

    if _is_system_admin_email(email):
        role_key = "admin"
        role_name = "Administrador"
        permissions = FACTORY_ROLE_PERMISSIONS["admin"]
        has_assignment = False
    return {
        "user_id": user_id,
        "email": email,
        "role": role_key,
        "role_name": role_name,
        "legacy_role": effective_role,
        "permissions": permissions,
        "has_role_assignment": has_assignment,
    }

def _has_module_permission(access_ctx: Dict[str, Any], module_key: str, action: str) -> bool:
    if action not in RBAC_ACTIONS:
        return False
    if _is_system_admin_email(access_ctx.get("email")):
        return True
    return bool((access_ctx.get("permissions") or {}).get(module_key, {}).get(action))

def require_module_permission(module_key: str, action: str):
    async def dependency(user_id: str = Depends(get_current_user_id)) -> Dict[str, Any]:
        access_ctx = await _get_access_context(user_id)
        if not _has_module_permission(access_ctx, module_key, action):
            raise HTTPException(status_code=403, detail=f"No tienes permiso para {action} en el módulo {module_key}.")
        return access_ctx
    return dependency

async def require_admin_access(user_id: str = Depends(get_current_user_id)) -> Dict[str, Any]:
    access_ctx = await _get_access_context(user_id)
    if access_ctx["legacy_role"] != "admin" and not _has_module_permission(access_ctx, "users", "update"):
        raise HTTPException(status_code=403, detail="Permisos insuficientes. Se requiere rol ADMIN.")
    return access_ctx

async def require_it_or_admin_access(user_id: str = Depends(get_current_user_id)) -> Dict[str, Any]:
    access_ctx = await _get_access_context(user_id)
    if access_ctx["role"] not in {"admin", "it"}:
        raise HTTPException(status_code=403, detail="Permisos insuficientes. Se requiere rol IT o ADMIN.")
    return access_ctx

def require_pending_import_monitor_access(
    internal_token: Optional[str] = Header(default=None, alias="X-MsMall-Internal-Token")
) -> Dict[str, Any]:
    """Authenticate the Railway pending-import monitor with a dedicated secret."""
    expected = str(os.getenv("PENDING_IMPORT_MONITOR_TOKEN") or "").strip()
    if len(expected) < 32:
        logger.error("PENDING_IMPORT_MONITOR_TOKEN is missing or shorter than 32 characters")
        raise HTTPException(
            status_code=503,
            detail="Autenticación interna del monitor no configurada.",
        )

    provided = str(internal_token or "").strip()
    if not provided or not secrets.compare_digest(provided, expected):
        raise HTTPException(status_code=401, detail="Credencial interna inválida.")

    return {"source": "pending_import_monitor", "role": "internal_service"}

async def require_audit_read_access(user_id: str = Depends(get_current_user_id)) -> Dict[str, Any]:
    access_ctx = await _get_access_context(user_id)
    if access_ctx["role"] not in {"admin", "it", "auditor"}:
        raise HTTPException(status_code=403, detail="Permisos insuficientes para consultar logs.")
    return access_ctx

def _get_user_mall_ids(user_id: str) -> List[str]:
    if not user_id or not supabase:
        return []
    try:
        res = (
            supabase.table("usuarios_malls")
            .select("mall_id")
            .eq("usuario_id", user_id)
            .execute()
        )
        return [row.get("mall_id") for row in (res.data or []) if row.get("mall_id")]
    except Exception as e:
        logger.warning(f"No se pudo cargar malls del usuario {user_id}: {e}")
        return []

def _ensure_operator_can_access_mall(operator_ctx: Dict[str, Any], mall_id: Optional[str]) -> None:
    if operator_ctx.get("role") == "admin":
        return
    if not mall_id:
        raise HTTPException(status_code=400, detail="La configuración no tiene mall_id asignado.")
    allowed_malls = _get_user_mall_ids(operator_ctx.get("user_id"))
    if mall_id not in allowed_malls:
        raise HTTPException(status_code=403, detail="No tienes permisos para operar sobre este mall.")

def _load_local_config_with_access(local_id: str, operator_ctx: Dict[str, Any]) -> Dict[str, Any]:
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")
    try:
        res = (
            supabase.table("locales")
            .select("*")
            .eq("id", local_id)
            .maybe_single()
            .execute()
        )
        local_cfg = res.data
    except Exception as e:
        logger.error(f"Error consultando local {local_id}: {e}")
        raise HTTPException(status_code=500, detail="Error consultando configuración del local")

    if not local_cfg:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")

    _ensure_operator_can_access_mall(operator_ctx, local_cfg.get("mall_id"))
    return local_cfg

def _load_local_config_for_exporter(local_id: str, exporter_ctx: TokenAuthContext) -> Dict[str, Any]:
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")
    try:
        res = (
            supabase.table("locales")
            .select("*")
            .eq("id", local_id)
            .maybe_single()
            .execute()
        )
        local_cfg = res.data
    except Exception as e:
        logger.error(f"Error consultando local {local_id} para exporter: {e}")
        raise HTTPException(status_code=500, detail="Error consultando configuración del local")

    if not local_cfg:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")

    validate_exporter_payload_mapping(
        str(local_cfg.get("mall_id") or ""),
        str(local_cfg.get("id") or local_id),
        exporter_ctx,
    )
    return local_cfg

def _load_local_config_for_pending_monitor(local_id: str) -> Dict[str, Any]:
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")
    try:
        res = (
            supabase.table("locales")
            .select("*")
            .eq("id", local_id)
            .maybe_single()
            .execute()
        )
        local_cfg = res.data
    except Exception as e:
        logger.error("Error consultando local para pending monitor %s: %s", local_id, e)
        raise HTTPException(status_code=500, detail="Error consultando configuración del local")

    if not local_cfg:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    if str(local_cfg.get("tipo_ejecucion") or "").strip().upper() != "AUTOMATICO":
        raise HTTPException(
            status_code=409,
            detail="El monitor solo puede ejecutar importadores automáticos.",
        )
    return local_cfg

def _reactivate_local_after_success(config_data: Dict[str, Any], *, source: str) -> None:
    """Reactivate only a local whose remote import has just proven successful."""
    status_value = str(config_data.get("processing_status") or "").strip().upper()
    failure_count = int(config_data.get("consecutive_failures") or 0)
    if status_value != "SUSPENDED_AUTH_ERROR" and failure_count < 5:
        return

    local_id = config_data.get("id")
    if not local_id:
        return
    try:
        _sensitive_ops_service().reactivate_local_processing(
            local_id=str(local_id),
            operator_ctx={"user_id": None, "role": "admin"},
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
            audit_metadata={"source": source, "automatic_reactivation": True},
        )
        logger.info(
            "Local %s reactivated after successful import source=%s",
            local_id,
            source,
        )
    except Exception as e:
        logger.warning(
            "Successful import could not reactivate local %s: %s",
            local_id,
            sanitize_sensitive_ops_error(e),
        )

def _apply_runtime_import_overrides(base_config: Dict[str, Any], runtime_config: Optional[Any]) -> Dict[str, Any]:
    if not runtime_config:
        return base_config

    runtime = runtime_config.dict(exclude_unset=True)
    allowed_override_keys = [
        "host", "puerto", "usuario", "password", "ruta_remota", "protocolo", "tipo_archivo",
        "mapping", "constants", "accion_post_procesado", "prefijo_renombrado",
        "sftp_host", "sftp_port", "sftp_user", "sftp_pass", "sftp_path", "sftp_protocol",
        "file_type", "mapping_config", "constants_config", "prefijo_backup",
        "has_header", "data_start_row"
    ]
    for key in allowed_override_keys:
        val = runtime.get(key)
        if val not in (None, ""):
            base_config[key] = val
    return base_config

def _normalize_import_config_payload(config_data: Dict[str, Any]) -> Dict[str, Any]:
    normalized = dict(config_data or {})
    if normalized.get("mapping"):
        normalized["mapping_config"] = dict(normalized["mapping"])
    else:
        normalized["mapping"] = normalized.get("mapping_config") or {}
    if normalized.get("constants"):
        normalized["constants_config"] = dict(normalized["constants"])
    else:
        normalized["constants"] = normalized.get("constants_config") or {}
    if not normalized.get("tipo_archivo"):
        normalized["tipo_archivo"] = normalized.get("file_type", "CSV")
    return normalized

def _parse_bool_value(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in ("true", "1", "si", "sí", "yes", "y"):
        return True
    if text in ("false", "0", "no", "n"):
        return False
    return None

def _extract_parsing_options(config: Optional[Dict[str, Any]]) -> Tuple[Optional[bool], Optional[int]]:
    payload = config or {}
    constants = payload.get("constants") or {}

    has_header = _parse_bool_value(payload.get("has_header"))
    if has_header is None:
        has_header = _parse_bool_value(constants.get("_has_header"))

    data_start_row = payload.get("data_start_row")
    if data_start_row in (None, ""):
        data_start_row = constants.get("_data_start_row")
    try:
        parsed_row = int(data_start_row) if data_start_row not in (None, "") else None
        if parsed_row is not None and parsed_row < 1:
            parsed_row = 1
    except Exception:
        parsed_row = None

    return has_header, parsed_row

def _looks_like_data_cell(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    if re.fullmatch(r"\d+", text):
        return True
    if re.fullmatch(r"\d{8}", text):
        return True
    if re.fullmatch(r"\d+([.,]\d+)?", text):
        return True
    if re.fullmatch(r"[A-Za-z]\d{1,4}", text):
        return True
    if re.fullmatch(r"[A-Za-z0-9_-]{1,6}", text):
        return True
    return False

def _looks_like_header_cell(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    header_tokens = (
        "fecha", "date", "factura", "invoice", "local", "codigo", "code",
        "total", "impuesto", "tax", "neto", "bruto", "monto", "amount", "hora"
    )
    return any(token in text for token in header_tokens)

def _should_treat_as_no_header(content: str, delimiter: str) -> bool:
    try:
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = [r for r in reader if any(str(c or "").strip() for c in r)]
        if not rows:
            return False

        first = rows[0]
        if not first:
            return False

        header_like = sum(1 for c in first if _looks_like_header_cell(c))
        first_data_like = sum(1 for c in first if _looks_like_data_cell(c))
        first_ratio = first_data_like / max(len(first), 1)

        # Single-line files (one transaction) still can be confidently identified as no-header.
        if len(rows) == 1:
            return header_like == 0 and first_ratio >= 0.7

        second = rows[1]
        second_data_like = sum(1 for c in second if _looks_like_data_cell(c))
        second_ratio = second_data_like / max(len(second), 1)

        return header_like == 0 and first_ratio >= 0.7 and second_ratio >= 0.7
    except Exception as e:
        logger.debug(f"No se pudo evaluar encabezado automáticamente: {e}")
        return False

def _detect_delimiter_and_header(sample: str) -> Tuple[str, bool]:
    sniffer = csv.Sniffer()
    delimiter = ","
    has_header = True
    candidates = [",", ";", "\t", "|"]

    try:
        dialect = sniffer.sniff(sample, delimiters="".join(candidates))
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    try:
        lines = [ln for ln in sample.splitlines() if ln.strip()][:10]
        if lines:
            counts = {d: sum(ln.count(d) for ln in lines) for d in candidates}
            best_delim = max(counts, key=counts.get)
            best_count = counts.get(best_delim, 0)
            current_count = counts.get(delimiter, 0)
            if best_count > 0 and (current_count == 0 or best_count >= current_count * 2):
                delimiter = best_delim
    except Exception:
        pass

    try:
        has_header = sniffer.has_header(sample)
    except Exception:
        has_header = True

    return delimiter, has_header

def _build_raw_preview_lines(content: str, max_lines: int = 12, max_chars: int = 220) -> List[str]:
    preview: List[str] = []
    for line in (content or "").splitlines():
        raw = str(line or "").rstrip("\r\n")
        if not raw.strip():
            continue
        if len(raw) > max_chars:
            raw = raw[:max_chars] + "..."
        preview.append(raw)
        if len(preview) >= max_lines:
            break
    return preview

def _clean_csv_header_name(name: Any) -> str:
    return str(name or "").replace("\ufeff", "").strip()

def _normalize_csv_row_keys(row: Dict[Any, Any]) -> Dict[str, Any]:
    if not isinstance(row, dict):
        return {}
    normalized: Dict[str, Any] = {}
    for key, value in row.items():
        clean_key = _clean_csv_header_name(key)
        if not clean_key:
            continue
        if clean_key not in normalized:
            normalized[clean_key] = value
    return normalized

def _clean_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if len(cleaned) >= 2 and cleaned[0] == cleaned[-1] and cleaned[0] in {"'", '"'}:
            cleaned = cleaned[1:-1].strip()
        return cleaned
    return value

def _normalize_text_for_csv(content: str) -> str:
    """
    Normalize uncommon line separators and control chars that can break csv.DictReader
    while still being displayed as multiple lines in previews.
    """
    text = str(content or "")
    # Remove null bytes that can break CSV parsing.
    text = text.replace("\x00", "")
    # Normalize Unicode line separators + classic CR styles to '\n'.
    text = text.replace("\u2028", "\n").replace("\u2029", "\n")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Additional record separators sometimes used by legacy exports.
    for sep in ["\x1e", "\x1d", "\x1c", "\x85", "\x0b", "\x0c"]:
        text = text.replace(sep, "\n")
    # Handle escaped newlines in a single-line payload (e.g., "\\n" literal separators).
    if "\n" not in text and text.count("\\n") >= 1:
        text = text.replace("\\n", "\n")
    return text

def _decode_remote_text(raw_bytes: bytes, is_json: bool = False) -> str:
    if raw_bytes is None:
        return ""

    candidates: List[Tuple[Tuple[int, int, int, int], str, str]] = []
    for enc in ["utf-8-sig", "utf-8", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1"]:
        try:
            decoded = raw_bytes.decode(enc)
        except Exception:
            continue

        replacement_count = decoded.count("�")
        if is_json:
            stripped = decoded.lstrip()
            json_hint = 1 if (stripped.startswith("{") or stripped.startswith("[")) else 0
            score = (json_hint, -replacement_count, len(decoded), -abs(len(raw_bytes) - len(decoded)))
        else:
            normalized = _normalize_text_for_csv(decoded)
            lines = [ln for ln in normalized.split("\n") if str(ln).strip()]
            first = lines[0] if lines else ""
            delim_counts = [first.count(d) for d in [",", ";", "\t", "|"]]
            max_delim = max(delim_counts) if delim_counts else 0
            structured = sum(1 for ln in lines[:500] if max([ln.count(d) for d in [",", ";", "\t", "|"]]) >= 1)
            score = (structured, len(lines), max_delim, -replacement_count)

        candidates.append((score, decoded, enc))

    if not candidates:
        return raw_bytes.decode("utf-8", errors="replace")

    candidates.sort(key=lambda item: item[0], reverse=True)
    best_score, best_decoded, best_enc = candidates[0]
    logger.info(f"Decodificación seleccionada: {best_enc} score={best_score} bytes={len(raw_bytes)}")
    return best_decoded

def _build_no_data_diagnostic(content: str, delimiter: Optional[str], has_header: Optional[bool]) -> str:
    normalized = _normalize_text_for_csv(content)
    lines = [ln for ln in normalized.split("\n") if str(ln).strip()]
    first_line = lines[0][:120] if lines else ""
    return (
        f"El archivo no contiene filas de data para importar. "
        f"[diag: bytes={len(content or '')}, lineas_no_vacias={len(lines)}, "
        f"delimitador='{delimiter or '?'}', has_header={has_header}, "
        f"primera_linea='{first_line}']"
    )

def _fallback_parse_csv_rows(
    content: str,
    has_header: bool,
    forced_data_start_row: Optional[int],
    preferred_delimiter: str = ","
) -> Tuple[List[Dict[str, Any]], bool, int]:
    """
    Secondary parser when DictReader fails to produce rows.
    Tries different delimiters and quote chars (including single-quote quoted TXT exports).
    Returns (rows, no_header_mode, line_offset).
    """
    delimiters: List[str] = []
    for d in [preferred_delimiter, ",", ";", "\t", "|"]:
        if d and d not in delimiters:
            delimiters.append(d)

    best_matrix: List[List[str]] = []
    best_delimiter = preferred_delimiter
    best_quote = '"'
    best_score: Tuple[int, int] = (0, 0)

    for delim in delimiters:
        for quote in ['"', "'"]:
            try:
                reader = csv.reader(
                    io.StringIO(content),
                    delimiter=delim,
                    skipinitialspace=True,
                    quotechar=quote
                )
                matrix_rows = [r for r in reader if any(str(c or "").strip() for c in r)]
                if not matrix_rows:
                    continue
                score = (len(matrix_rows), max(len(r) for r in matrix_rows))
                if score > best_score:
                    best_score = score
                    best_matrix = matrix_rows
                    best_delimiter = delim
                    best_quote = quote
            except Exception:
                continue

    if not best_matrix:
        return [], (not has_header), (1 if not has_header else 2)

    logger.info(
        f"Fallback CSV parser aplicado. delimiter='{best_delimiter}' quotechar='{best_quote}' "
        f"rows={len(best_matrix)} has_header={has_header}"
    )

    raw_rows: List[Dict[str, Any]] = []
    if has_header:
        header_row = [_clean_csv_header_name(c) for c in best_matrix[0]]
        max_cols = max(len(r) for r in best_matrix)
        normalized_headers = [
            (header_row[idx] if idx < len(header_row) and header_row[idx] else f"col_{idx + 1}")
            for idx in range(max_cols)
        ]
        data_rows = best_matrix[1:]
        line_offset = 2
        if forced_data_start_row and forced_data_start_row > 2:
            skip_count = forced_data_start_row - 2
            if skip_count < len(data_rows):
                data_rows = data_rows[skip_count:]
                line_offset = forced_data_start_row
            else:
                logger.warning(
                    f"Fallback parser: data_start_row={forced_data_start_row} fuera de rango. "
                    "Usando inicio por defecto (línea 2)."
                )

        for r in data_rows:
            padded = list(r) + [""] * (max_cols - len(r))
            raw_rows.append(dict(zip(normalized_headers, padded)))
        return raw_rows, False, line_offset

    matrix_rows = best_matrix
    line_offset = 1
    if forced_data_start_row and forced_data_start_row > 1:
        skip_count = forced_data_start_row - 1
        if skip_count < len(matrix_rows):
            matrix_rows = matrix_rows[skip_count:]
            line_offset = forced_data_start_row
        else:
            logger.warning(
                f"Fallback parser (no header): data_start_row={forced_data_start_row} fuera de rango. "
                "Usando inicio por defecto (línea 1)."
            )
    if matrix_rows:
        max_cols = max(len(r) for r in matrix_rows)
        synthetic_headers = [f"col_{idx}" for idx in range(1, max_cols + 1)]
        for r in matrix_rows:
            padded = list(r) + [""] * (max_cols - len(r))
            raw_rows.append(dict(zip(synthetic_headers, padded)))
    return raw_rows, True, line_offset

def _emergency_parse_delimited_rows(
    content: str,
    has_header: bool,
    forced_data_start_row: Optional[int]
) -> Tuple[List[Dict[str, Any]], bool, int]:
    """
    Last-resort parser based on plain line splitting.
    Useful when CSV parser cannot tokenize but raw preview clearly shows delimited lines.
    """
    text = _normalize_text_for_csv(content)
    lines = [ln.strip() for ln in text.split("\n") if str(ln).strip()]
    if not lines:
        return [], (not has_header), (1 if not has_header else 2)

    delimiter_candidates = [",", ";", "\t", "|"]
    first_line = lines[0]
    delimiter = max(delimiter_candidates, key=lambda d: first_line.count(d))
    if first_line.count(delimiter) <= 0:
        return [], (not has_header), (1 if not has_header else 2)

    matrix: List[List[str]] = []
    for ln in lines:
        row = [str(cell).strip() for cell in ln.split(delimiter)]
        if any(cell.strip() for cell in row):
            matrix.append(row)

    if not matrix:
        return [], (not has_header), (1 if not has_header else 2)

    logger.warning(
        f"Parser de emergencia activado. delimiter='{delimiter}' rows={len(matrix)} has_header={has_header}"
    )

    if has_header:
        header_row = [_clean_csv_header_name(_clean_cell_value(c)) for c in matrix[0]]
        max_cols = max(len(r) for r in matrix)
        normalized_headers = [
            (header_row[idx] if idx < len(header_row) and header_row[idx] else f"col_{idx + 1}")
            for idx in range(max_cols)
        ]

        data_rows = matrix[1:]
        line_offset = 2
        if forced_data_start_row and forced_data_start_row > 2:
            skip_count = forced_data_start_row - 2
            if skip_count < len(data_rows):
                data_rows = data_rows[skip_count:]
                line_offset = forced_data_start_row
            else:
                logger.warning(
                    f"Emergency parser: data_start_row={forced_data_start_row} fuera de rango. "
                    "Usando inicio por defecto (línea 2)."
                )

        raw_rows: List[Dict[str, Any]] = []
        for r in data_rows:
            padded = list(r) + [""] * (max_cols - len(r))
            cleaned_row = [_clean_cell_value(cell) for cell in padded]
            raw_rows.append(dict(zip(normalized_headers, cleaned_row)))
        return raw_rows, False, line_offset

    matrix_rows = matrix
    line_offset = 1
    if forced_data_start_row and forced_data_start_row > 1:
        skip_count = forced_data_start_row - 1
        if skip_count < len(matrix_rows):
            matrix_rows = matrix_rows[skip_count:]
            line_offset = forced_data_start_row
        else:
            logger.warning(
                f"Emergency parser (no header): data_start_row={forced_data_start_row} fuera de rango. "
                "Usando inicio por defecto (línea 1)."
            )

    raw_rows: List[Dict[str, Any]] = []
    if matrix_rows:
        max_cols = max(len(r) for r in matrix_rows)
        synthetic_headers = [f"col_{idx}" for idx in range(1, max_cols + 1)]
        for r in matrix_rows:
            padded = list(r) + [""] * (max_cols - len(r))
            cleaned_row = [_clean_cell_value(cell) for cell in padded]
            raw_rows.append(dict(zip(synthetic_headers, cleaned_row)))
    return raw_rows, True, line_offset

async def get_current_mall(
    x_mall_id: Optional[str] = Header(None, alias="X-Mall-Id"),
    user_id: str = Depends(get_current_user_id)
):
    # 1. Si viene el header, usarlo (validando acceso - simplificado por ahora, RLS protege data)
    if x_mall_id:
        return x_mall_id

    # 2. Si no viene header, buscar por defecto en DB
    try:
        res = supabase.table("usuarios_malls").select("mall_id").eq("usuario_id", user_id).execute()
        malls = res.data
        if len(malls) == 1:
            return malls[0]['mall_id']
        elif len(malls) > 1:
            raise HTTPException(status_code=400, detail="Ambiguous context. Please select a mall (X-Mall-Id).")
            
        raise HTTPException(status_code=403, detail="No mall assigned to user.")
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error resolving tenant: {e}")
        raise HTTPException(status_code=500, detail="Error resolving tenant context")

# --- Schemas ---
class IngestionResponse(BaseModel):
    status: str
    message: str
    records_processed: int
    batch_id: str

class SaleReportSchema(BaseModel):
    local_id: str
    local_nombre: str
    total_bruto: float
    total_impuestos: float
    total_neto: float
    mall_nombre: str

class StoreSchema(BaseModel):
    id: str
    mall_id: str
    codigo_interno: str
    nombre: str
    email: Optional[str] = None
    email_secundario: Optional[str] = None
    rubro: Optional[str] = None
    created_at: str
    responsable: str
    contrato_no: str
    piso: str
    tipo_negocio: str
    mts: str
    porciento_renta: str
    upsert_activo: bool = False
    mall_nombre: Optional[str] = "Mall Plaza"
    fecha_corte_importacion: Optional[str] = None

STORE_WRITE_FIELDS = {
    "mall_id", "codigo_interno", "nombre", "email", "email_secundario", "rubro", "responsable",
    "contrato_no", "piso", "tipo_negocio", "mts", "porciento_renta",
    "upsert_activo", "renta_fija", "breakpoint_venta", "porcentaje_variable",
    "fecha_corte_importacion",
}

STORE_NUMERIC_FIELDS = {
    "mts", "porciento_renta", "renta_fija", "breakpoint_venta", "porcentaje_variable",
}


def _sanitize_store_write_payload(payload: Dict[str, Any], *, existing_mall_id: Optional[str] = None) -> Dict[str, Any]:
    data = {
        key: value
        for key, value in (payload or {}).items()
        if key in STORE_WRITE_FIELDS
    }
    if existing_mall_id:
        data["mall_id"] = existing_mall_id
    if "nombre" in data:
        data["nombre"] = str(data.get("nombre") or "").strip()
    if "codigo_interno" in data:
        data["codigo_interno"] = str(data.get("codigo_interno") or "").strip()
    if data.get("email") == "":
        data["email"] = None
    if data.get("email_secundario") == "":
        data["email_secundario"] = None
    for field in STORE_NUMERIC_FIELDS:
        if field not in data:
            continue
        raw_value = data.get(field)
        if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
            data[field] = None
            continue
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"{field} debe contener un número válido")
        if numeric_value != numeric_value or numeric_value in (float("inf"), float("-inf")):
            raise HTTPException(status_code=400, detail=f"{field} debe contener un número válido")
        data[field] = numeric_value
    if "fecha_corte_importacion" in data:
        raw_cutoff = str(data.get("fecha_corte_importacion") or "").strip()
        if raw_cutoff:
            try:
                datetime.strptime(raw_cutoff, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="fecha_corte_importacion debe tener formato YYYY-MM-DD")
            data["fecha_corte_importacion"] = raw_cutoff
        else:
            data["fecha_corte_importacion"] = None
    return data


def _normalize_store_catalog_key(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip()).lower()
    return "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )


def _validate_store_catalog_values(mall_id: str, data: Dict[str, Any]) -> None:
    fields = [field for field in ("tipo_negocio", "rubro") if field in data]
    if not fields:
        return

    requested = {
        field: _normalize_store_catalog_key(data.get(field))
        for field in fields
    }
    missing_value = next((field for field, key in requested.items() if not key), None)
    if missing_value:
        raise HTTPException(status_code=400, detail=f"{missing_value} debe seleccionarse desde Catálogos Locales")

    try:
        response = (
            supabase.table("store_field_options")
            .select("field_name,value")
            .eq("mall_id", mall_id)
            .in_("field_name", fields)
            .execute()
        )
    except Exception as exc:
        logger.error("Error validando catalogos de locales mall=%s: %s", mall_id, exc)
        raise HTTPException(status_code=400, detail="Catálogos Locales no está disponible para validar el local")

    allowed: Dict[str, Set[str]] = {field: set() for field in fields}
    for row in response.data or []:
        field_name = str(row.get("field_name") or "")
        if field_name in allowed:
            allowed[field_name].add(_normalize_store_catalog_key(row.get("value")))

    invalid = [
        field for field, key in requested.items()
        if key not in allowed.get(field, set())
    ]
    if invalid:
        labels = {
            "tipo_negocio": "Tipo de Negocio",
            "rubro": "Rubro General",
        }
        raise HTTPException(
            status_code=400,
            detail=f"{labels[invalid[0]]} debe existir en Catálogos Locales antes de guardar el local",
        )


class CustomFieldOptionPayload(BaseModel):
    id: Optional[str] = None
    field_definition_id: Optional[str] = None
    label: str
    value: str
    sort_order: int = 0
    active: bool = True
    parent_option_id: Optional[str] = None


class CustomFieldDefinitionCreateRequest(BaseModel):
    mall_id: str
    key: str
    label: str
    data_type: str
    widget_type: str
    required: bool = False
    active: bool = True
    sort_order: int = 0
    parent_field_id: Optional[str] = None
    options: List[CustomFieldOptionPayload] = []


class CustomFieldDefinitionUpdateRequest(BaseModel):
    key: Optional[str] = None
    label: Optional[str] = None
    data_type: Optional[str] = None
    widget_type: Optional[str] = None
    required: Optional[bool] = None
    active: Optional[bool] = None
    sort_order: Optional[int] = None
    parent_field_id: Optional[str] = None
    options: Optional[List[CustomFieldOptionPayload]] = None


class LocalCustomFieldValuePayload(BaseModel):
    field_definition_id: str
    value_text: Optional[str] = None
    value_number: Optional[float] = None
    value_date: Optional[str] = None
    selected_option_id: Optional[str] = None


class LocalCustomFieldValueUpsertRequest(BaseModel):
    values: List[LocalCustomFieldValuePayload]

class ResendTestMessageRequest(BaseModel):
    to: str
    subject: Optional[str] = None
    message: Optional[str] = None

class ResendSenderUpdateRequest(BaseModel):
    from_email: str
    from_name: str

class MissingDaysEmailPreviewRequest(BaseModel):
    mall_name: str
    local_name: str
    fecha_inicio: str
    fecha_fin: str
    missing_details: List[Dict[str, Any]]
    report_url: Optional[str] = None

class MissingDaysEmailSettingsRequest(BaseModel):
    mall_id: str
    notification_type: str = "missing_days_audit"
    enabled: bool = False
    weekdays: List[int] = []
    send_time: str = "08:00"
    lookback_days: int = 7
    send_only_with_gaps: bool = True
    cc_emails: List[str] = []
    subject_template: Optional[str] = None
    body_template: Optional[str] = None

class MissingDaysSendNowRequest(BaseModel):
    mall_id: str
    notification_type: str = "missing_days_audit"

class CopilotSettingsRequest(BaseModel):
    enabled: bool = False
    provider: str = "openai"
    model: Optional[str] = None
    api_key: Optional[str] = None
    clear_api_key: bool = False

class CopilotChatMessage(BaseModel):
    role: str
    content: str

class CopilotChatRequest(BaseModel):
    mall_id: str
    message: str
    history: List[CopilotChatMessage] = []

class CopilotEmailSendRequest(BaseModel):
    mall_id: str
    draft_id: str

class RemoteRequest(BaseModel):
    local_id: Optional[str] = None
    protocolo: str = "SFTP"
    host: str
    puerto: int = 22
    usuario: str = ""
    password: Optional[str] = None
    ruta: str = "/"
    tipo_archivo: str = "CSV"
    has_header: Optional[bool] = None
    data_start_row: Optional[int] = None
    provider: Optional[str] = None

class RemoteConnectionCreateRequest(BaseModel):
    mall_id: str
    nombre: str
    protocolo: str = "SFTP"
    host: str
    puerto: int = 22
    usuario: str
    password: str
    ruta_base: Optional[str] = None

class RemoteConnectionUpdateRequest(BaseModel):
    nombre: Optional[str] = None
    protocolo: Optional[str] = None
    host: Optional[str] = None
    puerto: Optional[int] = None
    usuario: Optional[str] = None
    password: Optional[str] = None
    ruta_base: Optional[str] = None

class RemoteConnectionResponse(BaseModel):
    id: str
    mall_id: str
    nombre: str
    protocolo: str
    host: str
    puerto: int
    usuario: str
    password: str = ""
    password_masked: Optional[str] = ""
    has_password: bool = False
    ruta_base: Optional[str] = None
    created_at: Optional[str] = None

class ConnectionRunSchema(BaseModel):
    id: str
    mall_id: str
    local_id: Optional[str] = None
    connection_id: Optional[str] = None
    run_type: str
    status: str
    error_code: Optional[str] = None
    error_message: Optional[str] = None
    started_at: str
    finished_at: str
    duration_ms: int = 0
    created_by: Optional[str] = None
    created_at: Optional[str] = None

class ConnectionStatusResponse(BaseModel):
    mall_id: str
    summary: Dict[str, int]
    recent_runs: List[Dict[str, Any]]
    connections: List[Dict[str, Any]]

class ConnectionFailuresResponse(BaseModel):
    mall_id: str
    date: str
    count: int
    failures: List[Dict[str, Any]]

class ConnectionRetryResponse(BaseModel):
    status: str
    connection_id: str
    mall_id: Optional[str] = None
    run: Dict[str, Any]
    retry_attempt: Optional[Dict[str, Any]] = None
    policy: Dict[str, Any]

class ImportConfigSchema(BaseModel):
    id: Optional[str] = None
    nombre: Optional[str] = None
    protocolo: str = "SFTP"
    host: Optional[str] = None
    puerto: Optional[int] = None
    usuario: Optional[str] = None
    password: Optional[str] = None
    ruta_remota: Optional[str] = None
    tipo_archivo: Optional[str] = "CSV"
    mapping: Dict[str, str] = {}
    constants: Dict[str, str] = {}  # Added for constant field values
    date_format: Optional[str] = "auto"  # Date format preference for fecha_venta
    has_header: Optional[bool] = None
    data_start_row: Optional[int] = None
    # Worker names fallback support is in normalization logic

class ExecuteManualRequest(BaseModel):
    config_id: str
    filename: str
    config: Optional[ImportConfigSchema] = None
    request_id: Optional[str] = None

class LoadLogSchema(BaseModel):
    id: Optional[str] = None
    fecha_hora: datetime
    mall_id: Optional[str] = None
    local_id: Optional[str] = None
    local_nombre: str
    archivo: str
    estado: str # 'exito', 'error', 'no_encontrado'
    mensaje: str
    batch_id: Optional[str] = None

class UserSchema(BaseModel):
    id: str
    nombre: str
    email: str
    rol: str
    estado: str
    ultimo_acceso: Optional[str] = None
    created_at: str

# --- Dependencias de Seguridad ---
async def verify_api_key(x_api_key: str = Header(..., alias="X-API-Key")):
    # En un entorno real, aquí consultaríamos en la DB si la API Key existe y está activa
    valid_keys = ["demo-key-123", "mall-plaza-admin-key", "costanera-center-key"]
    if x_api_key not in valid_keys:
        logger.warning(f"Intento de acceso fallido con API Key: {x_api_key}")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="API Key inválida o no autorizada"
        )
    return x_api_key

def insert_load_log(
    local_nombre: str,
    archivo: str,
    estado: str,
    mensaje: str,
    batch_id: Optional[str] = None,
    detalles: Optional[List[Dict]] = None,
    mall_id: Optional[str] = None,
    local_id: Optional[str] = None,
    mall_nombre: Optional[str] = None,
    canal: Optional[str] = None,
    records_processed: Optional[int] = None,
    error_count: Optional[int] = None,
    metadata: Optional[Dict[str, Any]] = None,
):
    """Inserts a log into Supabase 'logs_carga' table."""
    if not supabase:
        logger.warning(f"Supabase not configured. Skipping log: {mensaje}")
        return
    
    try:
        log_data = build_load_log_payload(
            local_nombre=local_nombre,
            archivo=archivo,
            estado=estado,
            mensaje=mensaje,
            batch_id=batch_id,
            detalles=detalles,
            mall_id=mall_id,
            mall_nombre=mall_nombre,
            local_id=local_id,
            canal=canal,
            records_processed=records_processed,
            error_count=error_count,
            metadata=metadata,
        )
        logger.info(f"Intentando guardar log en Supabase: {local_nombre} - {archivo} - {estado}")
        insert_load_log_row(supabase, log_data, logger=logger)
        logger.info("Log guardado exitosamente.")
    except Exception as e:
        logger.error(f"Error CRÍTICO insertando log en Supabase: {e}")
        logger.error(f"Data intentada: {log_data}")

def _split_transform_fields(raw_fields: Any) -> List[str]:
    if isinstance(raw_fields, list):
        return [str(field).strip() for field in raw_fields if str(field or "").strip()]
    return [part.strip() for part in str(raw_fields or "").split(",") if part.strip()]


def _clean_generated_invoice_piece(value: Any) -> str:
    text = str(value or "").strip().strip("'\"")
    text = re.sub(r"\s+", "", text)
    text = text.replace("/", "").replace("\\", "").replace("-", "")
    return text


def _format_generated_invoice(local_code: Any, sale_date: Any, sequence: int) -> str:
    local_part = _clean_generated_invoice_piece(local_code)
    date_part = _clean_generated_invoice_piece(str(sale_date or "").replace("-", ""))
    return f"{local_part}{date_part}{sequence:04d}"


def _parse_mapped_decimal(value: Any, decimal_separator: Any = ".") -> float:
    if value is None:
        return 0.0

    text = str(value).strip().strip("'\"")
    if not text:
        return 0.0

    text = re.sub(r"\s+", "", text)
    text = text.replace("$", "").replace("RD", "").replace("rd", "")
    if decimal_separator == ",":
        text = text.replace(".", "")
        if text.count(",") > 1:
            sign = "-" if text.startswith("-") else ""
            unsigned = text[1:] if sign else text
            parts = unsigned.split(",")
            if len(parts) >= 3 and all(len(part) == 3 for part in parts[-2:]):
                digits = "".join(parts)
                if len(digits) > 6:
                    text = f"{sign}{digits[:-6]}.{digits[-6:]}"
                else:
                    text = f"{sign}0.{digits.zfill(6)}"
            else:
                text = "".join(parts[:-1]) + "." + parts[-1]
                if sign and not text.startswith("-"):
                    text = f"-{text}"
        else:
            text = text.replace(",", ".")
    else:
        text = text.replace(",", "")

    try:
        return float(text)
    except Exception:
        return 0.0


def _parse_import_cutoff_date(raw: Any) -> Optional[str]:
    text = str(raw or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _is_import_date_closed(sale_date: str, cutoff_date: Optional[str]) -> bool:
    return bool(sale_date and cutoff_date and sale_date <= cutoff_date)


def process_file_content(content: str, filename: str, config: Dict[str, Any], batch_id: str, mall_id: str = None):
    """
    Parses content based on config mapping and inserts into Supabase 'ventas' table.
    Returns (success_count, errors_list).
    """
    mapping = config.get("mapping", {})
    constants = config.get("constants", {})
    decimal_separator = constants.get("_decimal_separator", ".")
    import_cutoff_date = _parse_import_cutoff_date(config.get("fecha_corte_importacion"))
    tipo_archivo = config.get("tipo_archivo", "CSV").upper()
    local_nombre = config.get("nombre", "Desconocido")
    effective_mall_id = mall_id or config.get("mall_id")
    normalized_content = _normalize_text_for_csv(content) if tipo_archivo != "JSON" else content
    
    records_to_insert = []
    errors = []
    
    # Pre-warm Supabase connection / cache (optional)
    try:
        supabase.table("ventas").select("count", count="exact").limit(0).execute()
    except:
        pass
    
    try:
        raw_rows = []
        forced_has_header, forced_data_start_row = _extract_parsing_options(config)
        no_header_mode = False
        line_offset = 2
        detected_delimiter: Optional[str] = None
        detected_has_header: Optional[bool] = None
        default_no_header_map = {
            "factura_numero": "col_1",
            "local_codigo": "col_2",
            "fecha_venta": "col_3",
            "total_bruto": "col_8",
            "total_impuestos": "col_9",
            "total_neto": "col_10",
        }
        if tipo_archivo == "JSON":
            try:
                data = json.loads(content)
                if isinstance(data, list):
                    raw_rows = data
                elif isinstance(data, dict):
                    # Try to find list inside
                    for k, v in data.items():
                        if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                            raw_rows = v
                            break
                    if not raw_rows:
                        raw_rows = [data]
                
                # Apply flattening using Pandas
                if raw_rows:
                    df = pd.json_normalize(raw_rows)
                    # Convert NaN to None for SQL safety
                    raw_rows = df.where(pd.notnull(df), None).to_dict(orient='records')
            except Exception as e:
                return 0, [{"linea": 0, "error": f"JSON inválido: {str(e)}"}]
        else:
            # Default CSV/TXT
            f = io.StringIO(normalized_content)
            sample = normalized_content[:4096]
            delimiter, has_header = _detect_delimiter_and_header(sample)
            detected_delimiter = delimiter
            detected_has_header = has_header
            parsed_rows_before_offset = 0
            if forced_has_header is not None:
                has_header = forced_has_header
                detected_has_header = has_header
            if has_header and _should_treat_as_no_header(normalized_content, delimiter):
                has_header = False
                detected_has_header = has_header
            f.seek(0)
            if has_header:
                reader = csv.DictReader(f, delimiter=delimiter, skipinitialspace=True)
                raw_rows = [_normalize_csv_row_keys(r) for r in reader]
                parsed_rows_before_offset = len(raw_rows)
                if forced_data_start_row and forced_data_start_row > 2:
                    skip_count = forced_data_start_row - 2
                    if skip_count < len(raw_rows):
                        raw_rows = raw_rows[skip_count:]
                        line_offset = forced_data_start_row
                    else:
                        logger.warning(
                            f"data_start_row={forced_data_start_row} fuera de rango para archivo '{filename}'. "
                            "Usando inicio por defecto (línea 2)."
                        )
            else:
                no_header_mode = True
                line_offset = 1
                reader = csv.reader(f, delimiter=delimiter, skipinitialspace=True)
                matrix_rows = [r for r in reader if any(str(c or "").strip() for c in r)]
                parsed_rows_before_offset = len(matrix_rows)
                if forced_data_start_row and forced_data_start_row > 1:
                    skip_count = forced_data_start_row - 1
                    if skip_count < len(matrix_rows):
                        matrix_rows = matrix_rows[skip_count:]
                        line_offset = forced_data_start_row
                    else:
                        logger.warning(
                            f"data_start_row={forced_data_start_row} fuera de rango para archivo '{filename}' (no header). "
                            "Usando inicio por defecto (línea 1)."
                        )
                if matrix_rows:
                    max_cols = max(len(r) for r in matrix_rows)
                    synthetic_headers = [f"col_{idx}" for idx in range(1, max_cols + 1)]
                    for r in matrix_rows:
                        padded = list(r) + [""] * (max_cols - len(r))
                        raw_rows.append(dict(zip(synthetic_headers, padded)))

        if tipo_archivo != "JSON" and not raw_rows:
            fallback_rows, fallback_no_header, fallback_line_offset = _fallback_parse_csv_rows(
                content=normalized_content,
                has_header=has_header,
                forced_data_start_row=forced_data_start_row,
                preferred_delimiter=delimiter
            )
            if fallback_rows:
                raw_rows = fallback_rows
                no_header_mode = fallback_no_header
                line_offset = fallback_line_offset

        if tipo_archivo != "JSON" and not raw_rows:
            emergency_rows, emergency_no_header, emergency_line_offset = _emergency_parse_delimited_rows(
                content=normalized_content,
                has_header=has_header,
                forced_data_start_row=forced_data_start_row
            )
            if emergency_rows:
                raw_rows = emergency_rows
                no_header_mode = emergency_no_header
                line_offset = emergency_line_offset

        if not raw_rows:
            return 0, [{
                "linea": 0,
                "error": _build_no_data_diagnostic(
                    normalized_content if tipo_archivo != "JSON" else content,
                    detected_delimiter,
                    detected_has_header
                )
            }]

        effective_mapping = dict(mapping or {})
        if no_header_mode:
            for field, col in default_no_header_map.items():
                if not effective_mapping.get(field):
                    effective_mapping[field] = col

        # Validar mapeo básico
        req_sys_fields = ['factura_numero', 'fecha_venta', 'local_codigo', 'total_bruto']
        missing_mapping = []
        for field in req_sys_fields:
            has_mapping = field in effective_mapping and effective_mapping[field]
            has_constant = field in constants and constants[field]
            has_transform = constants.get(f"_{field}_mode") in ("generated_sequence", "concat")
            if not (has_mapping or has_constant or has_transform):
                missing_mapping.append(field)
        
        if missing_mapping:
            logger.error(f"Mapeo incompleto. Faltan: {', '.join(missing_mapping)}")
            return 0, [{"linea": 0, "error": f"Mapeo incompleto. Faltan: {', '.join(missing_mapping)}"}]

        for i, row in enumerate(raw_rows):
            try:
                line_no = i + line_offset
                record = {}
                normalized_row = _normalize_csv_row_keys(row) if isinstance(row, dict) else {}
                lowered_row = {k.lower(): v for k, v in normalized_row.items()}
                # 1. Apply Mapping
                for sys_field, header in effective_mapping.items():
                    mapped_value = None
                    header_key = _clean_csv_header_name(header) if header is not None else ""

                    if header_key in normalized_row:
                        mapped_value = normalized_row[header_key]
                    elif header_key.lower() in lowered_row:
                        mapped_value = lowered_row[header_key.lower()]
                    elif no_header_mode:
                        fallback_col = None
                        if header_key.isdigit():
                            numeric_col = f"col_{int(header_key)}"
                            if numeric_col in normalized_row:
                                fallback_col = numeric_col
                        elif header_key.lower().startswith("col_"):
                            explicit_col = header_key.lower()
                            if explicit_col in normalized_row:
                                fallback_col = explicit_col

                        if not fallback_col:
                            fallback_col = default_no_header_map.get(sys_field)

                        if fallback_col and fallback_col in normalized_row:
                            mapped_value = normalized_row[fallback_col]
                        elif fallback_col and fallback_col.lower() in lowered_row:
                            mapped_value = lowered_row[fallback_col.lower()]

                    if mapped_value is not None:
                        record[sys_field] = _clean_cell_value(mapped_value)
                
                # 2. Apply Constants (exclude meta-constants that are not DB columns)
                for k, v in constants.items():
                    # Internal parsing/config flags (prefixed with "_") are not DB columns.
                    if not str(k).startswith("_"):
                        record[k] = v
                
                if i == 0:
                    logger.info(f"Muestra mapeo primer registro: {record}")
                
                # 3. Validation & Type Casting (Shared for both formats)
                if not record.get('fecha_venta'):
                     errors.append({"linea": line_no, "error": "Falta fecha_venta"})
                     continue

                local_code = str(record.get('local_codigo') or "").strip().strip("'\"")
                if not local_code:
                    errors.append({
                        "linea": line_no,
                        "error": "Falta local_codigo. No se puede cargar una venta sin un código de local válido."
                    })
                    continue
                record['local_codigo'] = local_code
                record['_source_line'] = line_no

                # Normalize date with shared format support used by manual and automatic imports.
                raw_date = str(record['fecha_venta']).strip().strip("'\"")
                explicit_format = constants.get('_date_format', 'auto')
                normalized_date = normalize_sale_date(raw_date, explicit_format)

                if normalized_date:
                    record['fecha_venta'] = normalized_date
                else:
                    errors.append({"linea": line_no, "error": f"Formato de fecha inválido: {raw_date}"})
                    continue

                if _is_import_date_closed(record["fecha_venta"], import_cutoff_date):
                    errors.append({
                        "linea": line_no,
                        "error": f"Fecha {record['fecha_venta']} pertenece a un periodo cerrado (cierre hasta {import_cutoff_date})."
                    })
                    continue

                def resolve_transform_value(part: str) -> str:
                    clean_part = str(part or "").strip()
                    if clean_part in ("numero_registro", "linea", "_line_number"):
                        return f"{i + 1:04d}"
                    if clean_part == "fecha_venta" and record.get("fecha_venta"):
                        return str(record.get("fecha_venta")).replace("-", "")
                    if clean_part in record:
                        return _clean_cell_value(record.get(clean_part))

                    header_key = _clean_csv_header_name(clean_part)
                    if header_key in normalized_row:
                        return _clean_cell_value(normalized_row[header_key])
                    if header_key.lower() in lowered_row:
                        return _clean_cell_value(lowered_row[header_key.lower()])
                    return ""

                for sys_field in list(req_sys_fields) + ["total_impuestos", "total_neto", "comprobante", "hora_transaccion"]:
                    transform_mode = constants.get(f"_{sys_field}_mode")
                    if transform_mode == "generated_sequence" and sys_field == "factura_numero":
                        record["factura_numero"] = _format_generated_invoice(
                            record.get("local_codigo"),
                            record.get("fecha_venta"),
                            i + 1
                        )
                    elif transform_mode == "concat":
                        transform_fields = _split_transform_fields(constants.get(f"_{sys_field}_concat_fields"))
                        separator = str(constants.get(f"_{sys_field}_concat_separator", "-"))
                        values = [
                            _clean_generated_invoice_piece(resolve_transform_value(part))
                            for part in transform_fields
                        ]
                        values = [value for value in values if value]
                        if values:
                            record[sys_field] = separator.join(values)
                
                # Ensure numeric types
                for num_field in ['total_bruto', 'total_impuestos', 'total_neto']:
                    record[num_field] = _parse_mapped_decimal(
                        record.get(num_field, 0.0),
                        decimal_separator
                    )
                
                # Validation: Reject if Total/Net is 0 but Tax > 0
                if record['total_bruto'] == 0:
                    if record['total_impuestos'] > 0:
                        errors.append({"linea": line_no, "error": f"Total Bruto es 0.00 pero tiene impuestos ({record['total_impuestos']}). Verifique el archivo."})
                        continue
                    if record['total_neto'] > 0:
                        errors.append({"linea": line_no, "error": f"Total Bruto es 0.00 pero tiene Neto ({record['total_neto']}). Verifique el archivo."})
                        continue

                if record['total_neto'] == 0 and record['total_bruto'] > 0 and record['total_impuestos'] > 0:
                     # This might be valid for tax-inclusive pricing where net wasn't calculated, 
                     # but user asked for alert on zero net.
                     # Let's be strict: if tax > 0, net should ideally be > 0.
                     # However, user specifically mentioned "alert if totalbruto or totalneto is zero"
                     errors.append({"linea": line_no, "error": f"Total Neto es 0.00 pero tiene Impuestos/Total. Verifique el archivo."})
                     continue
                
                if record.get('total_bruto', 0) == 0:
                     # Allow 0 total only if everything else is 0 (cancel/void), or logic elsewhere handles it,
                     # but broadly warning is good.
                     pass

                records_to_insert.append(record)
            except Exception as row_e:
                errors.append({"linea": line_no, "error": str(row_e)})

    except Exception as e:
        logger.error(f"Error procesando contenido: {e}")
        return 0, [{"linea": 0, "error": str(e)}]

    try:
        # --- DB SCHEMA MAPPING & RESOLUTION ---
        final_records = []
        
        # 1. Resolve Local UUIDs cache
        local_codigos = {
            str(r.get('local_codigo')).strip().upper()
            for r in records_to_insert
            if r.get('local_codigo')
        }
        local_map = {} # codigo -> {id, mall_id}
        
        if local_codigos and supabase:
            try:
                # Query locales table to find UUIDs for these codes
                query = (
                    supabase.table("locales")
                    .select("id, codigo_interno, mall_id")
                    .in_("codigo_interno", list(local_codigos))
                )
                if effective_mall_id:
                    query = query.eq("mall_id", effective_mall_id)
                res = query.execute()
                for loc in res.data:
                    code_key = str(loc.get('codigo_interno') or '').strip().upper()
                    if code_key:
                        local_map[code_key] = {'id': loc['id'], 'mall_id': loc.get('mall_id')}
            except Exception as e:
                logger.warning(f"Error resolviendo local_ids: {e}")

        # 2. Transform Keys
        for r in records_to_insert:
            new_r = r.copy()
            source_line = int(new_r.pop('_source_line', 0) or 0)
            
            # Map System Fields -> DB Columns
            if 'factura_numero' in new_r:
                new_r['factura_no'] = new_r.pop('factura_numero')
            if 'fecha_venta' in new_r:
                new_r['fecha'] = new_r.pop('fecha_venta')
            if 'factura_no' in new_r and new_r['factura_no'] is not None:
                factura_limpia = str(new_r['factura_no']).strip().strip("'\"")
                new_r['factura_no'] = factura_limpia if factura_limpia else None
            
            # Normalizar campos de hora (hora, hora_transaccion)
            for time_col in ['hora', 'hora_transaccion']:
                if time_col in new_r and new_r[time_col]:
                    val = str(new_r[time_col]).strip().strip("'\"")
                    if val.isdigit():
                        if int(val) < 24:
                            new_r[time_col] = f"{int(val):02d}:00:00"
                        elif len(val) in [5, 6]:
                            # HHMMSS -> HH:MM:SS
                            vh = val.zfill(6)
                            hh, mm, ss = int(vh[0:2]), int(vh[2:4]), int(vh[4:6])
                            # Validation: Clamp to valid ranges if needed (user data often has 60s or bad clocks)
                            mm = min(mm, 59)
                            ss = min(ss, 59)
                            new_r[time_col] = f"{hh:02d}:{mm:02d}:{ss:02d}"
                        elif len(val) > 6:
                            # 1118234 -> 11:18:23 / 7 digits
                            # Take first 6 as HHMMSS
                            vh = val.zfill(6)
                            hh, mm, ss = int(vh[0:2]), int(vh[2:4]), int(vh[4:6])
                            mm = min(mm, 59)
                            ss = min(ss, 59)
                            new_r[time_col] = f"{hh:02d}:{mm:02d}:{ss:02d}"
                    elif 'AM' in val.upper() or 'PM' in val.upper():
                        # Example: "10:02 AM" must not become "10:02 AM:00".
                        # Parse first, then truncate to hour as requested (HH:00:00).
                        ampm_val = re.sub(r'\s+', ' ', val.replace('.', '')).strip()
                        ampm_val = re.sub(r'(?i)\s*([AP]M)$', r' \1', ampm_val).strip()
                        parsed_dt = None
                        for fmt in ("%I:%M:%S %p", "%I:%M %p", "%I %p"):
                            try:
                                parsed_dt = datetime.strptime(ampm_val, fmt)
                                break
                            except Exception:
                                continue
                        if parsed_dt:
                            new_r[time_col] = f"{parsed_dt.hour:02d}:00:00"
                    elif val.count(':') == 1:
                        hh_mm = val.split(':', 1)
                        try:
                            hh = int(hh_mm[0])
                            mm = int(hh_mm[1])
                            if 0 <= hh <= 23:
                                mm = min(max(mm, 0), 59)
                                new_r[time_col] = f"{hh:02d}:{mm:02d}:00"
                        except Exception:
                            pass

            # Resolve Local ID & Mall ID
            l_code = str(new_r.get('local_codigo')).strip().strip("'\"").upper() if new_r.get('local_codigo') else None
            if l_code:
                if l_code in local_map:
                    local_info = local_map[l_code]
                    new_r['local_id'] = local_info['id']
                    if local_info.get('mall_id'):
                        new_r['mall_id'] = local_info['mall_id']
                    else:
                        # Fallback to context mall_id if local doesn't have one
                        if effective_mall_id:
                            new_r['mall_id'] = effective_mall_id
                            logger.info(f"Using context mall_id: {effective_mall_id} for local {l_code}")
                    del new_r['local_codigo'] # Remove text code, keep UUID
                else:
                    msg = (
                        f"Código local '{l_code}' no encontrado"
                        f"{f' en el mall {effective_mall_id}' if effective_mall_id else ''}."
                    )
                    errors.append({"linea": source_line, "error": msg})
                    logger.warning(msg)
                    continue
            else:
                msg = "Falta local_codigo. No se puede cargar una venta sin un código de local válido."
                errors.append({"linea": source_line, "error": msg})
                logger.warning(msg)
                continue
            
            final_records.append(new_r)
        
        records_to_insert = final_records

        def _dedupe_key(row: Dict[str, Any]) -> Optional[str]:
            local_id = row.get("local_id")
            fecha = row.get("fecha")
            factura = str(row.get("factura_no") or "").strip()
            if not local_id or not fecha or not factura:
                return None
            return f"{local_id}|{fecha}|{factura}"

        def _upsert_chunk_without_unique_constraint(chunk_rows: List[Dict[str, Any]]) -> None:
            keyed_rows: Dict[str, Dict[str, Any]] = {}
            no_key_rows: List[Dict[str, Any]] = []
            for row in chunk_rows:
                key = _dedupe_key(row)
                if key:
                    keyed_rows[key] = row
                else:
                    no_key_rows.append(row)

            existing_map: Dict[str, str] = {}
            if keyed_rows:
                local_ids = list({str(r.get("local_id")) for r in keyed_rows.values() if r.get("local_id")})
                fechas = list({str(r.get("fecha")) for r in keyed_rows.values() if r.get("fecha")})
                if local_ids and fechas:
                    existing_rows = (
                        supabase.table("ventas")
                        .select("id, local_id, fecha, factura_no")
                        .in_("local_id", local_ids)
                        .in_("fecha", fechas)
                        .execute()
                    ).data or []

                    for ex in existing_rows:
                        ex_key = f"{ex.get('local_id')}|{ex.get('fecha')}|{str(ex.get('factura_no') or '').strip()}"
                        if ex.get("id"):
                            existing_map[ex_key] = ex["id"]

            updates: List[Tuple[str, Dict[str, Any]]] = []
            inserts: List[Dict[str, Any]] = []
            for key, row in keyed_rows.items():
                existing_id = existing_map.get(key)
                if existing_id:
                    updates.append((existing_id, row))
                else:
                    inserts.append(row)
            inserts.extend(no_key_rows)

            for existing_id, row in updates:
                payload = {k: v for k, v in row.items() if k != "id"}
                supabase.table("ventas").update(payload).eq("id", existing_id).execute()

            if inserts:
                supabase.table("ventas").insert(inserts).execute()

        # Insertion into Supabase
        if records_to_insert and supabase:
            # Remove duplicates within the same import batch (last row wins) before writing to DB.
            deduped_records: Dict[str, Dict[str, Any]] = {}
            no_key_records: List[Dict[str, Any]] = []
            for row in records_to_insert:
                row_key = _dedupe_key(row)
                if row_key:
                    deduped_records[row_key] = row
                else:
                    no_key_records.append(row)
            records_to_insert = list(deduped_records.values()) + no_key_records

            # Batch upsert in chunks of 100
            for i in range(0, len(records_to_insert), 100):
                chunk = records_to_insert[i:i+100]

                if i == 0:
                    fechas_muestra = [r.get('fecha') for r in chunk[:3]]
                    logger.info(f"Insertando/actualizando ventas con fechas: {fechas_muestra}")
                    logger.info(f"Muestra registro completo: {chunk[0]}")

                try:
                    # Preferred path: true DB-level upsert on (local_id, fecha, factura_no).
                    res = supabase.table("ventas").upsert(
                        chunk,
                        on_conflict="local_id,fecha,factura_no"
                    ).execute()
                    logger.info(f"Respuesta upsert ventas: {res}")
                except Exception as e:
                    msg = str(e).lower()
                    if "no unique" in msg or "on conflict" in msg:
                        logger.warning(
                            "No existe constraint única para upsert en ventas(local_id,fecha,factura_no). "
                            "Aplicando fallback de update/insert por aplicación."
                        )
                        _upsert_chunk_without_unique_constraint(chunk)
                    else:
                        logger.error(f"Error insertando/upsert chunk: {e}")
                        raise e
                
        inserted_count = len(records_to_insert)
        if inserted_count > 0:
            invalidated = _invalidate_dashboard_cache(mall_id)
            logger.info(
                "Dashboard BI cache invalidated after sales import (mall=%s, keys=%s)",
                mall_id,
                invalidated,
            )
        return inserted_count, errors

    except Exception as e:
        logger.error(f"Error in process_file_content: {e}")
        return 0, [{"linea": 0, "error": str(e)}]

@app.get("/")
async def root():
    return {"message": "MSMALL API is online", "docs": "/docs"}

# --- API DE CONSUMO: INGESTA DE VENTAS ---
@app.post("/api/v1/ingesta", response_model=IngestionResponse, status_code=status.HTTP_201_CREATED)
async def ingesta_ventas(
    file: UploadFile = File(...), 
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
    mall_id: str = Depends(get_current_mall)
):
    """
    Endpoint principal para que los locales envíen sus ventas diarias.
    Acepta un archivo y procesa según la configuración del local.
    
    El mall_id se detecta automáticamente del contexto del usuario autenticado.
    No es necesario enviar X-Mall-Id header - se infiere del usuario logueado.
    """
    if not SUPABASE_SERVICE_ROLE_KEY:
        logger.error("La ingesta autenticada requiere SUPABASE_SERVICE_ROLE_KEY en el backend.")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="La ingesta segura no está disponible. Contacta al administrador del sistema.",
        )

    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    batch_id = str(uuid4())
    try:
        content_bytes = await file.read()
        content = content_bytes.decode('utf-8-sig', errors='replace')
        
        # En una app real, buscaríamos la configuración del local asociado a la API Key
        # Por ahora, usamos una configuración genérica o vacía si no tenemos el link.
        # Para este MVP, asumiremos que si es ingesta directa, viene en formato estándar.
        config = {
            "nombre": "Ingesta API",
            "tipo_archivo": "CSV" if file.filename.endswith(".csv") else "TXT",
            "mapping": {
                "factura_numero": "factura_numero",
                "fecha_venta": "fecha_venta",
                "local_codigo": "local_codigo",
                "total_bruto": "total_bruto",
                "total_neto": "total_neto",
                "total_impuestos": "total_impuestos"
            }
        }
        
        count, errors = process_file_content(content, file.filename, config, batch_id, mall_id)
        
        estado = "exito" if count > 0 and not errors else "parcial" if count > 0 else "error"
        mensaje = f"Procesado: {count} registros."
        if errors: mensaje += f" Errores: {len(errors)}"
        
        insert_load_log(
            config["nombre"],
            file.filename,
            estado,
            mensaje,
            batch_id,
            errors,
            mall_id=mall_id,
            canal="API",
            records_processed=count,
            error_count=len(errors or []),
            metadata={"source": "direct_ingest_api"},
        )
        
        return {
            "status": "success" if count > 0 else "error",
            "message": mensaje,
            "records_processed": count,
            "batch_id": batch_id
        }

    except Exception as e:
        logger.error(f"Error procesando ingesta: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# --- EXPLORACIÓN DE DIRECTORIOS LOCALES ---
def _default_local_explorer_root() -> str:
    import os
    import sys

    # Optional override for local installs / containers.
    env_root = str(os.getenv("LOCAL_EXPLORER_ROOT") or "").strip()
    if env_root and os.path.isdir(env_root):
        return os.path.abspath(env_root)

    if os.name == "nt":
        system_drive = str(os.getenv("SystemDrive") or "C:").rstrip("\\/")
        candidate = f"{system_drive}\\"
        return candidate if os.path.isdir(candidate) else "C:\\"

    if sys.platform == "darwin" and os.path.isdir("/Users"):
        return "/Users"

    for candidate in ["/home", os.path.expanduser("~"), "/"]:
        if candidate and os.path.isdir(candidate):
            return os.path.abspath(candidate)
    return "/"


def _resolve_local_explorer_path(requested_path: Optional[str]) -> str:
    import os

    raw = str(requested_path or "").strip()
    if not raw:
        return _default_local_explorer_root()

    normalized = os.path.abspath(raw)
    cwd_path = os.path.abspath(os.getcwd())
    # Keep "/" navigable (clicking ".." from /Users or /home should reach filesystem root).
    default_markers = {".", "./", "/app", cwd_path}

    # When the UI opens the browser for the first time it often sends '.' which becomes /app
    # in containerized environments. Prefer a user-friendly root for LOCAL browsing.
    if raw in default_markers or normalized in default_markers:
        return _default_local_explorer_root()

    if os.path.exists(normalized):
        return normalized

    # If requested path doesn't exist, fall back to a sensible root instead of /app.
    return _default_local_explorer_root()


@app.get("/api/v1/explorar-directorio")
async def explorar_directorio(
    path: Optional[str] = Query(None, alias="ruta"),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    """
    Endpoint para listar directorios locales. 
    Permite al usuario navegar por carpetas para configurar la importación.
    """
    try:
        import os
        # Resolver ruta inicial amigable según OS (evita caer en /app por defecto).
        target_path = _resolve_local_explorer_path(path)
            
        items = []
        # Añadir opción para subir de nivel
        parent = os.path.dirname(target_path)
        if parent != target_path:
            items.append({"nombre": "..", "ruta": parent, "es_dir": True})

        for item in os.listdir(target_path):
            full_path = os.path.join(target_path, item)
            if os.path.isdir(full_path) and not item.startswith('.'):
                items.append({
                    "nombre": item,
                    "ruta": full_path,
                    "es_dir": True
                })
        
        return {
            "ruta_actual": target_path,
            "items": sorted(items, key=lambda x: x["nombre"].lower())
        }
    except Exception as e:
        logger.error(f"Error explorando directorio {path}: {str(e)}")
        # Return real error for debugging
        raise HTTPException(status_code=500, detail=f"Error remoto: {str(e)}")

# --- UTILIDADES DE CONEXIÓN REMOTA ---
from concurrent.futures import ThreadPoolExecutor

executor = ThreadPoolExecutor(max_workers=20) # Aumentado de 5 a 20 para evitar agotamiento

def _normalize_remote_host(host: str) -> str:
    """
    Normalize host values entered from UI:
    - trims whitespace
    - strips protocol prefixes
    - strips trailing path fragments
    """
    normalized = (host or "").strip()
    if normalized.startswith("sftp://"):
        normalized = normalized[len("sftp://"):]
    elif normalized.startswith("ftp://"):
        normalized = normalized[len("ftp://"):]
    if "/" in normalized:
        normalized = normalized.split("/", 1)[0]
    return normalized

def _candidate_hosts(host: str) -> List[str]:
    normalized = _normalize_remote_host(host)
    if not normalized:
        return []
    candidates = [normalized]
    if normalized.startswith("www.") and len(normalized) > 4:
        candidates.append(normalized[4:])
    return candidates

def get_sftp_client(host, port, user, password):
    last_error = None
    for candidate in _candidate_hosts(host):
        try:
            transport = paramiko.Transport((candidate, int(port)))
            transport.banner_timeout = 20
            transport.auth_timeout = 25
            transport.connect(username=user, password=password)
            transport.set_keepalive(30)
            sftp = paramiko.SFTPClient.from_transport(transport)
            return transport, sftp
        except Exception as e:
            last_error = e
            logger.warning(f"SFTP connect failed for host '{candidate}': {e}")

    if last_error:
        raise last_error
    raise ValueError("Host remoto inválido o vacío para conexión SFTP")

def get_ftp_client(host, port, user, password):
    last_error = None
    for candidate in _candidate_hosts(host):
        try:
            ftp = FTP()
            ftp.connect(candidate, int(port), timeout=25)
            ftp.login(user, password)
            ftp.set_pasv(True)
            return ftp
        except Exception as e:
            last_error = e
            logger.warning(f"FTP connect failed for host '{candidate}': {e}")
    if last_error:
        raise last_error
    raise ValueError("Host remoto inválido o vacío para conexión FTP")


def _remote_connection_error_message(protocol: str, exc: Exception, duration: float) -> str:
    raw_message = str(exc or "").strip()
    normalized = raw_message.lower()
    if str(protocol or "").strip().upper() == "SFTP":
        if "no existing session" in normalized or "error reading ssh protocol banner" in normalized:
            return (
                f"SFTP no disponible ({duration:.2f}s): el puerto responde, pero el servidor no completa "
                "la negociación SSH. Revise o reinicie el servicio SSH/SFTP y sus límites de sesiones."
            )
        if isinstance(exc, paramiko.AuthenticationException) or "authentication failed" in normalized:
            return f"Autenticación SFTP rechazada ({duration:.2f}s). Verifique usuario y contraseña."
        if isinstance(exc, (socket.timeout, TimeoutError)) or "timed out" in normalized:
            return f"Timeout SFTP ({duration:.2f}s): el servidor no respondió durante la negociación SSH."
    return f"Error ({duration:.2f}s): {raw_message or type(exc).__name__}"


def _is_webservice_protocol(value: Any) -> bool:
    return str(value or "").strip().upper() in {"API", "WEBSERVICE"}


def _test_remote_connection_sync(req: RemoteRequest):
    logger.info(f"Probando conexión remota sync a {req.host}:{req.puerto} ({req.protocolo})")
    start_time = time.time()
    try:
        protocol = str(req.protocolo or "").strip().upper()
        if protocol == "API":
            previous_day = date.today() - timedelta(days=1)
            api_config = _api_config_from_remote_request(
                req,
                previous_day.isoformat(),
                previous_day.isoformat(),
            )
            provider = api_provider_name(api_config)
            rows, _ = (
                fetch_bundaberg_sales(api_config)
                if provider == "bundaberg"
                else fetch_studio_g_sales(api_config)
            )
            duration = time.time() - start_time
            return {
                "status": "success",
                "message": (
                    f"API {('Bundaberg' if provider == 'bundaberg' else 'Studio G')} autenticada y consulta de ventas validada "
                    f"({len(rows)} registro(s), {duration:.2f}s)"
                ),
            }
        if protocol == "SFTP":
            ssh, sftp = get_sftp_client(req.host, req.puerto, req.usuario, req.password)
            sftp.close()
            ssh.close()
        elif protocol == "FTP":
            ftp = get_ftp_client(req.host, req.puerto, req.usuario, req.password)
            ftp.quit()
        else:
            raise ValueError(f"Protocolo no soportado: {req.protocolo}")
        duration = time.time() - start_time
        logger.info(f"Conexión exitosa en {duration:.2f}s")
        return {"status": "success", "message": f"Conexión exitosa ({duration:.2f}s)"}
    except Exception as e:
        duration = time.time() - start_time
        logger.error(f"Error conexión remota después de {duration:.2f}s: {e}")
        return {"status": "error", "message": _remote_connection_error_message(req.protocolo, e, duration)}


def _studio_g_config_from_remote_request(
    req: RemoteRequest,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
) -> Dict[str, Any]:
    today = date.today().isoformat()
    return {
        "id": "studio-g-preview",
        "mall_id": "00000000-0000-0000-0000-000000000000",
        "nombre": "Studio G API",
        "sftp_protocol": "API",
        "sftp_host": req.host,
        "sftp_user": req.usuario,
        "sftp_pass": req.password,
        "sftp_path": req.ruta,
        "_webservice_timeout_seconds": "20",
        "constants_config": {
            "provider": "studio_g",
            "_studio_g_fecha_inicio": fecha_inicio or today,
            "_studio_g_fecha_fin": fecha_fin or today,
        },
    }


def _api_config_from_remote_request(
    req: RemoteRequest,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
) -> Dict[str, Any]:
    provider = str(req.provider or "").strip().lower()
    if provider in {"bundaberg", "agora", "agora_bundaberg"} or (
        "sibs2.com" in req.host.lower() and "api_agora" in req.host.lower()
    ):
        today = date.today().isoformat()
        return {
            "id": "bundaberg-preview",
            "mall_id": "00000000-0000-0000-0000-000000000000",
            "nombre": "Bundaberg API",
            "sftp_protocol": "API",
            "sftp_host": req.host,
            "sftp_user": "",
            "sftp_pass": req.password,
            "sftp_path": req.ruta,
            "_webservice_timeout_seconds": "20",
            "constants_config": {
                "provider": "bundaberg",
                "_api_fecha_inicio": fecha_inicio or today,
                "_api_fecha_fin": fecha_fin or today,
            },
        }
    return _studio_g_config_from_remote_request(req, fecha_inicio, fecha_fin)


def _api_preview_rows(req: RemoteRequest) -> List[Dict[str, Any]]:
    provider_config = _api_config_from_remote_request(req)
    if api_provider_name(provider_config) != "bundaberg":
        return _studio_g_preview_rows(req)

    today = date.today()
    try:
        rows, _ = fetch_bundaberg_sales(
            _api_config_from_remote_request(req, today.isoformat(), today.isoformat())
        )
    except Exception as exc:
        logger.warning(
            "Bundaberg no pudo consultar la fecha actual para vista previa: %s",
            sanitize_sensitive_ops_error(exc),
        )
        rows = []
    if rows:
        return rows

    history_start = today - timedelta(days=STUDIO_G_PREVIEW_HISTORY_DAYS)
    rows, _ = fetch_bundaberg_sales(
        _api_config_from_remote_request(req, history_start.isoformat(), today.isoformat())
    )
    return rows


def _studio_g_preview_rows(req: RemoteRequest) -> List[Dict[str, Any]]:
    today = date.today()
    try:
        rows, _ = fetch_studio_g_sales(
            _studio_g_config_from_remote_request(req, today.isoformat(), today.isoformat())
        )
    except Exception as exc:
        logger.warning(
            "Studio G no pudo consultar la fecha actual para vista previa: %s",
            sanitize_sensitive_ops_error(exc),
        )
        rows = []
    if rows:
        return rows

    history_start = today - timedelta(days=STUDIO_G_PREVIEW_HISTORY_DAYS)
    rows, _ = fetch_studio_g_sales(
        _studio_g_config_from_remote_request(req, history_start.isoformat(), today.isoformat())
    )
    if rows:
        return rows

    return [{
        "fecha": today.isoformat(),
        "factura_no": "STUDIOG-MUESTRA",
        "comprobante": "STUDIOG-MUESTRA",
        "hora_transaccion": "12:00:00",
        "total_bruto": 0,
        "total_impuestos": 0,
        "total_neto": 0,
    }]


def _remote_request_with_saved_password(
    req: RemoteRequest,
    operator_ctx: Dict[str, Any],
) -> RemoteRequest:
    if not str(req.password or "").strip() and req.local_id:
        stored_config = _load_local_config_with_access(req.local_id, operator_ctx)
        stored_protocol = str(stored_config.get("sftp_protocol") or "").strip().upper()
        requested_protocol = str(req.protocolo or "").strip().upper()
        stored_host = str(stored_config.get("sftp_host") or "").strip().rstrip("/").lower()
        requested_host = str(req.host or "").strip().rstrip("/").lower()
        if stored_protocol == requested_protocol and stored_host == requested_host:
            stored_password = str(stored_config.get("sftp_pass") or "")
            if stored_password:
                return req.model_copy(update={"password": stored_password})
    return req


@app.post("/api/v1/remote/test")
async def test_remote_connection(
    req: RemoteRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    req = _remote_request_with_saved_password(req, operator_ctx)
    loop = asyncio.get_event_loop()
    try:
        # Timeout de 30s para no bloquear el worker de FastAPI indefinidamente
        return await asyncio.wait_for(
            loop.run_in_executor(executor, _test_remote_connection_sync, req),
            timeout=30.0
        )
    except asyncio.TimeoutError:
        logger.error(f"Timeout en test_remote_connection para {req.host}")
        raise HTTPException(status_code=504, detail="El servidor remoto no respondió a tiempo (Backend Timeout)")
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error(f"Error CRITICO en test_remote_connection: {e}\n{tb}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)} -> {tb}")

def _list_remote_files_sync(req: RemoteRequest):
    try:
        if str(req.protocolo or "").strip().upper() == "API":
            provider = api_provider_name(_api_config_from_remote_request(req))
            return {
                "ruta_actual": req.ruta,
                "items": [{
                    "nombre": "BUNDABERG_API" if provider == "bundaberg" else "STUDIO_G_API",
                    "ruta": req.ruta,
                    "es_dir": False,
                }],
            }
        items = []
        if req.protocolo == "SFTP":
            ssh, sftp = get_sftp_client(req.host, req.puerto, req.usuario, req.password)
            try:
                # Check directly if path exists or listdir
                try:
                    for attr in sftp.listdir_attr(req.ruta):
                        items.append({
                            "nombre": attr.filename,
                            "ruta": f"{req.ruta.rstrip('/')}/{attr.filename}",
                            "es_dir": attr.st_mode is not None and (attr.st_mode & 0o40000) == 0o40000
                        })
                except FileNotFoundError:
                    # Fallback to root or return empty?
                    # logic from before seemed to just fail. 
                    # Let's catch and maybe list root if requested path fails?
                    # For now just re-raise or empty
                     if req.ruta != '.':
                         # Try listing root as fallback
                         for attr in sftp.listdir_attr('.'):
                             items.append({
                                "nombre": attr.filename,
                                "ruta": f"./{attr.filename}",
                                "es_dir": attr.st_mode is not None and (attr.st_mode & 0o40000) == 0o40000
                             })
                         # Signal that we fell back?
                     else:
                         raise
            finally:
                sftp.close()
                ssh.close()
        elif req.protocolo == "FTP":
            ftp = get_ftp_client(req.host, req.puerto, req.usuario, req.password)
            try:
                ftp.cwd(req.ruta)
                entries = []
                try:
                    entries = list(ftp.mlsd()) # name, facts
                    for name, facts in entries:
                         if name in ['.', '..']: continue
                         items.append({
                             "nombre": name,
                             "ruta": f"{req.ruta.rstrip('/')}/{name}",
                             "es_dir": facts.get('type') == 'dir'
                         })
                except:
                    names = ftp.nlst()
                    for name in names:
                         items.append({
                             "nombre": name,
                             "ruta": f"{req.ruta.rstrip('/')}/{name}",
                             "es_dir": '.' not in name 
                         })
            finally:
                ftp.quit()
                
        return {"ruta_actual": req.ruta, "items": sorted(items, key=lambda x: x['nombre'])}
    except Exception as e:
        logger.error(f"Error listando remoto: {e}")
        # Return empty list instead of 500? Or just raise 500
        raise e

@app.post("/api/v1/remote/list")
async def list_remote_files(
    req: RemoteRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(executor, _list_remote_files_sync, req)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/remote-connections", response_model=List[RemoteConnectionResponse])
async def get_remote_connections(
    mall_id: str = Query(..., alias="mall_id"),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return _sensitive_ops_service().list_remote_connections(
            mall_id=mall_id,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing remote connections: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudieron cargar las conexiones remotas.")

@app.post("/api/v1/remote-connections", response_model=RemoteConnectionResponse)
async def create_remote_connection(
    payload: RemoteConnectionCreateRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return _sensitive_ops_service().create_remote_connection(
            payload=payload.dict(),
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error creating remote connection: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo guardar la conexión remota.")

@app.patch("/api/v1/remote-connections/{connection_id}", response_model=RemoteConnectionResponse)
async def update_remote_connection(
    connection_id: str,
    payload: RemoteConnectionUpdateRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return _sensitive_ops_service().update_remote_connection(
            connection_id=connection_id,
            payload=payload.dict(exclude_unset=True),
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except KeyError:
        raise HTTPException(status_code=404, detail="Conexión remota no encontrada.")
    except Exception as e:
        logger.error(f"Error updating remote connection {connection_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo actualizar la conexión remota.")

@app.delete("/api/v1/remote-connections/{connection_id}", status_code=status.HTTP_200_OK)
async def delete_remote_connection(
    connection_id: str,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        _sensitive_ops_service().delete_remote_connection(
            connection_id=connection_id,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
        return {"status": "success", "message": "Conexión remota eliminada correctamente."}
    except HTTPException:
        raise
    except KeyError:
        raise HTTPException(status_code=404, detail="Conexión remota no encontrada.")
    except Exception as e:
        logger.error(f"Error deleting remote connection {connection_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo eliminar la conexión remota.")

@app.get("/api/v1/load-logs", status_code=status.HTTP_200_OK)
async def get_load_logs_secure(
    mall_id: Optional[str] = Query(None, alias="mall_id"),
    local_id: Optional[str] = Query(None, alias="local_id"),
    start_date: Optional[str] = Query(None, alias="start_date"),
    end_date: Optional[str] = Query(None, alias="end_date"),
    limit: int = Query(50, ge=1, le=200),
    operator_ctx: Dict[str, Any] = Depends(require_audit_read_access),
):
    try:
        effective_mall_id = mall_id
        if not effective_mall_id:
            user_malls = _get_user_mall_ids(operator_ctx.get("user_id"))
            if len(user_malls) == 1:
                effective_mall_id = user_malls[0]
            elif len(user_malls) > 1:
                raise HTTPException(status_code=400, detail="Ambiguous context. Please select a mall (mall_id).")
            raise HTTPException(status_code=403, detail="No mall assigned to user.")
        return _sensitive_ops_service().list_load_logs(
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
            mall_id=effective_mall_id,
            local_id=local_id,
            start_date=start_date,
            end_date=end_date,
            limit=limit,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing load logs: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudieron cargar los logs de carga.")

def _clear_load_logs_via_service(mall_id: str, operator_ctx: Dict[str, Any]) -> Dict[str, Any]:
    return _sensitive_ops_service().clear_load_logs(
        mall_id=mall_id,
        operator_ctx=operator_ctx,
        ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
    )

@app.delete("/api/v1/load-logs", status_code=status.HTTP_200_OK)
async def clear_load_logs_secure(
    mall_id: str = Query(..., alias="mall_id"),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return _clear_load_logs_via_service(mall_id, operator_ctx)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error clearing load logs for mall {mall_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo limpiar el historial de cargas.")

@app.delete("/api/v1/audit/logs", status_code=status.HTTP_200_OK)
async def clear_load_logs(
    mall_id: str = Query(..., alias="mall_id"),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    """
    Clears load audit logs only for the selected mall.
    """
    try:
        return _clear_load_logs_via_service(mall_id, operator_ctx)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error clearing logs for mall {mall_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo limpiar el historial de auditoría.")

@app.post("/api/v1/locales/{local_id}/reactivate-processing", status_code=status.HTTP_200_OK)
async def reactivate_local_processing(
    local_id: str,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return _sensitive_ops_service().reactivate_local_processing(
            local_id=local_id,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except KeyError:
        raise HTTPException(status_code=404, detail="Local no encontrado.")
    except Exception as e:
        logger.error(f"Error reactivating local {local_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo reactivar el local.")

@app.get("/api/v1/connections/status", response_model=ConnectionStatusResponse)
async def get_connections_status(
    mall_id: str = Query(..., alias="mall_id"),
    operator_ctx: Dict[str, Any] = Depends(require_audit_read_access)
):
    try:
        return _connection_monitor_service().get_status_summary(
            mall_id=mall_id,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting connection status for mall {mall_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo obtener el estado de conexiones.")

@app.get("/api/v1/connections/failures", response_model=ConnectionFailuresResponse)
async def get_connections_failures(
    mall_id: str = Query(..., alias="mall_id"),
    date_str: str = Query(..., alias="date"),
    operator_ctx: Dict[str, Any] = Depends(require_audit_read_access)
):
    try:
        return _connection_monitor_service().get_failures_by_date(
            mall_id=mall_id,
            run_date=date_str,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error getting connection failures mall={mall_id} date={date_str}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudieron cargar las fallas de conexiones.")

@app.post("/api/v1/connections/retry-failed", status_code=status.HTTP_200_OK)
async def retry_failed_connections_batch(
    mall_id: str = Query(..., alias="mall_id"),
    date_str: str = Query(..., alias="date"),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return await asyncio.to_thread(
            _connection_monitor_service().execute_batch_retry_failed,
            mall_id=mall_id,
            run_date=date_str,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RetryPolicyBlocked as e:
        raise HTTPException(
            status_code=429,
            detail={
                "message": e.message,
                "reason": e.code,
                "retry_after_seconds": e.retry_after_seconds,
            }
        )
    except Exception as e:
        logger.error(f"Error batch retry failed connections mall={mall_id} date={date_str}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudieron ejecutar los reintentos en lote.")

@app.post("/api/v1/connections/{connection_id}/retry", response_model=ConnectionRetryResponse)
async def retry_connection_manual(
    connection_id: str,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        return await asyncio.to_thread(
            _connection_monitor_service().execute_manual_retry,
            connection_id=connection_id,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except HTTPException:
        raise
    except KeyError:
        raise HTTPException(status_code=404, detail="Conexión remota no encontrada.")
    except RetryPolicyBlocked as e:
        raise HTTPException(
            status_code=429,
            detail={
                "message": e.message,
                "reason": e.code,
                "retry_after_seconds": e.retry_after_seconds,
                "attempt_no": e.attempt_no,
            }
        )
    except Exception as e:
        logger.error(f"Error manual retry connection {connection_id}: {sanitize_sensitive_ops_error(e)}")
        raise HTTPException(status_code=500, detail="No se pudo ejecutar el reintento de conexión.")

def _read_remote_headers_sync(req: RemoteRequest):
    content = ""
    if req.protocolo == "SFTP":
        ssh, sftp = get_sftp_client(req.host, req.puerto, req.usuario, req.password)
        try:
            with sftp.open(req.ruta, 'r') as f:
                if req.tipo_archivo in ["JSON", "XML"]:
                    content = f.read().decode('utf-8')
                else:
                    head = [next(f) for _ in range(2)]
                    content = "".join(head)
        finally:
            sftp.close()
            ssh.close()
    elif req.protocolo == "FTP":
        ftp = get_ftp_client(req.host, req.puerto, req.usuario, req.password)
        try:
            bio = io.BytesIO()
            ftp.retrbinary(f"RETR {req.ruta.split('/')[-1]}", bio.write)
            bio.seek(0)
            content = bio.read().decode('utf-8')
        finally:
            ftp.quit()

    headers = []
    if req.tipo_archivo in ["CSV", "TXT"]:
            # logic ...
            try:
                dialect = csv.Sniffer().sniff(content)
                reader = csv.reader(io.StringIO(content), dialect)
            except:
                reader = csv.reader(io.StringIO(content))
            headers = next(reader)
            
    elif req.tipo_archivo == "JSON":
        data = json.loads(content)
        if isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
        elif isinstance(data, dict):
            found_list = False
            for k, v in data.items():
                if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                        headers = list(v[0].keys())
                        found_list = True
                        break
            if not found_list:
                headers = list(data.keys())
                
    elif req.tipo_archivo == "XML":
        data = xmltodict.parse(content)
        def find_keys(d):
            for k, v in d.items():
                if isinstance(v, list): 
                    if len(v) > 0 and isinstance(v[0], dict):
                        return list(v[0].keys())
                elif isinstance(v, dict):
                    res = find_keys(v)
                    if res: return res
            return list(d.keys())
        
        headers = find_keys(data)

    return {"headers": headers}

@app.post("/api/v1/remote/headers")
async def read_remote_headers(
    req: RemoteRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(executor, _read_remote_headers_sync, req)
    except Exception as e:
         logger.error(f"Error leyendo headers remotos: {e}")
         raise HTTPException(status_code=500, detail=str(e))


# --- Mantenimiento de Usuarios y Locales (Mantenidos para funcionalidad UI) ---
@app.get("/api/v1/usuarios", response_model=List[UserSchema])
async def get_users():
    return [
        {"id": "1", "nombre": "Admin Auditor", "email": "admin@msmall.com", "rol": "admin", "estado": "activo", "created_at": "2024-01-01", "ultimo_acceso": "Hace 5 min"},
        {"id": "2", "nombre": "Roberto Carlos", "email": "rcarlos@mallplaza.com", "rol": "mall_manager", "estado": "activo", "created_at": "2024-02-15", "ultimo_acceso": "Ayer"}
    ]

@app.get("/api/v1/locales", response_model=List[StoreSchema])
async def get_stores():
    return [
        {
          "id": "64d82d1a-8893-4913-a9c5-d79b3221710e",
          "mall_id": "c23e99b6-8feb-4be8-8842-86c263bc5cad",
          "codigo_interno": "l002",
          "nombre": "Adidas",
          "email": "notificaciones@adidas.example",
          "email_secundario": "respaldo@adidas.example",
          "rubro": "Deporte",
          "created_at": "2026-01-27T15:33:02",
          "responsable": "Jose Perez",
          "contrato_no": "99812-91283",
          "piso": "P2-L123",
          "tipo_negocio": "Ropa Deportiva",
          "mts": "150.00",
          "porciento_renta": "2.00"
        }
    ]


@app.post("/api/v1/locales", status_code=status.HTTP_201_CREATED)
async def create_store_backend(
    payload: Dict[str, Any] = Body(...),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")
    data = _sanitize_store_write_payload(payload)
    mall_id = str(data.get("mall_id") or "").strip()
    if not mall_id:
        raise HTTPException(status_code=400, detail="mall_id requerido")
    if not data.get("nombre") or not data.get("codigo_interno"):
        raise HTTPException(status_code=400, detail="nombre y codigo_interno son requeridos")
    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    _validate_store_catalog_values(mall_id, data)
    try:
        response = supabase.table("locales").insert(data).execute()
        if not response.data:
            raise ValueError("No se recibió el local creado desde Supabase")
        row = response.data[0]
        return row
    except Exception as e:
        logger.error("Error creando local mall=%s user=%s: %s", mall_id, operator_ctx.get("user_id"), e)
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/v1/locales/{local_id}")
async def update_store_backend(
    local_id: str,
    payload: Dict[str, Any] = Body(...),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")
    existing = supabase.table("locales").select("id,mall_id").eq("id", local_id).maybe_single().execute().data
    if not existing:
        raise HTTPException(status_code=404, detail="Local no encontrado")
    mall_id = str(existing.get("mall_id") or "")
    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    data = _sanitize_store_write_payload(payload, existing_mall_id=mall_id)
    data.pop("mall_id", None)
    if "nombre" in data and not data.get("nombre"):
        raise HTTPException(status_code=400, detail="nombre es requerido")
    if "codigo_interno" in data and not data.get("codigo_interno"):
        raise HTTPException(status_code=400, detail="codigo_interno es requerido")
    if not data:
        raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")
    _validate_store_catalog_values(mall_id, data)
    try:
        response = supabase.table("locales").update(data).eq("id", local_id).execute()
        if not response.data:
            raise ValueError("No se recibió el local actualizado desde Supabase")
        row = response.data[0]
        return row
    except Exception as e:
        logger.error("Error actualizando local=%s user=%s: %s", local_id, operator_ctx.get("user_id"), e)
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/v1/locales/{local_id}", status_code=status.HTTP_200_OK)
async def delete_store_backend(
    local_id: str,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")
    existing = supabase.table("locales").select("id,mall_id").eq("id", local_id).maybe_single().execute().data
    if not existing:
        raise HTTPException(status_code=404, detail="Local no encontrado")
    _ensure_operator_can_access_mall(operator_ctx, str(existing.get("mall_id") or ""))
    try:
        supabase.table("locales").delete().eq("id", local_id).execute()
        return {"status": "success"}
    except Exception as e:
        logger.error("Error eliminando local=%s user=%s: %s", local_id, operator_ctx.get("user_id"), e)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/locales/custom-fields")
async def list_local_custom_fields(
    mall_id: str = Query(...),
    include_inactive: bool = Query(True),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    return _local_custom_fields_service().list_definitions(mall_id, include_inactive=include_inactive)


@app.post("/api/v1/locales/custom-fields")
async def create_local_custom_field(
    request: CustomFieldDefinitionCreateRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    return _local_custom_fields_service().create_definition(
        request.dict(),
        operator_ctx=operator_ctx,
        ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
    )


@app.patch("/api/v1/locales/custom-fields/{field_id}")
async def update_local_custom_field(
    field_id: str,
    request: CustomFieldDefinitionUpdateRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    return _local_custom_fields_service().update_definition(
        field_id,
        request.dict(exclude_unset=True),
        operator_ctx=operator_ctx,
        ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
    )


@app.get("/api/v1/locales/{local_id}/custom-fields")
async def get_local_custom_fields(
    local_id: str,
    include_inactive: bool = Query(False),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    return _local_custom_fields_service().get_local_fields(
        local_id,
        operator_ctx=operator_ctx,
        ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        include_inactive=include_inactive,
    )


@app.put("/api/v1/locales/{local_id}/custom-fields")
async def upsert_local_custom_fields(
    local_id: str,
    request: LocalCustomFieldValueUpsertRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    return _local_custom_fields_service().upsert_local_values(
        local_id,
        [value.dict(exclude_unset=True) for value in request.values],
        operator_ctx=operator_ctx,
        ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
    )

# --- AI & INSIGHTS ENDPOINTS ---
@app.get("/api/v1/insights/alerts")
async def get_intelligent_alerts(local_id: Optional[str] = None):
    """Fetch a real antifraud snapshot for a store."""
    if not supabase or not local_id:
        return {
            "status": "no_data",
            "source": "none",
            "alerts": [],
            "summary": {
                "risk_state": "NO_DATA",
                "risk_label": "Sin Datos",
                "description": "Selecciona un local con ventas para evaluar el semaforo.",
                "last_evaluated_at": None,
                "has_recent_run": False,
                "risk_score": 0,
                "alerts_count": 0,
                "analysis_window_days": 30,
            },
        }
    try:
        return await asyncio.to_thread(
            lambda: _analytics_service().get_alert_snapshot(local_id, allow_live_refresh=True)
        )
    except Exception as exc:
        logger.error("Error obteniendo snapshot IA para local %s: %s", local_id, exc)
        return {
            "status": "error",
            "source": "none",
            "alerts": [],
            "summary": {
                "risk_state": "NO_DATA",
                "risk_label": "Sin Datos",
                "description": "No se pudo evaluar el semaforo en este momento.",
                "last_evaluated_at": None,
                "has_recent_run": False,
                "risk_score": 0,
                "alerts_count": 0,
                "analysis_window_days": 30,
            },
        }

@app.get("/api/v1/insights/benchmarking/{local_id}")
async def get_benchmarking(local_id: str):
    """Compare local performance vs category average based on real data."""
    if not supabase: return None
    try:
        # 1. Get Store Info
        store_res = supabase.table("locales").select("id, nombre, rubro").eq("id", local_id).single().execute()
        if not store_res.data: return None
        
        # 2. Get Sales ATV
        sales_res = supabase.table("ventas").select("total_bruto").eq("local_id", local_id).execute()
        if not sales_res.data:
            return {
                "local_name": store_res.data['nombre'],
                "local_value": 0, "category_avg": 0, "status": "Sin datos",
                "atv_local": 0, "atv_category": 0, "atv_growth": "0%"
            }
        
        local_total = sum(float(r['total_bruto']) for r in sales_res.data)
        atv_local = local_total / len(sales_res.data)
        
        # 3. Get Category Average
        rubro = store_res.data.get('rubro')
        atv_category = atv_local # Default
        if rubro:
            cat_stores = supabase.table("locales").select("id").eq("rubro", rubro).execute()
            cat_ids = [s['id'] for s in cat_stores.data]
            cat_sales = supabase.table("ventas").select("total_bruto").in_("local_id", cat_ids).execute()
            if cat_sales.data:
                atv_category = sum(float(r['total_bruto']) for r in cat_sales.data) / len(cat_sales.data)

        return {
            "local_name": store_res.data['nombre'],
            "local_value": local_total,
            "category_avg": atv_category * len(sales_res.data),
            "status": "Líder" if atv_local > atv_category else "Promedio",
            "atv_local": round(atv_local, 2),
            "atv_category": round(atv_category, 2),
            "atv_growth": "0%"
        }
    except Exception as e:
        logger.error(f"Error benchmarking: {e}")
        return None

@app.get("/api/v1/insights/efficiency/{local_id}")
async def get_efficiency(local_id: str):
    """Calculate Real Estate Efficiency metrics from real store and sales data."""
    if not supabase: return None
    try:
        # 1. Get Store MTS
        store_res = supabase.table("locales").select("mts, nombre, porciento_renta").eq("id", local_id).single().execute()
        if not store_res.data: return None
        
        mts = float(store_res.data.get('mts') or 1.0)
        
        # 2. Sum Sales
        sales_res = supabase.table("ventas").select("total_neto").eq("local_id", local_id).execute()
        total_sales = sum(float(r['total_neto']) for r in sales_res.data) if sales_res.data else 0
        
        if total_sales == 0:
            return {
                "sales_per_m2": 0, "occupancy_cost_ratio": 0, "is_healthy": True, "risk_level": "BAJO", "message": "Sin datos"
            }

        # 3. Simple Mock Renta (unless added to DB)
        renta_fija = 2500  # Placeholder
        gastos_comunes = 600 # Placeholder
        
        sales_per_m2 = total_sales / mts
        occupancy_cost_ratio = (renta_fija + gastos_comunes) / total_sales
        
        return {
            "sales_per_m2": round(sales_per_m2, 2),
            "occupancy_cost_ratio": round(occupancy_cost_ratio * 100, 2),
            "is_healthy": occupancy_cost_ratio < 0.15,
            "risk_level": "BAJO" if occupancy_cost_ratio < 0.15 else "MEDIO" if occupancy_cost_ratio < 0.20 else "ALTO",
            "message": "Operación saludable"
        }
    except Exception as e:
        logger.error(f"Error efficiency: {e}")
        return None

@app.get("/api/v1/insights/heatmap/{local_id}")
async def get_heatmap(local_id: str):
    """Generate sales intensity heatmap data from real transaction times."""
    if not supabase: return []
    cache_key = f"insights:heatmap:{local_id}"
    cached = _cache_get(cache_key)
    if cached is not _CACHE_MISS:
        return cached
    try:
        rpc_res = supabase.rpc("get_insights_heatmap", {"local_id_param": local_id}).execute()
        if rpc_res.data:
            result = rpc_res.data
            _cache_set(cache_key, result, TTL_HEATMAP)
            return result
        _cache_set(cache_key, [], TTL_HEATMAP)
        return []
    except Exception as rpc_err:
        logger.warning(f"Heatmap RPC unavailable, fallback to python aggregation: {rpc_err}")
    try:
        res = supabase.table("ventas").select("fecha, hora_transaccion").eq("local_id", local_id).limit(2000).execute()
        if not res.data:
            _cache_set(cache_key, [], TTL_HEATMAP)
            return []
        
        counts = {}
        days_map = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}
        
        for r in res.data:
            dt = datetime.strptime(r['fecha'], '%Y-%m-%d')
            day_name = days_map[dt.weekday()]
            hora_str = r.get('hora_transaccion') or '12:00:00'
            hour_val = int(hora_str.split(':')[0])
            # Match UI blocks
            block = (hour_val // 2) * 2
            if block < 10: block = 10
            if block > 22: block = 22
            key = (day_name, f"{block:02d}:00")
            counts[key] = counts.get(key, 0) + 1
            
        max_count = max(counts.values()) if counts else 1
        result = []
        days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        hours = ["10:00", "12:00", "14:00", "16:00", "18:00", "20:00", "22:00"]
        for d in days:
            for h in hours:
                val = (counts.get((d, h), 0) / max_count) * 100
                result.append({"dia": d, "hora": h, "valor": round(val, 2)})
        _cache_set(cache_key, result, TTL_HEATMAP)
        return result
    except:
        return []

@app.get("/api/v1/insights/ranking")
async def get_ranking(metric: str, mall_id: Optional[str] = Query(None, alias="mall_id")):
    """Get ranking of all stores for a specific metric based on real database data."""
    if not supabase: return []
    cache_key = f"insights:ranking:{metric}:{mall_id or 'all'}"
    cached = _cache_get(cache_key)
    if cached is not _CACHE_MISS:
        return cached
    try:
        rpc_res = supabase.rpc("get_insights_ranking", {
            "metric_param": metric,
            "mall_id_param": mall_id
        }).execute()
        if rpc_res.data:
            # Ensure JSON-serializable primitive types
            normalized = []
            for row in rpc_res.data:
                normalized.append({
                    "id": row.get("id"),
                    "nombre": row.get("nombre"),
                    "valor": float(row.get("valor") or 0),
                    "extra": row.get("extra")
                })
            _cache_set(cache_key, normalized, TTL_RANKING)
            return normalized
        _cache_set(cache_key, [], TTL_RANKING)
        return []
    except Exception as rpc_err:
        logger.warning(f"Ranking RPC unavailable, fallback to python aggregation: {rpc_err}")
    try:
        # 1. Fetch all stores
        query = supabase.table("locales").select("id, nombre, mts, rubro")
        if mall_id:
            query = query.eq("mall_id", mall_id)
        
        stores_res = query.execute()
        if not stores_res.data: return []
        
        # 2. Fetch all sales
        sales_query = supabase.table("ventas").select("local_id, total_bruto, total_neto")
        if mall_id:
            sales_query = sales_query.eq("mall_id", mall_id)
        sales_res = sales_query.execute()
        
        # Aggregate
        sales_data = {} # id -> {bruto, neto, cnt}
        for s in sales_res.data:
            lid = s['local_id']
            if lid not in sales_data: sales_data[lid] = {'bruto': 0, 'neto': 0, 'cnt': 0}
            sales_data[lid]['bruto'] += float(s['total_bruto'])
            sales_data[lid]['neto'] += float(s['total_neto'])
            sales_data[lid]['cnt'] += 1
            
        ranking = []
        for s in stores_res.data:
            stats = sales_data.get(s['id'], {'bruto': 0, 'neto': 0, 'cnt': 0})
            
            valor = 0
            extra = s.get('rubro') or "General"
            
            if metric == 'sales_per_m2':
                mts = float(s.get('mts') or 1.0)
                valor = stats['neto'] / mts
                extra = f"{mts} m²"
            elif metric == 'occupancy_cost':
                # Use a placeholder for rent until it's in DB
                costos = 3000 
                valor = (costos / stats['neto'] * 100) if stats['neto'] > 0 else 0
                extra = "Saludable" if (0 < valor < 15) else "Riesgo" if valor >= 15 else "Sin Ventas"
            
            ranking.append({
                "id": s['id'],
                "nombre": s['nombre'],
                "valor": round(valor, 2),
                "extra": extra
            })
            
        # Sort desc
        ranking.sort(key=lambda x: x['valor'], reverse=True)
        _cache_set(cache_key, ranking, TTL_RANKING)
        return ranking
    except Exception as e:
        logger.error(f"Error in ranking: {e}")
        return []

class CubeRequest(BaseModel):
    fecha_inicio: str
    fecha_fin: str
    agrupacion: str = "DIA" # DIA, SEMANA, MES
    metrica: str = "total_neto" # total_neto, total_bruto, transacciones
    local_id: Optional[str] = None
    custom_dimension_key: Optional[str] = None
    custom_filters: Optional[Dict[str, Any]] = None

# --- INTELLIGENT AUTO-MAPPING ---
SYSTEM_FIELDS_SYNONYMS = {
    "factura_numero": ["invoice", "factura", "doc_num", "documento", "folio", "ticket", "recibo", "invoiceNumber", "invoice_id"],
    "fecha_venta": ["date", "fecha", "time", "dia", "issued", "created", "invoiceDate"],
    "local_codigo": ["store", "local", "tienda", "sucursal", "code", "id_local", "storeCode", "terminalCode"],
    "total_bruto": ["gross", "bruto", "total", "amount", "monto", "venta", "precio", "importe", "grandTotal", "totals.grandTotal", "paymentTotal"],
    "total_impuestos": ["tax", "impuesto", "iva", "vat", "tributes", "taxTotal", "totals.taxTotal", "taxAmount"],
    "total_neto": ["net", "neto", "subtotal", "base", "subTotal", "totals.subTotal"],
    "comprobante": ["ticket", "vourcher", "comprobante", "recibo", "doc_type", "ncf", "fiscalData.ncf"],
    "hora_transaccion": ["time", "hora", "trans_hour", "momento"]
}


def _remote_analysis_timeout_seconds(filename: Optional[str], tipo_archivo: Optional[str] = None) -> float:
    normalized_type = str(tipo_archivo or "").strip().upper()
    normalized_name = str(filename or "").strip().lower()
    if normalized_type == "JSON" or normalized_name.endswith(".json"):
        return 420.0
    return 180.0




def _perform_mapping_analysis(decoded_content, filename, tipo_archivo=None, force_has_header: Optional[bool] = None, data_start_row: Optional[int] = None):
    headers: List[str] = []
    sample_row: Dict[str, Any] = {}
    normalized_decoded_content = _normalize_text_for_csv(decoded_content)
    raw_preview_lines = _build_raw_preview_lines(normalized_decoded_content)
    detected_delimiter: Optional[str] = None
    detected_has_header: Optional[bool] = None

    # Normalize tipo_archivo if provided
    current_type = tipo_archivo.upper() if tipo_archivo else None
    if not current_type:
        if filename.lower().endswith('.json'):
            current_type = "JSON"
        elif filename.lower().endswith('.csv') or filename.lower().endswith('.txt'):
            current_type = "CSV"

    def _payload(
        csv_headers: List[str],
        suggested_mapping: Dict[str, Any],
        sample: Dict[str, Any],
        detected_headers: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        return {
            "csv_headers": csv_headers,
            "headers": csv_headers,
            "detected_headers": detected_headers if detected_headers is not None else csv_headers,
            "suggested_mapping": suggested_mapping,
            "sample_row": sample,
            "raw_preview_lines": raw_preview_lines,
            "analysis_type": current_type,
            "detected_delimiter": detected_delimiter,
            "detected_has_header": detected_has_header
        }

    # 1. Detect Format and Extract Headers/Sample
    if current_type == "CSV" or current_type == "TXT" or not current_type:
        sample_str = normalized_decoded_content[:4096]
        delimiter, has_header = _detect_delimiter_and_header(sample_str)
        if force_has_header is not None:
            has_header = force_has_header
        if has_header and _should_treat_as_no_header(normalized_decoded_content, delimiter):
            has_header = False
        detected_delimiter = delimiter
        detected_has_header = has_header

        f = io.StringIO(normalized_decoded_content)
        if has_header:
            reader = csv.DictReader(f, delimiter=delimiter, skipinitialspace=True)
            try:
                all_rows = [_normalize_csv_row_keys(r) for r in reader]
                if data_start_row and data_start_row > 2:
                    all_rows = all_rows[data_start_row - 2:]
                row1 = all_rows[0] if all_rows else None
                headers = [_clean_csv_header_name(h) for h in (reader.fieldnames or []) if _clean_csv_header_name(h)]
                sample_row = row1 or {}
                if not row1:
                    return _payload(headers, {}, {}, headers)
            except Exception:
                return _payload([], {}, {}, [])
        else:
            raw_reader = csv.reader(f, delimiter=delimiter, skipinitialspace=True)
            matrix_rows = [r for r in raw_reader if any(str(c or "").strip() for c in r)]
            if data_start_row and data_start_row > 1:
                matrix_rows = matrix_rows[data_start_row - 1:]
            if not matrix_rows:
                return _payload([], {}, {}, [])

            max_cols = max(len(r) for r in matrix_rows)
            headers = [f"col_{idx}" for idx in range(1, max_cols + 1)]
            first_row = list(matrix_rows[0]) + [""] * (max_cols - len(matrix_rows[0]))
            sample_row = dict(zip(headers, first_row))

    if current_type == "JSON":
        try:
            data = json.loads(decoded_content)

            target_data = data
            if isinstance(data, dict):
                if "invoices" in data:
                    target_data = data["invoices"]
                else:
                    found_list = False
                    for _, v in data.items():
                        if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                            target_data = v
                            found_list = True
                            break
                    if not found_list:
                        target_data = [data]
            elif isinstance(data, list):
                target_data = data
            else:
                target_data = [data]

            df = pd.json_normalize(target_data)

            if df.empty:
                headers = []
            else:
                headers = list(df.columns)
                if len(df) > 0:
                    sample_row = df.iloc[0].where(pd.notnull(df.iloc[0]), None).to_dict()

        except Exception as e:
            logger.error(f"Error parsing JSON analysis: {e}")

    if not headers:
        return _payload([], {}, {}, [])

    # 2. Fuzzy Match System Fields
    suggested_mapping: Dict[str, Any] = {}
    no_header_columns = all(isinstance(h, str) and h.lower().startswith("col_") for h in headers)

    if no_header_columns:
        positional_defaults = {
            "factura_numero": "col_1",
            "local_codigo": "col_2",
            "fecha_venta": "col_3",
            "total_bruto": "col_8",
            "total_impuestos": "col_9",
            "total_neto": "col_10",
        }
        for sys_field, col_name in positional_defaults.items():
            if col_name in headers:
                suggested_mapping[sys_field] = {
                    "csv_header": col_name,
                    "confidence": 95,
                    "is_confident": True
                }
        return _payload(headers, suggested_mapping, sample_row, headers)

    for sys_field, synonyms in SYSTEM_FIELDS_SYNONYMS.items():
        query_list = [sys_field] + synonyms
        best_match = None
        best_score = 0

        for h in headers:
            for q in query_list:
                if h == q:
                    best_match = h
                    best_score = 100
                    break
            if best_score == 100:
                break

        if best_score < 100:
            for query in query_list:
                match, score = process.extractOne(query, headers, scorer=fuzz.token_sort_ratio) or (None, 0)
                if score > best_score:
                    best_score = score
                    best_match = match

        if best_score > 60:
            suggested_mapping[sys_field] = {
                "csv_header": best_match,
                "confidence": best_score,
                "is_confident": best_score > 80
            }

    return _payload(headers, suggested_mapping, sample_row, headers)

@app.post("/api/v1/mapping/analyze")
async def analyze_mapping(file: UploadFile = File(...)):
    """
    Analyzes a sample file (CSV/JSON) and suggests mapping to system fields using fuzzy logic.
    """
    try:
        content = await file.read()
        # Use utf-8-sig to handle BOM which is common in Windows/Excel generated files
        decoded = content.decode('utf-8-sig', errors='replace')
        return _perform_mapping_analysis(decoded, file.filename) # For upload we usually trust extension or could add more logic
    except Exception as e:
        logger.error(f"Error analyzing mapping: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/mapping/analyze-remote")
async def analyze_remote_mapping(
    req: RemoteRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    """
    Connects to SFTP/FTP, reads a sample of the file, and returns mapping suggestions.
    """
    try:
        content = ""
        loop = asyncio.get_event_loop()
        analysis_timeout_seconds = _remote_analysis_timeout_seconds(req.ruta, req.tipo_archivo)
        
        def _read_remote_sample():
            # Determine if we should read all (JSON) or sample (CSV)
            is_json = req.tipo_archivo == "JSON" or req.ruta.lower().endswith('.json')
            read_size = -1 if is_json else 32768 # Read all for JSON, 32KB for CSV (increased from 8KB)

            if str(req.protocolo or "").strip().upper() == "API":
                return json.dumps(_api_preview_rows(req), ensure_ascii=False)
            if req.protocolo == "SFTP":
                ssh, sftp = get_sftp_client(req.host, req.puerto, req.usuario, req.password)
                try:
                    target_path = req.ruta
                    # If route points to directory, fail with deterministic message.
                    st = sftp.stat(target_path)
                    if stat.S_ISDIR(st.st_mode):
                        raise FileNotFoundError(f"La ruta '{target_path}' es un directorio. Seleccione un archivo.")
                    with sftp.open(target_path, 'rb') as f:
                        payload = f.read() if is_json else f.read(read_size)
                        return _decode_remote_text(payload, is_json=is_json)
                finally:
                    sftp.close()
                    ssh.close()
            elif req.protocolo == "FTP":
                ftp = get_ftp_client(req.host, req.puerto, req.usuario, req.password)
                try:
                    target_path = (req.ruta or "").replace("\\", "/")
                    target_name = posixpath.basename(target_path)
                    target_dir = posixpath.dirname(target_path)
                    used_dir = _ftp_enter_target_dir(ftp, target_dir or ".")
                    raw_bytes, _ = _download_ftp_file_bytes(
                        ftp=ftp,
                        requested_filename=target_name,
                        remote_path=used_dir,
                        max_bytes=None if is_json else read_size
                    )
                    return _decode_remote_text(raw_bytes, is_json=is_json)
                finally:
                    ftp.quit()
            return ""

        # Usar wait_for para evitar hangs si el archivo es gigante o la red falla
        content = await asyncio.wait_for(
            loop.run_in_executor(executor, _read_remote_sample),
            timeout=analysis_timeout_seconds
        )
        if not content:
            return {"headers": [], "suggested_mapping": {}, "sample_row": {}}
            
        return _perform_mapping_analysis(
            content,
            req.ruta,
            req.tipo_archivo,
            force_has_header=req.has_header,
            data_start_row=req.data_start_row
        )
        
    except asyncio.TimeoutError:
        raise HTTPException(
            status_code=504,
            detail="Timeout analizando archivo remoto (el archivo podria ser grande y la primera carga puede tardar varios minutos)",
        )
    except Exception as e:
        logger.error(f"Error analyzing remote mapping: {e}")
        raise HTTPException(status_code=500, detail=str(e))

import posixpath

def _ftp_enter_target_dir(ftp: FTP, remote_path: str) -> str:
    """
    Try to enter remote_path; if it points to a file or is invalid, fallback to parent.
    Returns the directory that was attempted/used.
    """
    normalized = (remote_path or ".").replace("\\", "/").strip() or "."
    try:
        ftp.cwd(normalized)
        return normalized
    except Exception:
        parent = posixpath.dirname(normalized) or "."
        try:
            ftp.cwd(parent)
            return parent
        except Exception:
            logger.warning(
                f"No se pudo cambiar a ruta FTP '{normalized}' ni a su padre '{parent}'. "
                "Se usará el directorio actual de sesión."
            )
            try:
                return ftp.pwd() or "."
            except Exception:
                return "."

def _parse_ftp_list_line(line: str) -> Optional[Tuple[str, bool]]:
    """
    Parses a LIST line and preserves full filename (including spaces/commas).
    Returns (name, is_dir) or None if parsing failed.
    """
    if not line:
        return None

    # Unix LIST format: perms links owner group size month day time/year name
    unix_parts = line.split(maxsplit=8)
    if len(unix_parts) == 9 and unix_parts[0] and unix_parts[0][0] in ("d", "-", "l"):
        name = unix_parts[8].strip()
        if name in (".", "..") or " -> " in name:
            return None
        return name, unix_parts[0][0] == "d"

    # Windows/DOS LIST format: MM-DD-YY HH:MMPM <DIR>|size name
    m = re.match(r"^\d{2}-\d{2}-\d{2,4}\s+\d{2}:\d{2}(?:AM|PM)\s+(\<DIR\>|\d+)\s+(.+)$", line, re.IGNORECASE)
    if m:
        marker = m.group(1).upper()
        name = m.group(2).strip()
        if name in (".", ".."):
            return None
        return name, marker == "<DIR>"

    # Best effort fallback preserving tail as filename.
    generic_parts = line.split(maxsplit=8)
    if len(generic_parts) >= 9:
        name = generic_parts[8].strip()
        if name and name not in (".", ".."):
            return name, False
    return None

def _resolve_ftp_filename(ftp: FTP, requested_filename: str, remote_path: str) -> str:
    """
    Resolves a potentially truncated FTP filename (e.g. due to LIST parsing) to the real file.
    """
    requested = (requested_filename or "").strip()
    remote_basename = posixpath.basename((remote_path or "").replace("\\", "/").strip())

    candidates: List[str] = []
    for candidate in [requested, posixpath.basename(requested), remote_basename]:
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    # Fast exact/basename match from current directory listing.
    try:
        names = ftp.nlst() or []
    except Exception:
        names = []

    normalized_names: List[str] = []
    for raw_name in names:
        if not raw_name:
            continue
        cleaned = raw_name.rstrip("/")
        if cleaned and cleaned not in (".", ".."):
            normalized_names.append(cleaned)

    for cand in candidates:
        for listed in normalized_names:
            if listed == cand or posixpath.basename(listed) == cand:
                return posixpath.basename(listed)

    # Suffix match fallback for cases like "Ventas XYZ, SRL29012026.txt" selected as "SRL29012026.txt".
    for cand in candidates:
        low = cand.lower()
        matches = [posixpath.basename(n) for n in normalized_names if n.lower().endswith(low)]
        if len(matches) == 1:
            return matches[0]

    return requested or remote_basename or ""

def _download_ftp_file_bytes(
    ftp: FTP,
    requested_filename: str,
    remote_path: str,
    max_bytes: Optional[int] = None
) -> Tuple[bytes, str]:
    """
    Download an FTP file trying robust candidate paths.
    Prefers non-empty payload when multiple candidates succeed.
    """
    requested_clean = posixpath.basename(str(requested_filename or "").strip().strip("'\""))
    resolved_filename = _resolve_ftp_filename(ftp, requested_clean, remote_path) or requested_clean
    remote_norm = (remote_path or "").replace("\\", "/").strip()
    remote_dir = posixpath.dirname(remote_norm) if remote_norm else ""
    remote_base = posixpath.basename(remote_norm) if remote_norm else ""

    listed_basenames: List[str] = []
    try:
        names = ftp.nlst() or []
        for raw_name in names:
            cleaned = posixpath.basename(str(raw_name or "").rstrip("/"))
            if cleaned and cleaned not in (".", ".."):
                listed_basenames.append(cleaned)
    except Exception:
        listed_basenames = []

    exact_matches = [
        name for name in listed_basenames
        if requested_clean and (name == requested_clean or name.lower() == requested_clean.lower())
    ]

    def _normalized_stem(name: str) -> str:
        stem = posixpath.splitext(posixpath.basename(name or ""))[0].strip().lower()
        if stem.startswith("pr_"):
            stem = stem[3:]
        for ch in [" ", "-", "_"]:
            stem = stem.replace(ch, "")
        return stem

    requested_stem = _normalized_stem(requested_clean)
    related_matches: List[str] = []
    for listed in listed_basenames:
        listed_stem = _normalized_stem(listed)
        if not listed_stem or not requested_stem:
            continue
        if listed_stem == requested_stem:
            related_matches.append(listed)
            continue
        # Keep permissive suffix relation for legacy names with extra prefixes/suffixes.
        if listed.lower().endswith(requested_clean.lower()) or requested_clean.lower().endswith(listed.lower()):
            related_matches.append(listed)

    candidates: List[str] = []
    if exact_matches:
        for candidate in exact_matches:
            if candidate not in candidates:
                candidates.append(candidate)
    for candidate in related_matches:
        if candidate not in candidates:
            candidates.append(candidate)
    for candidate in [requested_clean, resolved_filename, remote_base]:
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    # Explicit PR_ / non-PR toggles.
    if requested_clean:
        if requested_clean.upper().startswith("PR_"):
            alt = requested_clean[3:]
            if alt and alt not in candidates:
                candidates.append(alt)
        else:
            alt = f"PR_{requested_clean}"
            if alt not in candidates:
                candidates.append(alt)

    # Also try absolute-like paths; some FTP servers behave better this way.
    for base_name in list(candidates):
        if not base_name:
            continue
        for base_dir in [remote_norm, remote_dir]:
            if not base_dir:
                continue
            joined = posixpath.join(base_dir, base_name)
            if joined not in candidates:
                candidates.append(joined)

    class _LimitedWriter:
        def __init__(self, limit: Optional[int] = None):
            self.bio = io.BytesIO()
            self.limit = limit
            self.written = 0

        def write(self, data: bytes):
            if self.limit is not None and self.written >= self.limit:
                raise StopIteration
            chunk = data
            if self.limit is not None:
                remaining = self.limit - self.written
                if remaining <= 0:
                    raise StopIteration
                chunk = data[:remaining]
            if chunk:
                self.bio.write(chunk)
                self.written += len(chunk)

    def _score_payload(data: bytes) -> Tuple[int, int, int]:
        if not data:
            return (0, 0, 0)
        text = _normalize_text_for_csv(_decode_remote_text(data, is_json=False))
        lines = [ln.strip() for ln in text.split("\n") if str(ln).strip()]
        if not lines:
            return (0, 0, len(data))
        first = lines[0]
        delimiters = [",", ";", "\t", "|"]
        best_delim = max(delimiters, key=lambda d: first.count(d))
        structured = sum(1 for ln in lines[:500] if ln.count(best_delim) >= 1)
        return (structured, len(lines), len(data))

    best_payload = b""
    best_candidate = ""
    best_score: Tuple[int, int, int] = (0, 0, 0)
    last_error: Optional[Exception] = None

    for candidate in candidates:
        try:
            writer = _LimitedWriter(limit=max_bytes)
            try:
                ftp.retrbinary(f"RETR {candidate}", writer.write)
            except StopIteration:
                pass  # expected for limited reads

            payload = writer.bio.getvalue()
            score = _score_payload(payload)
            if score > best_score:
                best_payload = payload
                best_candidate = candidate
                best_score = score
            if payload:
                logger.info(f"Descarga FTP candidata: {candidate} ({len(payload)} bytes) score={score}")
        except Exception as retr_err:
            last_error = retr_err

    if best_candidate:
        logger.info(
            f"Descarga FTP seleccionada: {best_candidate} bytes={len(best_payload)} score={best_score}"
        )
        return best_payload, posixpath.basename(best_candidate)

    raise FileNotFoundError(
        f"No se pudo descargar '{requested_filename}' por FTP. Último error: {last_error}"
    )

def _list_remote_files(config: Dict[str, Any]):
    # Normalize keys (handles frontend names and worker-style names from Supabase 'locales')
    protocolo = config.get("protocolo") or config.get("sftp_protocol", "SFTP")
    host = config.get("host") or config.get("sftp_host")
    puerto = config.get("puerto") or config.get("sftp_port", 22)
    usuario = config.get("usuario") or config.get("sftp_user")
    password = config.get("password") or config.get("sftp_pass")
    ruta = config.get("ruta_remota") or config.get("sftp_path", ".")
    tipo_archivo = config.get("tipo_archivo") or config.get("file_type", "CSV")
    logger.info(f"[DEBUG_AUTH] User: '{usuario}', PassLen: {len(password) if password else 0}, Host: '{host}', Port: {puerto}, Path: '{ruta}'")
    if str(protocolo or "").strip().upper() == "API":
        provider = api_provider_name(config)
        return [{
            "nombre": "BUNDABERG_API" if provider == "bundaberg" else "STUDIO_G_API",
            "fecha": datetime.utcnow().isoformat(),
            "tamano": 0,
        }]
    
    # Allow all supported extensions to be listed, to prevent confusion if config doesn't match file
    # ext = ".csv" if tipo_archivo == "CSV" else ".txt" if tipo_archivo == "TXT" else ".json"
    supported_exts = (".csv", ".txt", ".json")
    
    if not host or not usuario:
        logger.error(f"Missing connection parameters: host={host}, user={usuario}")
        return []

    files = []
    if protocolo == "SFTP":
        ssh, sftp = get_sftp_client(host, puerto, usuario, password)
        try:
            # Si la ruta es un archivo, listar su directorio contenedor
            try:
                logger.info(f"Haciendo stat de: {ruta}")
                st = sftp.stat(ruta)
                logger.info(f"Stat OK. Mode: {st.st_mode}, IsDir: {stat.S_ISDIR(st.st_mode)}")
                if not stat.S_ISDIR(st.st_mode):
                    ruta = posixpath.dirname(ruta) or "."
                    logger.info(f"Ruta ajustada (era archivo): {ruta}")
            except Exception as e:
                 logger.warning(f"Stat falló para {ruta}: {e}")
                 pass

            try:
                print(f"[DEBUG] Listando ruta: {ruta}")
                raw_list = sftp.listdir(ruta)
                print(f"[DEBUG] Raw listdir output ({len(raw_list)}): {raw_list}")
            except Exception as e:
                print(f"[DEBUG] Raw listdir falló: {e}")
            for attr in sftp.listdir_attr(ruta):
                print(f"[DEBUG] Encontrado: {attr.filename} (Dir: {stat.S_ISDIR(attr.st_mode)})")
                if not stat.S_ISDIR(attr.st_mode):
                    if attr.filename.lower().endswith(supported_exts):
                        print(f"[DEBUG] -> Aceptado: {attr.filename}")
                        files.append({
                            "nombre": attr.filename,
                            "fecha": datetime.fromtimestamp(attr.st_mtime).isoformat(),
                            "tamano": attr.st_size
                        })
                    else:
                        print(f"[DEBUG] -> Ignorado (extensión): {attr.filename}")
        finally:
            sftp.close()
            ssh.close()
    elif protocolo == "FTP":
        ftp = get_ftp_client(host, puerto, usuario, password)
        try:
            _ftp_enter_target_dir(ftp, ruta)

            seen_names = set()

            def _append_file(name: str, size: int = 0):
                clean_name = posixpath.basename((name or "").rstrip("/"))
                if not clean_name or clean_name in seen_names:
                    return
                if not clean_name.lower().endswith(supported_exts):
                    return
                seen_names.add(clean_name)
                files.append({
                    "nombre": clean_name,
                    "fecha": datetime.now().isoformat(), # FTP dates vary by server/listing format
                    "tamano": size
                })

            # Prefer MLSD when available (structured output and safe filenames).
            try:
                for name, facts in ftp.mlsd():
                    if name in (".", ".."):
                        continue
                    if (facts or {}).get("type") == "dir":
                        continue
                    file_size = int((facts or {}).get("size", 0) or 0)
                    _append_file(name, file_size)
            except Exception:
                pass

            # Fallback 1: NLST (usually returns complete names).
            if not files:
                try:
                    for name in ftp.nlst() or []:
                        _append_file(name, 0)
                except Exception:
                    pass

            # Fallback 2: LIST with robust parsing preserving spaces/commas.
            if not files:
                lines: List[str] = []
                ftp.retrlines("LIST", lines.append)
                for line in lines:
                    parsed = _parse_ftp_list_line(line)
                    if not parsed:
                        continue
                    name, is_dir = parsed
                    if is_dir:
                        continue
                    _append_file(name, 0)
        finally:
            ftp.quit()
    return sorted(files, key=lambda x: x["fecha"], reverse=True)

def _build_remote_listing_config(
    request_config: Dict[str, Any],
    stored_config: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    # Saved imports always use the database as the source of truth. Secrets are
    # intentionally absent from the browser payload, and a stale/default path
    # must not replace the route already configured for an automatic worker.
    source = dict(stored_config) if stored_config else dict(request_config)
    return _normalize_import_config_payload(source)


@app.post("/api/v1/remote/list-files")
async def list_files_endpoint(
    config: ImportConfigSchema,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    try:
        logger.info(f"Recibida solucitud de listado para: {config.host}:{config.puerto} (Protocolo: {config.protocolo})")
        request_config = config.dict()
        db_config = None
        if config.id:
            db_config = _load_local_config_with_access(config.id, operator_ctx)
        config_dict = _build_remote_listing_config(request_config, db_config)

        loop = asyncio.get_event_loop()
        files = await asyncio.wait_for(
            loop.run_in_executor(executor, _list_remote_files, config_dict),
            timeout=30.0
        )
        return files
    except asyncio.TimeoutError:
        logger.error(f"Timeout listando archivos para {config.nombre}")
        raise HTTPException(status_code=504, detail="El servidor remoto no respondió a tiempo (List Timeout)")
    except Exception as e:
        logger.error(f"Error listing files: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def _execute_manual_endpoint_impl(
    req: ExecuteManualRequest,
    operator_ctx: Optional[Dict[str, Any]] = None,
    exporter_ctx: Optional[TokenAuthContext] = None,
    internal_ctx: Optional[Dict[str, Any]] = None,
):
    logger.info(f"Ejecutando manual para {req.config_id} - Archivo: {req.filename}")
    execution_source = (
        "manual_remote_import_exporter"
        if exporter_ctx
        else "pending_import_monitor"
        if internal_ctx
        else "manual_remote_import"
    )
    batch_id = str(uuid4())
    request_id = (req.request_id or "").strip() if req.request_id else None
    request_cache_key = f"manual_exec:{request_id}" if request_id else None
    manual_exec_lock_acquired = False
    
    try:
        if request_cache_key:
            cached_result = _cache_get(request_cache_key)
            if cached_result is not _CACHE_MISS:
                logger.info(f"Devolviendo resultado cacheado para request_id={request_id}")
                return cached_result

            with _INFLIGHT_MANUAL_EXEC_LOCK:
                if request_id in _INFLIGHT_MANUAL_EXEC:
                    raise HTTPException(
                        status_code=409,
                        detail="La importación para esta solicitud ya está en proceso. Espera unos segundos e intenta de nuevo."
                    )
                _INFLIGHT_MANUAL_EXEC.add(request_id)
                manual_exec_lock_acquired = True

        def _cache_and_return(payload: Dict[str, Any]):
            if request_cache_key:
                _cache_set(request_cache_key, payload, ttl=1800)
            return payload

        # Source of truth: config del local desde DB + overrides permitidos del request.
        if exporter_ctx:
            config_data = _load_local_config_for_exporter(req.config_id, exporter_ctx)
        elif internal_ctx:
            if req.config is not None:
                raise HTTPException(
                    status_code=400,
                    detail="El monitor interno no admite sobrescribir la configuración del local.",
                )
            config_data = _load_local_config_for_pending_monitor(req.config_id)
        else:
            if not operator_ctx:
                raise HTTPException(status_code=401, detail="Se requiere autenticación para ejecutar importación manual")
            config_data = _load_local_config_with_access(req.config_id, operator_ctx)
        config_data = _apply_runtime_import_overrides(config_data, req.config)
        config_data = _normalize_import_config_payload(config_data)

        if not config_data.get("mall_id"):
            raise HTTPException(
                status_code=400,
                detail="La configuración del local no tiene mall_id asignado. No se puede importar de forma segura."
            )

        # Normalizar para _list_remote_files (y para esta lógica local)
        local_nombre = config_data.get("nombre") or "Desconocido"
        host = config_data.get("host") or config_data.get("sftp_host")
        puerto = config_data.get("puerto") or config_data.get("sftp_port", 22)
        usuario = config_data.get("usuario") or config_data.get("sftp_user")
        password = config_data.get("password") or config_data.get("sftp_pass")
        ruta_remota = config_data.get("ruta_remota") or config_data.get("sftp_path", ".")
        protocolo = config_data.get("protocolo") or config_data.get("sftp_protocol", "SFTP")
        source_filename = req.filename

        if _is_webservice_protocol(protocolo):
            result = await asyncio.to_thread(process_webservice_import, config_data, write_load_log=False)
            records_processed = int(result.get("records_processed") or 0)
            status_value = str(result.get("status") or ("success" if result.get("ok") else "error"))
            details = result.get("details") or []
            error_count = len(details)
            log_status = (
                "parcial"
                if status_value == "partial"
                or (error_count > 0 and records_processed > 0)
                else "exito" if status_value in {"success", "ok"} or result.get("ok")
                else "error"
            )
            log_channel = str(result.get("canal") or protocolo or "API").strip().upper()
            log_source_name = result.get("source_name") or "API ventas"
            insert_load_log(
                local_nombre,
                log_source_name,
                log_status,
                result.get("message") or f"{log_channel} ejecutado.",
                batch_id,
                details,
                mall_id=config_data.get("mall_id"),
                local_id=config_data.get("id"),
                canal=log_channel,
                records_processed=records_processed,
                error_count=error_count,
                metadata={
                    "source": "manual_api_webservice_import",
                    "worker_source": result.get("worker_source") or "worker_api_import",
                    "provider": result.get("provider") or api_provider_name(config_data),
                    "records_received": result.get("records_received"),
                    "duplicate_skipped": result.get("duplicate_skipped"),
                    "fallback_strategy": "daily" if result.get("failed_dates") else None,
                    "failed_dates": result.get("failed_dates") or [],
                    "error_type": result.get("error_type"),
                },
            )
            risk_snapshot = None
            if records_processed > 0:
                _reactivate_local_after_success(config_data, source=execution_source)
                risk_snapshot = await _run_local_risk_analysis_async(
                    config_data.get("id"),
                    trigger="manual_api_webservice_import",
                )
            return _cache_and_return({
                "status": status_value,
                "message": result.get("message") or "API ejecutada.",
                "records_processed": records_processed,
                "batch_id": batch_id,
                "errors": details,
                "renaming_error": None,
                "risk_summary": (risk_snapshot or {}).get("summary"),
            })

        def _build_prefixed_name(filename: str, prefix: str) -> str:
            base_name = posixpath.basename((filename or "").strip())
            if not base_name:
                return base_name
            upper_name = base_name.upper()
            if upper_name.startswith("PR_"):
                base_name = base_name[3:]
            elif upper_name.startswith("ERR_"):
                base_name = base_name[4:]
            return f"{prefix}{base_name}"

        def _rename_source_file(prefix: str) -> Optional[str]:
            nonlocal source_filename
            if not source_filename:
                return None

            if protocolo == "SFTP":
                ssh, sftp = get_sftp_client(host, puerto, usuario, password)
                try:
                    target_dir = ruta_remota
                    try:
                        st = sftp.stat(ruta_remota)
                        if not stat.S_ISDIR(st.st_mode):
                            target_dir = posixpath.dirname(ruta_remota) or "."
                    except Exception:
                        pass

                    old_path = posixpath.join(target_dir, source_filename)
                    new_name = _build_prefixed_name(source_filename, prefix)
                    new_path = posixpath.join(target_dir, new_name)
                    logger.info(f"Renombrando {old_path} -> {new_path}")
                    sftp.rename(old_path, new_path)
                    source_filename = new_name
                    return new_name
                finally:
                    try:
                        sftp.close()
                    except Exception:
                        pass
                    try:
                        ssh.close()
                    except Exception:
                        pass

            if protocolo == "FTP":
                ftp = get_ftp_client(host, puerto, usuario, password)
                try:
                    used_dir = _ftp_enter_target_dir(ftp, ruta_remota)
                    rename_source = _resolve_ftp_filename(ftp, source_filename, used_dir) or source_filename
                    new_name = _build_prefixed_name(rename_source, prefix)
                    logger.info(f"Renombrando {rename_source} -> {new_name}")
                    ftp.rename(rename_source, new_name)
                    source_filename = new_name
                    return new_name
                finally:
                    try:
                        ftp.quit()
                    except Exception:
                        pass

            return None
        
        # 2. Conectar y Descargar
        content = ""
        try:
            logger.info(f"Conectando a {host}:{puerto} via {protocolo} (User: {usuario})")
            if protocolo == "SFTP":
                ssh, sftp = get_sftp_client(host, puerto, usuario, password)
                try:
                    # Normalizar ruta: si es un archivo, usar el padre
                    target_dir = ruta_remota
                    try:
                        st = sftp.stat(ruta_remota)
                        if not stat.S_ISDIR(st.st_mode):
                            target_dir = posixpath.dirname(ruta_remota) or "."
                    except Exception as e:
                        logger.warning(f"No se pudo determinar si la ruta es archivo/directorio: {e}")

                    full_path = posixpath.join(target_dir, source_filename)
                    logger.info(f"Intentando abrir archivo SFTP: {full_path}")
                    with sftp.open(full_path, 'rb') as f:
                        raw_bytes = f.read()
                        content = _decode_remote_text(
                            raw_bytes,
                            is_json=source_filename.lower().endswith('.json')
                        )
                    
                    # Log file size and first chars for verification
                    logger.info(f"✅ Archivo leído: {full_path} | Tamaño: {len(content)} bytes | Primeros 100 chars: {content[:100]}")
                finally:
                    sftp.close()
                    ssh.close()
            elif protocolo == "FTP":
                ftp = get_ftp_client(host, puerto, usuario, password)
                try:
                    used_dir = _ftp_enter_target_dir(ftp, ruta_remota)
                    raw_bytes, resolved_name = _download_ftp_file_bytes(
                        ftp=ftp,
                        requested_filename=req.filename,
                        remote_path=used_dir,
                        max_bytes=None
                    )
                    source_filename = resolved_name or req.filename
                    content = _decode_remote_text(
                        raw_bytes,
                        is_json=source_filename.lower().endswith('.json')
                    )
                    
                    # Log file size and first chars for verification
                    logger.info(f"✅ Archivo FTP leído: {source_filename} | Tamaño: {len(content)} bytes | Primeros 100 chars: {content[:100]}")
                finally:
                    ftp.quit()
        except Exception as ce:
            error_msg = str(ce)
            logger.error(f"Error en ejecución manual ({protocolo} {host}): {error_msg}")
            insert_load_log(
                local_nombre,
                source_filename,
                "error",
                f"Error de conexión: {error_msg}",
                batch_id,
                mall_id=config_data.get("mall_id"),
                local_id=config_data.get("id"),
                canal=protocolo,
                records_processed=0,
                error_count=1,
                metadata={"source": execution_source, "connection_error": error_msg},
            )
            raise HTTPException(status_code=500, detail=f"Error de conexión remota ({protocolo}): {error_msg}")

        # 3. Procesar Contenido
        if not str(content or "").strip():
            detalles_errores = [{"linea": 0, "error": "Archivo remoto descargado vacío (0 bytes o sin texto)."}]
            insert_load_log(
                local_nombre,
                source_filename,
                "error",
                "Importación manual completada. 0 registros cargados. Se encontraron 1 errores de validación/mapeo.",
                batch_id,
                detalles_errores,
                mall_id=config_data.get("mall_id"),
                local_id=config_data.get("id"),
                canal=protocolo,
                records_processed=0,
                error_count=len(detalles_errores),
                metadata={"source": execution_source, "empty_payload": True},
            )
            renaming_error = None
            try:
                _rename_source_file("ERR_")
            except Exception as rename_err:
                renaming_error = str(rename_err)
                logger.error(f"Error renombrando archivo a ERR_: {rename_err}")
            return _cache_and_return({
                "status": "error",
                "message": "Importación manual completada. 0 registros cargados. Se encontraron 1 errores de validación/mapeo.",
                "records_processed": 0,
                "batch_id": batch_id,
                "errors": detalles_errores,
                "renaming_error": renaming_error
            })

        registros_exito, detalles_errores = process_file_content(
            content,
            source_filename,
            config_data,
            batch_id,
            config_data.get("mall_id")
        )

        no_data_detected = (registros_exito == 0 and not detalles_errores)
        if no_data_detected:
            estado = "no_encontrado"
            mensaje = "Importación manual completada. El archivo no contiene registros de data (vacío o solo encabezado)."
        else:
            estado = "exito" if registros_exito > 0 and not detalles_errores else "parcial" if registros_exito > 0 else "error"
            mensaje = f"Importación manual completada. {registros_exito} registros cargados."
            if detalles_errores:
                mensaje += f" Se encontraron {len(detalles_errores)} errores de validación/mapeo."

        # 4. Registrar Log en Monitor
        insert_load_log(
            local_nombre,
            source_filename,
            estado,
            mensaje,
            batch_id,
            detalles_errores,
            mall_id=config_data.get("mall_id"),
            local_id=config_data.get("id"),
            canal=protocolo,
            records_processed=registros_exito,
            error_count=len(detalles_errores or []),
            metadata={"source": execution_source},
        )

        risk_snapshot = None
        if registros_exito > 0:
            _reactivate_local_after_success(config_data, source=execution_source)
            risk_snapshot = await _run_local_risk_analysis_async(
                config_data.get("id"),
                trigger=execution_source,
            )

        # 5. Renombrar resultado: PR_ (éxito con registros) o ERR_ (fallo/no-data)
        logger.info(f"Evaluando renombrado: registros_exito={registros_exito} (Tipo: {type(registros_exito)})")
        target_prefix = "PR_" if (isinstance(registros_exito, int) and registros_exito > 0) else "ERR_"
        renaming_error = None
        try:
            _rename_source_file(target_prefix)
        except Exception as rename_err:
            renaming_error = str(rename_err)
            logger.error(f"Error al renombrar archivo pos-importación ({target_prefix}): {rename_err}")
            mensaje += f" (Advertencia: No se pudo renombrar el archivo: {rename_err})"

        final_status = "partial" if estado == "parcial" else ("success" if registros_exito > 0 else "error")
        return _cache_and_return({
            "status": final_status,
            "message": mensaje,
            "records_processed": registros_exito,
            "batch_id": batch_id,
            "errors": detalles_errores,
            "renaming_error": renaming_error,
            "risk_summary": (risk_snapshot or {}).get("summary"),
        })
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        import traceback
        traceback.print_exc()
        logger.error(f"Error en ejecución manual: {e}")
        if 'local_nombre' in locals():
            insert_load_log(
                local_nombre,
                source_filename,
                "error",
                str(e),
                batch_id,
                mall_id=config_data.get("mall_id") if 'config_data' in locals() else None,
                local_id=config_data.get("id") if 'config_data' in locals() else None,
                canal=protocolo if 'protocolo' in locals() else None,
                records_processed=0,
                error_count=1,
                metadata={"source": execution_source, "exception": str(e)},
            )
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if request_id and manual_exec_lock_acquired:
            with _INFLIGHT_MANUAL_EXEC_LOCK:
                _INFLIGHT_MANUAL_EXEC.discard(request_id)


@app.post("/api/v1/remote/execute-manual")
async def execute_manual_endpoint(
    req: ExecuteManualRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    return await _execute_manual_endpoint_impl(req=req, operator_ctx=operator_ctx)


@app.post("/api/v1/remote/execute-manual/exporter")
async def execute_manual_exporter_endpoint(
    req: ExecuteManualRequest,
    exporter_ctx: TokenAuthContext = Depends(
        require_token_auth("export:write", token_types={TOKEN_TYPE_EXPORTER})
    )
):
    return await _execute_manual_endpoint_impl(req=req, exporter_ctx=exporter_ctx)


@app.post("/api/v1/remote/execute-manual/internal")
async def execute_manual_internal_endpoint(
    req: ExecuteManualRequest,
    internal_ctx: Dict[str, Any] = Depends(require_pending_import_monitor_access),
):
    return await _execute_manual_endpoint_impl(req=req, internal_ctx=internal_ctx)


@app.post("/api/v1/analytics/cubo")
async def get_sales_cube(request: CubeRequest, mall_id: str = Depends(get_current_mall)):
    """
    Endpoint para generar el Cubo de Ventas (Matriz) usando datos reales de Supabase (Service Role).
    """
    try:
        custom_fields_service = _local_custom_fields_service()

        def _normalize_cube_totals_row(row: Dict[str, Any]) -> Dict[str, Any]:
            bruto = float(row.get("total_bruto") or 0)
            impuestos = float(row.get("total_impuestos") or 0) if row.get("total_impuestos") is not None else 0.0
            neto = float(row.get("total_neto") or 0)

            eps = 0.05
            as_is_delta = abs(neto - (bruto + impuestos))
            swapped_delta = abs(bruto - (neto + impuestos))
            if swapped_delta + eps < as_is_delta:
                row["total_bruto"] = neto
                row["total_neto"] = bruto
            return row

        # 1. Fetch Locales (Store Map) - Filtered by Mall
        stores_res = supabase.table("locales").select("id, nombre").eq("mall_id", mall_id).execute()
        stores = stores_res.data or []
        store_map = {str(s['id']): s['nombre'] for s in stores}
        allowed_local_ids = list(store_map.keys())

        if request.local_id:
            if str(request.local_id) not in store_map:
                # Prevent cross-tenant access and return deterministic empty matrix.
                return {"columns": ["local_nombre", "TOTAL_FILA"], "data": [], "grand_totals": {}}
            allowed_local_ids = [str(request.local_id)]

        snapshot = custom_fields_service.build_snapshot(mall_id, allowed_local_ids, include_inactive=False)
        if request.custom_filters:
            allowed_local_ids = custom_fields_service.filter_local_ids_by_custom_filters(
                allowed_local_ids,
                snapshot,
                request.custom_filters,
            )

        if not allowed_local_ids:
            return {"columns": ["local_nombre", "TOTAL_FILA"], "data": [], "grand_totals": {}}
        
        # 2. Fetch Sales within date range - Filtered by Mall
        # Note: Using service role key bypasses RLS
        # Important: filter by local_id list (derived from mall) instead of ventas.mall_id,
        # because some legacy rows may have null/incorrect mall_id while local_id is valid.
        # Supabase select has page limits; fetch all rows in batches.
        sales_data = []
        page_size = 1000
        page = 0
        while True:
            sales_res = (
                supabase.table("ventas")
                .select("*")
                .in_("local_id", allowed_local_ids)
                .gte("fecha", request.fecha_inicio)
                .lte("fecha", request.fecha_fin)
                .order("fecha")
                .range(page * page_size, (page + 1) * page_size - 1)
                .execute()
            )
            chunk = sales_res.data or []
            if not chunk:
                break
            sales_data.extend(_normalize_cube_totals_row(dict(row)) for row in chunk)
            if len(chunk) < page_size:
                break
            page += 1
        
        if not sales_data:
            # Return empty structure if no sales found
            return {"columns": ["local_nombre", "TOTAL_FILA"], "data": [], "grand_totals": {}}

        # 3. Convert to DataFrame
        df = pd.DataFrame(sales_data)
        
        # 4. Map Store Names
        # Ensure local_id is string for mapping
        df['local_id'] = df['local_id'].astype(str)
        df['local_nombre'] = df['local_id'].map(store_map).fillna("Desconocido (" + df['local_id'] + ")")
        
        # 5. Ensure numeric types for metrics
        df['total_bruto'] = pd.to_numeric(df['total_bruto'], errors='coerce').fillna(0)
        df['total_neto'] = pd.to_numeric(df['total_neto'], errors='coerce').fillna(0)
        df['transacciones'] = 1 # Each row is a transaction? Or aggregate? 
        # Assuming each row in 'ventas' is a transaction/daily summary. 
        # If 'ventas' is granular (ticket), count=1. If daily summary, we might need a 'count' column if exists, 
        # but usually 'ventas' tables in these systems are per-ticket or per-day. 
        # Looking at previous code, it seems granular or aggregated. Let's assume 1 row = 1 transaction for now if no other field.
        # Check if 'cantidad_transacciones' exists in DB? Previous view didn't show it.
        # Let's count rows as transactions.
        
        # 6. Generate Cube using existing logic or custom hierarchical grouping.
        result = custom_fields_service.build_cube_response(
            df,
            grouping=request.agrupacion,
            metric=request.metrica,
            start_date=request.fecha_inicio,
            end_date=request.fecha_fin,
            snapshot=snapshot,
            custom_dimension_key=request.custom_dimension_key,
        )
        return result
        
    except Exception as e:
        logger.error(f"Error generando cubo: {e}")
        # Return empty safe response instead of 500 to avoid breaking UI on minor data errors?
        # No, better to let UI know something went wrong, or return empty.
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v1/remote/analyze-file")
async def analyze_remote_file(
    req: ExecuteManualRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    """
    Analyzes a specific file from the remote server for mapping suggestions.
    Similar to analyze_remote_mapping but uses config+filename instead of direct path.
    """
    try:
        config_data = _load_local_config_with_access(req.config_id, operator_ctx)
        config_data = _apply_runtime_import_overrides(config_data, req.config)
        config_data = _normalize_import_config_payload(config_data)

        # Normalize configuration
        host = config_data.get("host") or config_data.get("sftp_host")
        puerto = config_data.get("puerto") or config_data.get("sftp_port", 22)
        usuario = config_data.get("usuario") or config_data.get("sftp_user")
        password = config_data.get("password") or config_data.get("sftp_pass")
        ruta_remota = config_data.get("ruta_remota") or config_data.get("sftp_path", ".")
        protocolo = config_data.get("protocolo") or config_data.get("sftp_protocol", "SFTP")
        analysis_filename = req.filename
        analysis_timeout_seconds = _remote_analysis_timeout_seconds(
            req.filename,
            config_data.get("tipo_archivo"),
        )
        
        loop = asyncio.get_event_loop()
        
        def _read_file_sample():
            nonlocal analysis_filename
            if protocolo == "SFTP":
                ssh, sftp = get_sftp_client(host, puerto, usuario, password)
                try:
                    # Normalize path
                    target_dir = ruta_remota
                    try:
                        st = sftp.stat(ruta_remota)
                        if not stat.S_ISDIR(st.st_mode):
                            target_dir = posixpath.dirname(ruta_remota) or "."
                    except:
                        pass
                    
                    full_path = posixpath.join(target_dir, req.filename)
                    with sftp.open(full_path, 'rb') as f:
                        analysis_filename = req.filename
                        # LEER TODO EL ARCHIVO SI ES JSON (necesario para parsear)
                        if analysis_filename.lower().endswith('.json'):
                            logger.info(f"Leyendo archivo COMPLETO (JSON): {analysis_filename}")
                            return _decode_remote_text(f.read(), is_json=True)
                        else:
                            return _decode_remote_text(f.read(65536), is_json=False)
                finally:
                    sftp.close()
                    ssh.close()
            elif protocolo == "FTP":
                ftp = get_ftp_client(host, puerto, usuario, password)
                try:
                    used_dir = _ftp_enter_target_dir(ftp, ruta_remota)
                    raw_bytes, resolved_name = _download_ftp_file_bytes(
                        ftp=ftp,
                        requested_filename=req.filename,
                        remote_path=used_dir,
                        max_bytes=65536
                    )
                    analysis_filename = resolved_name or req.filename
                    return _decode_remote_text(raw_bytes, is_json=False)
                finally:
                    ftp.quit()
            return ""

        content = await asyncio.wait_for(
            loop.run_in_executor(executor, _read_file_sample),
            timeout=analysis_timeout_seconds
        )
        
        if not content:
            return {"csv_headers": [], "suggested_mapping": {}, "sample_row": {}, "current_mapping": {}}
        
        forced_has_header, forced_data_start_row = _extract_parsing_options(config_data)
        analysis = _perform_mapping_analysis(
            content,
            analysis_filename,
            force_has_header=forced_has_header,
            data_start_row=forced_data_start_row
        )
        
        # Add current mapping from config if exists
        analysis["current_mapping"] = config_data.get("mapping", {})
        
        return analysis
        
    except asyncio.TimeoutError:
        raise HTTPException(
            status_code=504,
            detail="Timeout analizando archivo remoto (el archivo podria ser grande y la primera carga puede tardar varios minutos)",
        )
    except Exception as e:
        logger.error(f"Error analyzing remote file: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/remote/unmark-file")
async def unmark_file(
    req: ExecuteManualRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access)
):
    """
    Removes the 'PR_' prefix from a processed file to allow reprocessing.
    Renames 'PR_filename.ext' back to 'filename.ext'
    """
    try:
        config_data = _load_local_config_with_access(req.config_id, operator_ctx)
        config_data = _apply_runtime_import_overrides(config_data, req.config)
        config_data = _normalize_import_config_payload(config_data)

        # Normalize configuration
        host = config_data.get("host") or config_data.get("sftp_host")
        puerto = config_data.get("puerto") or config_data.get("sftp_port", 22)
        usuario = config_data.get("usuario") or config_data.get("sftp_user")
        password = config_data.get("password") or config_data.get("sftp_pass")
        ruta_remota = config_data.get("ruta_remota") or config_data.get("sftp_path", ".")
        protocolo = config_data.get("protocolo") or config_data.get("sftp_protocol", "SFTP")
        
        # Check if filename has PR_ prefix
        if not req.filename.startswith("PR_"):
            return {
                "status": "info",
                "message": f"El archivo '{req.filename}' no tiene el prefijo PR_, no requiere desmarcado."
            }
        
        # Calculate new name (remove PR_ prefix)
        new_filename = req.filename[3:]  # Remove first 3 characters "PR_"
        
        loop = asyncio.get_event_loop()
        
        def _rename_file():
            if protocolo == "SFTP":
                ssh, sftp = get_sftp_client(host, puerto, usuario, password)
                try:
                    # Determine directory
                    target_dir = ruta_remota
                    try:
                        st = sftp.stat(ruta_remota)
                        if not stat.S_ISDIR(st.st_mode):
                            target_dir = posixpath.dirname(ruta_remota) or "."
                    except:
                        pass
                    
                    old_path = posixpath.join(target_dir, req.filename)
                    new_path = posixpath.join(target_dir, new_filename)
                    
                    logger.info(f"Unmarking: {old_path} -> {new_path}")
                    sftp.rename(old_path, new_path)
                    return new_filename
                finally:
                    sftp.close()
                    ssh.close()
                    
            elif protocolo == "FTP":
                ftp = get_ftp_client(host, puerto, usuario, password)
                try:
                    used_dir = _ftp_enter_target_dir(ftp, ruta_remota)

                    resolved_old_name = _resolve_ftp_filename(ftp, req.filename, used_dir) or req.filename
                    resolved_new_name = resolved_old_name[3:] if resolved_old_name.startswith("PR_") else new_filename

                    logger.info(f"Unmarking: {resolved_old_name} -> {resolved_new_name}")
                    ftp.rename(resolved_old_name, resolved_new_name)
                    return resolved_new_name
                finally:
                    ftp.quit()
            
            return None
        
        result = await asyncio.wait_for(
            loop.run_in_executor(executor, _rename_file),
            timeout=30.0
        )
        
        if result:
            return {
                "status": "success",
                "message": f"Archivo desmarcado exitosamente",
                "old_name": req.filename,
                "new_name": result
            }
        else:
            raise HTTPException(status_code=500, detail="Error renombrando archivo")
            
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="Timeout conectando al servidor remoto")
    except Exception as e:
        logger.error(f"Error unmarking file: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _get_system_health_value(key: str) -> Optional[str]:
    if not supabase:
        return None
    try:
        rows = (
            supabase.table("system_health")
            .select("value")
            .eq("key", key)
            .order("last_update", desc=True)
            .limit(1)
            .execute()
        ).data or []
        row = rows[0] if rows else {}
        value = row.get("value")
        return str(value).strip() if value is not None else None
    except Exception as exc:
        logger.warning("No se pudo leer system_health[%s]: %s", key, exc)
        return None


def _upsert_system_health_value_sync(key: str, value: str) -> None:
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado.")
    timestamp = datetime.utcnow().isoformat()
    supabase.table("system_health").delete().eq("key", key).execute()
    supabase.table("system_health").insert({
        "key": key,
        "value": value,
        "last_update": timestamp,
    }).execute()


def _normalize_copilot_provider(value: Optional[str]) -> str:
    provider = str(value or "openai").strip().lower()
    if provider in {"chatgpt", "gpt", "open_ai"}:
        provider = "openai"
    if provider not in {"openai", "gemini"}:
        raise HTTPException(status_code=400, detail="Proveedor de Copilot invalido.")
    return provider


def _copilot_api_key_name(provider: str) -> str:
    return COPILOT_GEMINI_API_KEY_KEY if provider == "gemini" else COPILOT_OPENAI_API_KEY_KEY


def _copilot_model_key(provider: str) -> str:
    return COPILOT_GEMINI_MODEL_KEY if provider == "gemini" else COPILOT_OPENAI_MODEL_KEY


def _mask_secret(value: Optional[str]) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if len(raw) <= 4:
        return "****"
    return f"****{raw[-4:]}"


def _copilot_enabled_from_value(value: Optional[str]) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _copilot_config_status() -> Dict[str, Any]:
    provider = _normalize_copilot_provider(_get_system_health_value(COPILOT_PROVIDER_KEY) or "openai")
    model = (
        _get_system_health_value(_copilot_model_key(provider))
        or COPILOT_DEFAULT_MODELS[provider]
    )
    api_key = _get_system_health_value(_copilot_api_key_name(provider)) or ""
    enabled = _copilot_enabled_from_value(_get_system_health_value(COPILOT_ENABLED_KEY))
    return {
        "enabled": enabled,
        "provider": provider,
        "model": model,
        "api_key_configured": bool(api_key),
        "api_key_masked": _mask_secret(api_key),
        "available": enabled and bool(api_key),
    }


def _save_copilot_settings(payload: CopilotSettingsRequest) -> Dict[str, Any]:
    provider = _normalize_copilot_provider(payload.provider)
    model = str(payload.model or "").strip() or COPILOT_DEFAULT_MODELS[provider]
    api_key = None if payload.api_key is None else str(payload.api_key).strip()

    _upsert_system_health_value_sync(COPILOT_ENABLED_KEY, "true" if payload.enabled else "false")
    _upsert_system_health_value_sync(COPILOT_PROVIDER_KEY, provider)
    _upsert_system_health_value_sync(_copilot_model_key(provider), model)

    if payload.clear_api_key:
        _upsert_system_health_value_sync(_copilot_api_key_name(provider), "")
    elif api_key:
        _upsert_system_health_value_sync(_copilot_api_key_name(provider), api_key)

    return _copilot_config_status()


def _truncate_text(value: Any, limit: int = 260) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return f"{text[:limit - 3]}..."


def _safe_date_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _safe_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _compact_copilot_log(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "fecha_hora": _safe_date_text(row.get("fecha_hora")),
        "local": row.get("local_nombre"),
        "estado": row.get("estado"),
        "archivo": row.get("archivo"),
        "mensaje": _truncate_text(row.get("mensaje"), 220),
        "records_processed": row.get("records_processed"),
        "error_count": row.get("error_count"),
        "canal": row.get("canal"),
    }


def _load_copilot_locales(mall_id: str) -> List[Dict[str, Any]]:
    preferred_columns = (
        "id,nombre,codigo_interno,email,rubro,tipo_negocio,processing_status,"
        "consecutive_failures,upsert_activo,ultima_ejecucion,resultado_ultimo,"
        "sftp_protocol,sftp_host,frecuencia_cron,hora_especifica"
    )
    fallback_columns = "id,nombre,codigo_interno,rubro,tipo_negocio,mall_id"
    try:
        response = (
            supabase.table("locales")
            .select(preferred_columns)
            .eq("mall_id", mall_id)
            .order("nombre")
            .limit(80)
            .execute()
        )
        return response.data or []
    except Exception as exc:
        logger.warning("Copilot locales preferred query failed: %s", sanitize_sensitive_ops_error(exc))
        response = (
            supabase.table("locales")
            .select(fallback_columns)
            .eq("mall_id", mall_id)
            .order("nombre")
            .limit(80)
            .execute()
        )
        return response.data or []


def _load_copilot_missing_days(locales: List[Dict[str, Any]], lookback_days: int = 7) -> Dict[str, Any]:
    local_ids = [str(row.get("id")) for row in locales if row.get("id")]
    end_date = datetime.utcnow().date()
    start_date = end_date - timedelta(days=max(1, lookback_days) - 1)
    expected_dates = {
        (start_date + timedelta(days=offset)).isoformat()
        for offset in range((end_date - start_date).days + 1)
    }

    if not local_ids:
        return {
            "lookback_days": lookback_days,
            "fecha_inicio": start_date.isoformat(),
            "fecha_fin": end_date.isoformat(),
            "status": "sin_locales",
            "locales_con_brechas": 0,
            "locales_completos": 0,
            "top_brechas": [],
        }

    dates_by_local: Dict[str, Set[str]] = {local_id: set() for local_id in local_ids}
    page_size = 5000
    max_pages = 4
    try:
        for page in range(max_pages):
            chunk = (
                supabase.table("ventas")
                .select("local_id,fecha")
                .in_("local_id", local_ids)
                .gte("fecha", start_date.isoformat())
                .lte("fecha", end_date.isoformat())
                .range(page * page_size, (page + 1) * page_size - 1)
                .execute()
            ).data or []
            for row in chunk:
                local_id = str(row.get("local_id") or "")
                normalized = _normalize_missing_days_sale_date(row.get("fecha"))
                if local_id and normalized:
                    dates_by_local.setdefault(local_id, set()).add(normalized)
            if len(chunk) < page_size:
                break
    except Exception as exc:
        logger.warning("Copilot missing days query failed: %s", sanitize_sensitive_ops_error(exc))
        return {
            "lookback_days": lookback_days,
            "fecha_inicio": start_date.isoformat(),
            "fecha_fin": end_date.isoformat(),
            "status": "no_disponible",
            "error": "No se pudo consultar ventas para dias de informacion.",
        }

    rows = []
    for local in locales:
        local_id = str(local.get("id") or "")
        actual = dates_by_local.get(local_id, set())
        missing = sorted(expected_dates - actual)
        rows.append({
            "local_id": local_id,
            "local": local.get("nombre"),
            "dias_con_informacion": len(actual & expected_dates),
            "dias_faltantes": len(missing),
            "fechas_faltantes": missing[:10],
        })

    with_gaps = [row for row in rows if row["dias_faltantes"] > 0]
    return {
        "lookback_days": lookback_days,
        "fecha_inicio": start_date.isoformat(),
        "fecha_fin": end_date.isoformat(),
        "status": "ok",
        "locales_con_brechas": len(with_gaps),
        "locales_completos": len(rows) - len(with_gaps),
        "dias_esperados_por_local": len(expected_dates),
        "top_brechas": sorted(with_gaps, key=lambda item: item["dias_faltantes"], reverse=True)[:15],
    }


def _load_copilot_sales_summary(locales: List[Dict[str, Any]], lookback_days: int = 30) -> Dict[str, Any]:
    local_ids = [str(row.get("id")) for row in locales if row.get("id")]
    end_date = datetime.utcnow().date()
    start_date = end_date - timedelta(days=max(1, lookback_days) - 1)

    if not local_ids:
        return {
            "lookback_days": lookback_days,
            "fecha_inicio": start_date.isoformat(),
            "fecha_fin": end_date.isoformat(),
            "status": "sin_locales",
            "ventas_totales_bruto": 0,
            "ventas_totales_neto": 0,
            "transacciones": 0,
            "ticket_promedio": 0,
            "top_locales": [],
            "ventas_por_dia": [],
        }

    store_map = {str(row.get("id")): row for row in locales if row.get("id")}
    sales: List[Dict[str, Any]] = []
    page_size = 5000
    max_pages = 8

    try:
        for page in range(max_pages):
            chunk = (
                supabase.table("ventas")
                .select("local_id,fecha,total_bruto,total_neto")
                .in_("local_id", local_ids)
                .gte("fecha", start_date.isoformat())
                .lte("fecha", end_date.isoformat())
                .order("fecha")
                .range(page * page_size, (page + 1) * page_size - 1)
                .execute()
            ).data or []
            sales.extend(chunk)
            if len(chunk) < page_size:
                break
    except Exception as exc:
        logger.warning("Copilot sales query failed: %s", sanitize_sensitive_ops_error(exc))
        return {
            "lookback_days": lookback_days,
            "fecha_inicio": start_date.isoformat(),
            "fecha_fin": end_date.isoformat(),
            "status": "no_disponible",
            "error": "No se pudo consultar ventas recientes.",
        }

    total_bruto = 0.0
    total_neto = 0.0
    by_store: Dict[str, Dict[str, Any]] = {}
    by_day: Dict[str, float] = {}
    by_business_type: Dict[str, float] = {}
    by_rubro: Dict[str, float] = {}

    for sale in sales:
        local_id = str(sale.get("local_id") or "")
        store = store_map.get(local_id) or {}
        store_name = store.get("nombre") or "Local sin nombre"
        business_type = str(store.get("tipo_negocio") or "Sin tipo de negocio").strip()
        rubro = str(store.get("rubro") or "Sin rubro").strip()
        sale_date = _normalize_missing_days_sale_date(sale.get("fecha")) or str(sale.get("fecha") or "")
        bruto = _safe_float(sale.get("total_bruto"))
        neto = _safe_float(sale.get("total_neto"))

        total_bruto += bruto
        total_neto += neto
        by_day[sale_date] = by_day.get(sale_date, 0.0) + bruto
        by_business_type[business_type] = by_business_type.get(business_type, 0.0) + bruto
        by_rubro[rubro] = by_rubro.get(rubro, 0.0) + bruto

        store_totals = by_store.setdefault(store_name, {
            "local": store_name,
            "total_bruto": 0.0,
            "total_neto": 0.0,
            "transacciones": 0,
        })
        store_totals["total_bruto"] += bruto
        store_totals["total_neto"] += neto
        store_totals["transacciones"] += 1

    transactions = len(sales)
    stores_ranked = sorted(by_store.values(), key=lambda item: item["total_bruto"], reverse=True)
    return {
        "lookback_days": lookback_days,
        "fecha_inicio": start_date.isoformat(),
        "fecha_fin": end_date.isoformat(),
        "status": "ok",
        "ventas_totales_bruto": total_bruto,
        "ventas_totales_neto": total_neto,
        "transacciones": transactions,
        "ticket_promedio": (total_bruto / transactions) if transactions else 0,
        "top_locales": stores_ranked[:10],
        "locales": stores_ranked[:80],
        "ventas_por_dia": [
            {"fecha": sale_date, "total_bruto": total}
            for sale_date, total in sorted(by_day.items())[-30:]
        ],
        "ventas_por_tipo_negocio": [
            {"tipo_negocio": label, "total_bruto": total}
            for label, total in sorted(by_business_type.items(), key=lambda item: item[1], reverse=True)[:10]
        ],
        "ventas_por_rubro": [
            {"rubro": label, "total_bruto": total}
            for label, total in sorted(by_rubro.items(), key=lambda item: item[1], reverse=True)[:10]
        ],
    }


def _is_big_data_copilot_question(message: str) -> bool:
    normalized = _normalize_store_catalog_key(message)
    terms = (
        "proyeccion",
        "cierre",
        "como van las ventas",
        "ventas este mes",
        "ventas del mes",
        "anomalia",
        "categoria",
        "creciendo",
        "caida",
        "desempeno del mall",
        "resumen ejecutivo",
        "no han reportado",
        "periodo incompleto",
        "comparar local",
    )
    return any(term in normalized for term in terms)


def _require_big_data_copilot_flags(mall_id: str) -> None:
    for feature in ("BIG_DATA_CORE", "BIG_DATA_COPILOT"):
        enabled = supabase.rpc(
            "is_mall_feature_enabled",
            {"requested_mall_id": mall_id, "requested_feature": feature},
        ).execute().data
        if enabled is not True:
            raise HTTPException(
                status_code=403,
                detail=f"{feature} no está activado para este mall.",
            )


def _build_big_data_copilot_context(
    mall_id: str, operator_ctx: Dict[str, Any]
) -> Dict[str, Any]:
    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    _require_big_data_copilot_flags(mall_id)
    today = datetime.utcnow().date()
    month_start = today.replace(day=1)
    summary = BigDataSprint2Service(supabase).executive_summary(
        mall_id, month_start, today
    )
    context = {
        "generated_at_utc": datetime.utcnow().isoformat(),
        "mall": {"id": mall_id},
        "periodo_analizado": {
            "inicio": month_start.isoformat(),
            "fin": today.isoformat(),
        },
        "big_data": summary,
        "restrictions": {
            "facts_calculations_inferences_separated": True,
            "projection_is_estimate": True,
            "causality_not_established": True,
            "source": "Sprint 1 aggregate tables",
        },
    }
    supabase.table("operations_events").insert(
        {
            "mall_id": mall_id,
            "event_type": "COPILOT_BIG_DATA_QUERY",
            "source": "COPILOT",
            "severity": "INFO",
            "processing_status": "PENDING",
            "payload": {
                "operator_id": operator_ctx.get("user_id"),
                "period": context["periodo_analizado"],
                "context_version": "big-data-copilot-v1",
            },
        }
    ).execute()
    return context


def _deterministic_big_data_answer(context: Dict[str, Any]) -> str:
    summary = context.get("big_data") or {}
    forecast = summary.get("forecast") or {}
    coverage = float(summary.get("coverage") or 0)
    if forecast.get("status") == "INSUFFICIENT_DATA":
        projection = "No hay información suficiente para calcular una proyección confiable."
    else:
        projection = (
            f"La proyección estimada de cierre es {float(forecast.get('expected_close') or 0):,.2f}, "
            f"con rango {float(forecast.get('lower_bound') or 0):,.2f}–"
            f"{float(forecast.get('upper_bound') or 0):,.2f} y confianza "
            f"{forecast.get('confidence', 'LOW')}."
        )
    quality = (
        "El período está incompleto; no se atribuyen causas comerciales."
        if summary.get("general_status") == "DATA_INCOMPLETE"
        else "La cobertura permite describir los hechos observados."
    )
    return (
        f"**Resumen del período**\n"
        f"- Hecho: venta acumulada {float(summary.get('accumulated_sales') or 0):,.2f}.\n"
        f"- Cobertura: {coverage:.1f}%. {quality}\n"
        f"- Estimación: {projection}\n"
        f"- Hallazgos activos: {len(summary.get('anomalies') or [])}.\n"
        f"- Las cifras provienen de agregados controlados; no se infiere causalidad."
    )


def _build_copilot_context(mall_id: str, operator_ctx: Dict[str, Any]) -> Dict[str, Any]:
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado.")
    _ensure_operator_can_access_mall(operator_ctx, mall_id)

    try:
        mall_row = (
            supabase.table("malls")
            .select("id,nombre")
            .eq("id", mall_id)
            .maybe_single()
            .execute()
        ).data or {}
    except Exception:
        mall_row = {"id": mall_id, "nombre": "Mall seleccionado"}

    locales = _load_copilot_locales(mall_id)
    compact_locales = [
        {
            "id": row.get("id"),
            "codigo": row.get("codigo_interno"),
            "nombre": row.get("nombre"),
            "email": row.get("email"),
            "rubro": row.get("rubro"),
            "tipo_negocio": row.get("tipo_negocio"),
            "processing_status": row.get("processing_status"),
            "consecutive_failures": row.get("consecutive_failures"),
            "upsert_activo": row.get("upsert_activo"),
            "ultima_ejecucion": _safe_date_text(row.get("ultima_ejecucion")),
            "resultado_ultimo": row.get("resultado_ultimo"),
            "protocolo": row.get("sftp_protocol"),
            "frecuencia": row.get("frecuencia_cron"),
            "hora_especifica": row.get("hora_especifica"),
        }
        for row in locales[:60]
    ]

    try:
        logs = (
            supabase.table("logs_carga")
            .select("*")
            .eq("mall_id", mall_id)
            .order("fecha_hora", desc=True)
            .limit(40)
            .execute()
        ).data or []
    except Exception as exc:
        logger.warning("Copilot load logs query failed: %s", sanitize_sensitive_ops_error(exc))
        logs = []

    status_counts: Dict[str, int] = {}
    for log in logs:
        status_key = str(log.get("estado") or "desconocido").strip().lower() or "desconocido"
        status_counts[status_key] = status_counts.get(status_key, 0) + 1

    try:
        connection_monitor = _connection_monitor_service().get_status_summary(
            mall_id=mall_id,
            operator_ctx=operator_ctx,
            ensure_operator_can_access_mall=_ensure_operator_can_access_mall,
        )
    except Exception as exc:
        logger.warning("Copilot connection monitor query failed: %s", sanitize_sensitive_ops_error(exc))
        connection_monitor = {"status": "no_disponible"}

    missing_days = _load_copilot_missing_days(locales)
    sales_summary = _load_copilot_sales_summary(locales)

    return {
        "generated_at_utc": datetime.utcnow().isoformat(),
        "mall": {
            "id": mall_id,
            "nombre": mall_row.get("nombre") or "Mall seleccionado",
        },
        "locales": {
            "total": len(locales),
            "con_importacion_activa": sum(1 for row in locales if row.get("upsert_activo") or row.get("sftp_host")),
            "en_proceso": sum(1 for row in locales if row.get("processing_status") == "BUSY"),
            "con_fallas_consecutivas": sum(1 for row in locales if int(row.get("consecutive_failures") or 0) > 0),
            "muestra": compact_locales,
        },
        "monitor_carga": {
            "total_logs_recientes": len(logs),
            "conteo_por_estado": status_counts,
            "logs_recientes": [_compact_copilot_log(row) for row in logs[:15]],
        },
        "monitor_conexiones": connection_monitor,
        "dias_informacion": missing_days,
        "ventas_recientes": sales_summary,
    }


def _normalize_copilot_report_format(message: str) -> Optional[str]:
    text = _normalize_store_catalog_key(message)
    if re.search(r"\b(pdf)\b", text):
        return "pdf"
    if re.search(r"\b(excel|xlsx|xls|hoja de calculo)\b", text):
        return "xlsx"
    return None


def _normalize_copilot_report_type(message: str) -> Optional[str]:
    text = _normalize_store_catalog_key(message)
    if any(term in text for term in ["venta", "sales", "facturacion", "ingreso"]):
        return "ventas"
    if any(term in text for term in ["faltante", "dias", "brecha", "informacion"]):
        return "dias_faltantes"
    if any(term in text for term in ["monitor", "carga", "log", "importacion", "error"]):
        return "monitor_carga"
    if any(term in text for term in ["local", "tienda", "listado", "establecimiento"]):
        return "locales"
    return "locales"


def _parse_copilot_report_request(message: str) -> Optional[Dict[str, str]]:
    report_format = _normalize_copilot_report_format(message)
    if not report_format:
        return None
    return {
        "format": report_format,
        "type": _normalize_copilot_report_type(message) or "locales",
    }


def _parse_copilot_email_request(message: str) -> Optional[Dict[str, Any]]:
    text = _normalize_store_catalog_key(message)
    if not any(term in text for term in ["correo", "email", "mail", "enviar", "mandar", "envialo", "enviarlo", "enviame"]):
        return None

    recipients = _normalize_email_list(re.findall(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", message), strict=False)
    wants_pdf = "pdf" in text
    wants_xlsx = any(term in text for term in ["excel", "xlsx", "xls", "hoja de calculo"])
    wants_html = "html" in text or any(term in text for term in ["cuerpo", "resumen", "correo"])
    if not wants_pdf and not wants_xlsx and not wants_html:
        wants_html = True
        wants_xlsx = True

    attachment_format = "pdf" if wants_pdf else "xlsx" if wants_xlsx else None
    return {
        "recipients": recipients,
        "report_type": _normalize_copilot_report_type(message) or "locales",
        "include_html": True,
        "attachment_format": attachment_format,
    }


def _report_value(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 2)
    return value if value is not None else ""


def _copilot_report_definition(report_type: str, context: Dict[str, Any]) -> Dict[str, Any]:
    mall_name = (context.get("mall") or {}).get("nombre") or "MsMall"
    generated_at = context.get("generated_at_utc") or datetime.utcnow().isoformat()

    if report_type == "ventas":
        sales = context.get("ventas_recientes") or {}
        rows = sales.get("locales") or sales.get("top_locales") or []
        return {
            "title": "Ventas recientes por local",
            "subtitle": f"{mall_name} | {sales.get('fecha_inicio', '')} al {sales.get('fecha_fin', '')}",
            "filename_base": "msmall_ventas_recientes",
            "sources": ["ventas_recientes"],
            "headers": ["Local", "Total Bruto", "Total Neto", "Transacciones", "Ticket Promedio"],
            "rows": [
                [
                    row.get("local") or row.get("name"),
                    _report_value(row.get("total_bruto") or row.get("total") or 0),
                    _report_value(row.get("total_neto") or 0),
                    row.get("transacciones") or 0,
                    _report_value((row.get("total_bruto") or row.get("total") or 0) / max(1, int(row.get("transacciones") or 0))),
                ]
                for row in rows
            ],
            "summary": [
                ["Total Bruto", _report_value(sales.get("ventas_totales_bruto") or 0)],
                ["Total Neto", _report_value(sales.get("ventas_totales_neto") or 0)],
                ["Transacciones", sales.get("transacciones") or 0],
                ["Ticket Promedio", _report_value(sales.get("ticket_promedio") or 0)],
            ],
            "generated_at": generated_at,
        }

    if report_type == "dias_faltantes":
        missing = context.get("dias_informacion") or {}
        rows = missing.get("top_brechas") or []
        return {
            "title": "Dias faltantes por local",
            "subtitle": f"{mall_name} | {missing.get('fecha_inicio', '')} al {missing.get('fecha_fin', '')}",
            "filename_base": "msmall_dias_faltantes",
            "sources": ["dias_informacion"],
            "headers": ["Local", "Dias con informacion", "Dias faltantes", "Fechas faltantes"],
            "rows": [
                [
                    row.get("local"),
                    row.get("dias_con_informacion") or 0,
                    row.get("dias_faltantes") or 0,
                    ", ".join(row.get("fechas_faltantes") or []),
                ]
                for row in rows
            ],
            "summary": [
                ["Locales con brechas", missing.get("locales_con_brechas") or 0],
                ["Locales completos", missing.get("locales_completos") or 0],
                ["Dias esperados por local", missing.get("dias_esperados_por_local") or 0],
            ],
            "generated_at": generated_at,
        }

    if report_type == "monitor_carga":
        monitor = context.get("monitor_carga") or {}
        rows = monitor.get("logs_recientes") or []
        return {
            "title": "Monitor de carga reciente",
            "subtitle": f"{mall_name} | ultimos {len(rows)} registros",
            "filename_base": "msmall_monitor_carga",
            "sources": ["monitor_carga"],
            "headers": ["Fecha/Hora", "Local", "Estado", "Canal", "Archivo", "Procesados", "Errores", "Mensaje"],
            "rows": [
                [
                    row.get("fecha_hora"),
                    row.get("local"),
                    row.get("estado"),
                    row.get("canal"),
                    row.get("archivo"),
                    row.get("records_processed") or 0,
                    row.get("error_count") or 0,
                    row.get("mensaje"),
                ]
                for row in rows
            ],
            "summary": [[key, value] for key, value in (monitor.get("conteo_por_estado") or {}).items()],
            "generated_at": generated_at,
        }

    locales = (context.get("locales") or {}).get("muestra") or []
    return {
        "title": "Listado de locales",
        "subtitle": f"{mall_name} | {len(locales)} locales incluidos",
        "filename_base": "msmall_listado_locales",
        "sources": ["locales"],
        "headers": ["Codigo", "Nombre", "Email", "Rubro", "Tipo de Negocio", "Estado Proceso", "Fallas", "Importacion Activa"],
        "rows": [
            [
                row.get("codigo"),
                row.get("nombre"),
                row.get("email"),
                row.get("rubro"),
                row.get("tipo_negocio"),
                row.get("processing_status"),
                row.get("consecutive_failures") or 0,
                "Si" if row.get("upsert_activo") else "No",
            ]
            for row in locales
        ],
        "summary": [
            ["Total locales", (context.get("locales") or {}).get("total") or len(locales)],
            ["Con importacion activa", (context.get("locales") or {}).get("con_importacion_activa") or 0],
            ["En proceso", (context.get("locales") or {}).get("en_proceso") or 0],
            ["Con fallas consecutivas", (context.get("locales") or {}).get("con_fallas_consecutivas") or 0],
        ],
        "generated_at": generated_at,
    }


def _build_copilot_excel(definition: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte"

    title_font = Font(size=14, bold=True, color="111827")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")

    sheet.append([definition["title"]])
    sheet["A1"].font = title_font
    sheet.append([definition.get("subtitle") or ""])
    sheet.append(["Generado", definition.get("generated_at") or datetime.utcnow().isoformat()])
    sheet.append([])

    summary = definition.get("summary") or []
    if summary:
        sheet.append(["Resumen"])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
        for label, value in summary:
            sheet.append([label, _report_value(value)])
        sheet.append([])

    headers = definition.get("headers") or []
    sheet.append(headers)
    header_row = sheet.max_row
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(row=header_row, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in definition.get("rows") or []:
        sheet.append([_report_value(value) for value in row])

    for column_cells in sheet.columns:
        max_length = 10
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 48)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_copilot_pdf(definition: Dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(definition["title"], styles["Title"]),
        Paragraph(definition.get("subtitle") or "", styles["Normal"]),
        Paragraph(f"Generado: {definition.get('generated_at') or datetime.utcnow().isoformat()}", styles["Normal"]),
        Spacer(1, 12),
    ]

    summary = definition.get("summary") or []
    if summary:
        story.append(Paragraph("Resumen", styles["Heading3"]))
        summary_table = Table([["Indicador", "Valor"], *[[_truncate_text(label, 60), _report_value(value)] for label, value in summary]])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.extend([summary_table, Spacer(1, 12)])

    headers = definition.get("headers") or []
    rows = definition.get("rows") or []
    table_rows = [headers] + [
        [_truncate_text(_report_value(value), 72) for value in row]
        for row in rows[:80]
    ]
    if len(table_rows) == 1:
        table_rows.append(["Sin datos disponibles"] + [""] * max(0, len(headers) - 1))

    table = Table(table_rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)
    doc.build(story)
    return output.getvalue()


def _build_copilot_report_html(definition: Dict[str, Any]) -> str:
    summary_rows = "".join(
        f"<tr><td>{html.escape(str(label))}</td><td><strong>{html.escape(str(_report_value(value)))}</strong></td></tr>"
        for label, value in (definition.get("summary") or [])
    )
    headers = definition.get("headers") or []
    body_rows = ""
    for row in (definition.get("rows") or [])[:80]:
        cells = "".join(f"<td>{html.escape(str(_report_value(value)))}</td>" for value in row)
        body_rows += f"<tr>{cells}</tr>"
    if not body_rows:
        body_rows = f"<tr><td colspan=\"{max(1, len(headers))}\">Sin datos disponibles.</td></tr>"
    header_cells = "".join(f"<th>{html.escape(str(header))}</th>" for header in headers)
    return f"""
    <div style="font-family:Arial,sans-serif;color:#0f172a;background:#f8fafc;padding:24px">
      <div style="max-width:920px;margin:auto;background:white;border:1px solid #e2e8f0;border-radius:18px;overflow:hidden">
        <div style="background:#111827;color:white;padding:20px 24px">
          <h1 style="margin:0;font-size:20px">{html.escape(str(definition.get('title') or 'Reporte MsMall'))}</h1>
          <p style="margin:6px 0 0;color:#cbd5e1">{html.escape(str(definition.get('subtitle') or ''))}</p>
        </div>
        <div style="padding:22px 24px">
          <p style="margin:0 0 16px;color:#64748b;font-size:13px">Generado: {html.escape(str(definition.get('generated_at') or datetime.utcnow().isoformat()))}</p>
          {f'<h2 style="font-size:15px">Resumen</h2><table style="width:100%;border-collapse:collapse;margin-bottom:20px">{summary_rows}</table>' if summary_rows else ''}
          <h2 style="font-size:15px">Detalle</h2>
          <div style="overflow-x:auto">
            <table style="width:100%;border-collapse:collapse;font-size:13px">
              <thead><tr style="background:#f1f5f9">{header_cells}</tr></thead>
              <tbody>{body_rows}</tbody>
            </table>
          </div>
          <p style="margin-top:20px;color:#94a3b8;font-size:12px">Enviado por Copilot MsMall. Fuente: {html.escape(', '.join(definition.get('sources') or []))}</p>
        </div>
      </div>
    </div>
    <style>
      th, td {{ border-bottom:1px solid #e2e8f0; padding:10px 12px; text-align:left; vertical-align:top; }}
      th {{ color:#475569; font-size:11px; text-transform:uppercase; letter-spacing:.04em; }}
    </style>
    """


def _cleanup_copilot_downloads(now: Optional[float] = None) -> None:
    current_time = now or time.time()
    expired = [
        key for key, item in _COPILOT_DOWNLOADS.items()
        if float(item.get("expires_at_epoch") or 0) <= current_time
    ]
    for key in expired:
        _COPILOT_DOWNLOADS.pop(key, None)


def _store_copilot_download(filename: str, mime_type: str, content: bytes) -> Dict[str, Any]:
    download_id = secrets.token_urlsafe(24)
    expires_at_epoch = time.time() + _COPILOT_DOWNLOAD_TTL_SECONDS
    safe_filename = re.sub(r"[^A-Za-z0-9_.-]+", "_", filename).strip("_") or "msmall_reporte"
    with _COPILOT_DOWNLOADS_LOCK:
        _cleanup_copilot_downloads()
        _COPILOT_DOWNLOADS[download_id] = {
            "filename": safe_filename,
            "mime_type": mime_type,
            "content": content,
            "expires_at_epoch": expires_at_epoch,
        }
    return {
        "id": download_id,
        "filename": safe_filename,
        "mime_type": mime_type,
        "download_url": f"/api/v1/copilot/download/{download_id}",
        "expires_at": datetime.utcfromtimestamp(expires_at_epoch).isoformat(),
    }


def _store_copilot_email_draft(draft: Dict[str, Any]) -> Dict[str, Any]:
    draft_id = secrets.token_urlsafe(24)
    expires_at_epoch = time.time() + _COPILOT_DOWNLOAD_TTL_SECONDS
    with _COPILOT_EMAIL_DRAFTS_LOCK:
        _cleanup_copilot_email_drafts()
        _COPILOT_EMAIL_DRAFTS[draft_id] = {
            **draft,
            "expires_at_epoch": expires_at_epoch,
        }
    return {
        "id": draft_id,
        "expires_at": datetime.utcfromtimestamp(expires_at_epoch).isoformat(),
    }


def _cleanup_copilot_email_drafts(now: Optional[float] = None) -> None:
    current_time = now or time.time()
    expired = [
        key for key, item in _COPILOT_EMAIL_DRAFTS.items()
        if float(item.get("expires_at_epoch") or 0) <= current_time
    ]
    for key in expired:
        _COPILOT_EMAIL_DRAFTS.pop(key, None)


def _build_copilot_email_draft(email_request: Dict[str, Any], context: Dict[str, Any], mall_id: str, operator_ctx: Dict[str, Any]) -> Dict[str, Any]:
    recipients = email_request.get("recipients") or []
    if not recipients:
        raise HTTPException(status_code=400, detail="Indica al menos un correo destinatario para enviar el reporte.")

    definition = _copilot_report_definition(email_request["report_type"], context)
    html_body = _build_copilot_report_html(definition)
    subject = f"{definition['title']} - MsMall"
    attachments = []
    attachment_format = email_request.get("attachment_format")
    if attachment_format:
        generated_date = datetime.utcnow().strftime("%Y%m%d_%H%M")
        if attachment_format == "pdf":
            content = _build_copilot_pdf(definition)
            filename = f"{definition['filename_base']}_{generated_date}.pdf"
            mime_type = "application/pdf"
        else:
            content = _build_copilot_excel(definition)
            filename = f"{definition['filename_base']}_{generated_date}.xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        attachments.append({"filename": filename, "content": content, "mime_type": mime_type})

    stored = _store_copilot_email_draft({
        "mall_id": mall_id,
        "recipients": recipients,
        "subject": subject,
        "html_body": html_body,
        "text": f"{definition['title']}\n{definition.get('subtitle') or ''}",
        "attachments": attachments,
        "report_type": email_request["report_type"],
        "row_count": len(definition.get("rows") or []),
        "sources": definition.get("sources") or [],
        "created_by": operator_ctx.get("email") or operator_ctx.get("user_id"),
    })
    return {
        **stored,
        "recipients": recipients,
        "subject": subject,
        "report_type": email_request["report_type"],
        "row_count": len(definition.get("rows") or []),
        "attachment_count": len(attachments),
        "sources": definition.get("sources") or [],
    }


def _generate_copilot_report_attachment(report_request: Dict[str, str], context: Dict[str, Any]) -> Dict[str, Any]:
    definition = _copilot_report_definition(report_request["type"], context)
    generated_date = datetime.utcnow().strftime("%Y%m%d_%H%M")
    if report_request["format"] == "pdf":
        filename = f"{definition['filename_base']}_{generated_date}.pdf"
        content = _build_copilot_pdf(definition)
        mime_type = "application/pdf"
    else:
        filename = f"{definition['filename_base']}_{generated_date}.xlsx"
        content = _build_copilot_excel(definition)
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    attachment = _store_copilot_download(filename, mime_type, content)
    attachment.update({
        "label": definition["title"],
        "report_type": report_request["type"],
        "format": report_request["format"],
        "sources": definition.get("sources") or [],
        "row_count": len(definition.get("rows") or []),
    })
    return attachment


@app.get("/api/v1/copilot/download/{download_id}")
async def download_copilot_report(download_id: str):
    with _COPILOT_DOWNLOADS_LOCK:
        _cleanup_copilot_downloads()
        item = _COPILOT_DOWNLOADS.get(download_id)
    if not item:
        raise HTTPException(status_code=404, detail="Reporte no encontrado o expirado.")

    filename = item["filename"]
    return StreamingResponse(
        io.BytesIO(item["content"]),
        media_type=item["mime_type"],
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/v1/copilot/email/send")
async def send_copilot_email(
    payload: CopilotEmailSendRequest,
    operator_ctx: Dict[str, Any] = Depends(require_audit_read_access),
):
    mall_id = str(payload.mall_id or "").strip()
    draft_id = str(payload.draft_id or "").strip()
    if not mall_id or not draft_id:
        raise HTTPException(status_code=400, detail="mall_id y draft_id son requeridos.")
    _ensure_operator_can_access_mall(operator_ctx, mall_id)

    with _COPILOT_EMAIL_DRAFTS_LOCK:
        _cleanup_copilot_email_drafts()
        draft = _COPILOT_EMAIL_DRAFTS.get(draft_id)
    if not draft or draft.get("mall_id") != mall_id:
        raise HTTPException(status_code=404, detail="Borrador de correo no encontrado o expirado.")

    recipients = draft.get("recipients") or []
    if not recipients:
        raise HTTPException(status_code=400, detail="El borrador no tiene destinatarios.")

    sent = []
    for recipient in recipients:
        result = await asyncio.to_thread(
            _send_resend_email,
            recipient,
            draft.get("subject") or "Reporte MsMall",
            draft.get("text") or "Reporte generado por Copilot MsMall.",
            draft.get("html_body"),
            None,
            draft.get("attachments") or [],
        )
        sent.append({"email": recipient, "resend_id": result.get("id")})

    with _COPILOT_EMAIL_DRAFTS_LOCK:
        _COPILOT_EMAIL_DRAFTS.pop(draft_id, None)

    return {
        "sent": sent,
        "subject": draft.get("subject"),
        "attachment_count": len(draft.get("attachments") or []),
    }


def _copilot_system_prompt() -> str:
    return (
        "Eres MsMall Copilot, el asistente operativo del sistema MsMall. "
        "Responde en español, de forma breve y accionable. Usa solamente el contexto JSON del sistema: "
        "ventas recientes, monitor de carga, monitor de conexiones, locales y dias de informacion. "
        "Si el contexto no contiene un dato solicitado, dilo claramente y sugiere donde revisarlo. "
        "No inventes cifras, locales, fechas ni estados. Cuando sea util, menciona la fuente del dato. "
        "Formato obligatorio: usa un titulo corto en negrita, luego lineas separadas con bullets. "
        "Para reportes numericos, incluye periodo, fuente y maximo 8 bullets. "
        "No respondas en un parrafo largo; evita tablas markdown porque el chat es angosto."
    )


def _sanitize_copilot_history(history: List[CopilotChatMessage]) -> List[Dict[str, str]]:
    sanitized = []
    for item in (history or [])[-8:]:
        role = str(item.role or "").strip().lower()
        if role not in {"user", "assistant"}:
            continue
        content = _truncate_text(item.content, 900)
        if content:
            sanitized.append({"role": role, "content": content})
    return sanitized


def _build_copilot_user_context(context: Dict[str, Any], message: str) -> str:
    context_json = json.dumps(context, ensure_ascii=False, default=str)
    return (
        "Contexto operacional de MsMall en JSON:\n"
        f"{context_json}\n\n"
        "Pregunta del usuario:\n"
        f"{message.strip()}"
    )


def _extract_llm_error(prefix: str, raw: str) -> str:
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            error = parsed.get("error")
            if isinstance(error, dict):
                return error.get("message") or error.get("status") or prefix
            if isinstance(error, str):
                return error
            if isinstance(parsed.get("message"), str):
                return parsed["message"]
    except Exception:
        pass
    return raw[:240] if raw else prefix


def _call_openai_copilot(api_key: str, model: str, context: Dict[str, Any], message: str, history: List[CopilotChatMessage]) -> str:
    messages = [{"role": "system", "content": _copilot_system_prompt()}]
    messages.extend(_sanitize_copilot_history(history))
    messages.append({"role": "user", "content": _build_copilot_user_context(context, message)})
    payload = {
        "model": model or COPILOT_DEFAULT_MODELS["openai"],
        "messages": messages,
        "temperature": 0.2,
        "max_tokens": 700,
    }
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": RESEND_USER_AGENT,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as response:
            parsed = json.loads(response.read().decode("utf-8"))
        return (
            parsed.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
            .strip()
        ) or "No recibi una respuesta del proveedor."
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(status_code=502, detail=f"OpenAI: {_extract_llm_error('Error consultando OpenAI.', raw)}")
    except urllib.error.URLError:
        raise HTTPException(status_code=502, detail="No se pudo conectar con OpenAI.")
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Copilot OpenAI error: %s", sanitize_sensitive_ops_error(exc))
        raise HTTPException(status_code=500, detail="Error inesperado consultando OpenAI.")


def _call_gemini_copilot(api_key: str, model: str, context: Dict[str, Any], message: str, history: List[CopilotChatMessage]) -> str:
    history_text = "\n".join(
        f"{item['role']}: {item['content']}"
        for item in _sanitize_copilot_history(history)
    )
    prompt = (
        f"{_build_copilot_user_context(context, message)}\n\n"
        f"Historial reciente:\n{history_text or 'Sin historial previo.'}"
    )
    payload = {
        "systemInstruction": {"parts": [{"text": _copilot_system_prompt()}]},
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.2,
            "maxOutputTokens": 700,
        },
    }
    safe_model = (model or COPILOT_DEFAULT_MODELS["gemini"]).strip()
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{safe_model}:generateContent?key={api_key}"
    req = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": RESEND_USER_AGENT,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as response:
            parsed = json.loads(response.read().decode("utf-8"))
        parts = (
            parsed.get("candidates", [{}])[0]
            .get("content", {})
            .get("parts", [])
        )
        return "\n".join(str(part.get("text") or "").strip() for part in parts if part.get("text")).strip() or "No recibi una respuesta del proveedor."
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(status_code=502, detail=f"Gemini: {_extract_llm_error('Error consultando Gemini.', raw)}")
    except urllib.error.URLError:
        raise HTTPException(status_code=502, detail="No se pudo conectar con Gemini.")
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Copilot Gemini error: %s", sanitize_sensitive_ops_error(exc))
        raise HTTPException(status_code=500, detail="Error inesperado consultando Gemini.")


def _call_copilot_provider(settings: Dict[str, Any], context: Dict[str, Any], message: str, history: List[CopilotChatMessage]) -> str:
    provider = _normalize_copilot_provider(settings.get("provider"))
    api_key = _get_system_health_value(_copilot_api_key_name(provider)) or ""
    if not api_key:
        raise HTTPException(status_code=503, detail="Copilot no tiene API key configurada.")
    if provider == "gemini":
        return _call_gemini_copilot(api_key, settings.get("model") or "", context, message, history)
    return _call_openai_copilot(api_key, settings.get("model") or "", context, message, history)


@app.get("/api/v1/admin/copilot/settings")
async def get_copilot_settings(admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    return _copilot_config_status()


@app.put("/api/v1/admin/copilot/settings")
async def save_copilot_settings(
    payload: CopilotSettingsRequest,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    try:
        return _save_copilot_settings(payload)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error guardando configuracion de Copilot: %s", sanitize_sensitive_ops_error(exc))
        raise HTTPException(status_code=500, detail="No se pudo guardar la configuracion de Copilot.")


@app.get("/api/v1/copilot/status")
async def get_copilot_status(
    mall_id: str = Query(...),
    operator_ctx: Dict[str, Any] = Depends(require_audit_read_access),
):
    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    return _copilot_config_status()


@app.post("/api/v1/copilot/chat")
async def chat_with_copilot(
    payload: CopilotChatRequest,
    operator_ctx: Dict[str, Any] = Depends(require_audit_read_access),
):
    mall_id = str(payload.mall_id or "").strip()
    message = str(payload.message or "").strip()
    if not mall_id:
        raise HTTPException(status_code=400, detail="mall_id es requerido.")
    if not message:
        raise HTTPException(status_code=400, detail="Escribe una pregunta para Copilot.")
    if len(message) > 1400:
        raise HTTPException(status_code=400, detail="La pregunta es demasiado larga.")

    big_data_question = _is_big_data_copilot_question(message)
    settings = _copilot_config_status()
    if not settings.get("enabled"):
        raise HTTPException(status_code=503, detail="Copilot MsMall esta desactivado.")
    if not settings.get("api_key_configured") and not big_data_question:
        raise HTTPException(status_code=503, detail="Copilot MsMall no tiene API key configurada.")

    context_builder = (
        _build_big_data_copilot_context if big_data_question else _build_copilot_context
    )
    context = await asyncio.to_thread(context_builder, mall_id, operator_ctx)
    email_request = _parse_copilot_email_request(message)
    if email_request:
        try:
            email_draft = await asyncio.to_thread(_build_copilot_email_draft, email_request, context, mall_id, operator_ctx)
        except HTTPException as exc:
            if exc.status_code == 400:
                return {
                    "answer": (
                        "**Falta destinatario**\n"
                        "- Puedo preparar el reporte en HTML y enviarlo por correo.\n"
                        "- Indica el email destino, por ejemplo: `envialo a operaciones@empresa.com`."
                    ),
                    "provider": settings["provider"],
                    "model": settings["model"],
                    "context_generated_at": context.get("generated_at_utc"),
                    "sources": [],
                    "attachments": [],
                    "email_actions": [],
                }
            raise
        attachment_text = "HTML"
        if email_draft.get("attachment_count"):
            attachment_text += " + adjunto"
        return {
            "answer": (
                f"**Correo preparado**\n"
                f"- Para: **{', '.join(email_draft['recipients'])}**\n"
                f"- Asunto: **{email_draft['subject']}**\n"
                f"- Formato: **{attachment_text}**\n"
                f"- Filas incluidas: **{email_draft.get('row_count', 0)}**\n"
                f"- Confirma con el boton para enviarlo."
            ),
            "provider": settings["provider"],
            "model": settings["model"],
            "context_generated_at": context.get("generated_at_utc"),
            "sources": email_draft.get("sources") or [],
            "attachments": [],
            "email_actions": [email_draft],
        }

    report_request = _parse_copilot_report_request(message)
    if report_request:
        attachment = await asyncio.to_thread(_generate_copilot_report_attachment, report_request, context)
        return {
            "answer": (
                f"**Reporte listo**\n"
                f"- Archivo: **{attachment['filename']}**\n"
                f"- Formato: **{report_request['format'].upper()}**\n"
                f"- Filas incluidas: **{attachment.get('row_count', 0)}**\n"
                f"- El enlace expira en aproximadamente 15 minutos."
            ),
            "provider": settings["provider"],
            "model": settings["model"],
            "context_generated_at": context.get("generated_at_utc"),
            "sources": attachment.get("sources") or [],
            "attachments": [attachment],
        }

    try:
        answer = await asyncio.to_thread(
            _call_copilot_provider,
            settings,
            context,
            message,
            payload.history or [],
        )
        provider = settings["provider"]
        model = settings["model"]
    except HTTPException:
        if not big_data_question:
            raise
        # Structured metrics remain useful when the provider is unavailable.
        answer = _deterministic_big_data_answer(context)
        provider = "deterministic"
        model = "big-data-fallback-v1"
    return {
        "answer": answer,
        "provider": provider,
        "model": model,
        "context_generated_at": context.get("generated_at_utc"),
        "sources": (
            ["big_data_aggregates", "operational_findings", "big_data_forecast"]
            if big_data_question
            else ["ventas_recientes", "monitor_carga", "monitor_conexiones", "locales", "dias_informacion"]
        ),
        "attachments": [],
    }


def _normalize_resend_sender_email(value: str) -> str:
    email = str(value or "").strip().lower()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        raise HTTPException(status_code=400, detail="Correo remitente invalido.")
    if not email.endswith(f"@{RESEND_DOMAIN}"):
        raise HTTPException(status_code=400, detail=f"El remitente debe usar el dominio {RESEND_DOMAIN}.")
    return email


def _normalize_resend_sender_name(value: str) -> str:
    name = str(value or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nombre remitente requerido.")
    if len(name) > 80:
        raise HTTPException(status_code=400, detail="Nombre remitente demasiado largo.")
    return name


def _resolve_resend_sender_config() -> Dict[str, str]:
    raw_email = _get_system_health_value(RESEND_SENDER_EMAIL_KEY) or os.getenv("RESEND_FROM_EMAIL") or RESEND_FROM_EMAIL
    raw_name = _get_system_health_value(RESEND_SENDER_NAME_KEY) or os.getenv("RESEND_FROM_NAME") or RESEND_FROM_NAME
    try:
        from_email = _normalize_resend_sender_email(raw_email)
    except HTTPException:
        logger.warning("Remitente Resend invalido en configuracion persistida: %s", raw_email)
        from_email = RESEND_FROM_EMAIL
    try:
        from_name = _normalize_resend_sender_name(raw_name)
    except HTTPException:
        logger.warning("Nombre remitente Resend invalido en configuracion persistida: %s", raw_name)
        from_name = RESEND_FROM_NAME
    return {"from_email": from_email, "from_name": from_name}


def _resend_config_status() -> Dict[str, Any]:
    sender = _resolve_resend_sender_config()
    return {
        "provider": "resend",
        "domain": RESEND_DOMAIN,
        "from_email": sender["from_email"],
        "from_name": sender["from_name"],
        "configured": bool(os.getenv(RESEND_API_KEY_ENV)),
        "api_key_env": RESEND_API_KEY_ENV,
    }


def _send_resend_email(
    to_email: str,
    subject: str,
    message: str,
    html_body: Optional[str] = None,
    cc_emails: Optional[List[str]] = None,
    attachments: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    api_key = os.getenv(RESEND_API_KEY_ENV)
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=f"Resend no esta configurado. Falta la variable {RESEND_API_KEY_ENV}.",
        )

    sender = _resolve_resend_sender_config()
    payload = {
        "from": f"{sender['from_name']} <{sender['from_email']}>",
        "to": [to_email],
        "subject": subject,
        "text": message,
        "html": html_body or f"<p>{html.escape(message or '').replace(chr(10), '<br />')}</p>",
    }
    cc_list = _normalize_email_list(cc_emails or [])
    if cc_list:
        payload["cc"] = cc_list
    attachment_payload = []
    for attachment in attachments or []:
        content = attachment.get("content") or b""
        if isinstance(content, str):
            content_bytes = content.encode("utf-8")
        else:
            content_bytes = bytes(content)
        attachment_payload.append({
            "filename": attachment.get("filename") or "msmall_reporte",
            "content": base64.b64encode(content_bytes).decode("ascii"),
            "content_type": attachment.get("mime_type") or "application/octet-stream",
        })
    if attachment_payload:
        payload["attachments"] = attachment_payload

    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": RESEND_USER_AGENT,
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        try:
            parsed = json.loads(raw)
            detail = parsed.get("message") or parsed.get("detail") or parsed.get("error") or raw
        except Exception:
            detail = raw or "Resend rechazo el envio."
        logger.warning("Resend API error %s: %s", e.code, detail)
        raise HTTPException(status_code=502, detail=f"Resend: {detail}")
    except urllib.error.URLError as e:
        logger.error("Resend network error: %s", e)
        raise HTTPException(status_code=502, detail="No se pudo conectar con Resend.")
    except Exception as e:
        logger.error("Unexpected Resend error: %s", e)
        raise HTTPException(status_code=500, detail="Error inesperado enviando con Resend.")


def _normalize_missing_days_notification_type(value: str) -> str:
    notification_type = str(value or MISSING_DAYS_NOTIFICATION_TYPE).strip()
    if notification_type not in MISSING_DAYS_NOTIFICATION_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de envio de dias faltantes invalido.")
    return notification_type


def _missing_days_default_templates(notification_type: str) -> tuple[str, str]:
    if notification_type == MISSING_DAYS_CONSOLIDATED_NOTIFICATION_TYPE:
        return DEFAULT_CONSOLIDATED_SUBJECT_TEMPLATE, DEFAULT_CONSOLIDATED_BODY_TEMPLATE
    return DEFAULT_MISSING_DAYS_SUBJECT_TEMPLATE, DEFAULT_MISSING_DAYS_BODY_TEMPLATE


def _default_missing_days_email_settings(
    mall_id: str,
    notification_type: str = MISSING_DAYS_NOTIFICATION_TYPE,
) -> Dict[str, Any]:
    notification_type = _normalize_missing_days_notification_type(notification_type)
    subject_template, body_template = _missing_days_default_templates(notification_type)
    return {
        "id": None,
        "mall_id": mall_id,
        "notification_type": notification_type,
        "enabled": False,
        "weekdays": [],
        "send_time": "08:00",
        "lookback_days": 7,
        "send_only_with_gaps": True,
        "cc_emails": [],
        "subject_template": subject_template,
        "body_template": body_template,
        "created_at": None,
        "updated_at": None,
    }


def _normalize_weekdays(weekdays: List[int]) -> List[int]:
    normalized = set()
    for day in weekdays or []:
        try:
            value = int(day)
        except (TypeError, ValueError):
            continue
        if 0 <= value <= 6:
            normalized.add(value)
    return sorted(normalized)


def _normalize_email_list(values: List[str], *, strict: bool = True) -> List[str]:
    emails = []
    if isinstance(values, str):
        values = re.split(r"[\n,;]+", values)
    for value in values or []:
        email = str(value or "").strip().lower()
        if not email:
            continue
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            if not strict:
                continue
            raise HTTPException(status_code=400, detail=f"Email invalido: {email}")
        emails.append(email)
    return sorted(set(emails))


def _normalize_send_time(value: str) -> str:
    candidate = str(value or "").strip()
    if re.match(r"^\d{2}:\d{2}$", candidate):
        candidate = f"{candidate}:00"
    if not re.match(r"^\d{2}:\d{2}:\d{2}$", candidate):
        raise HTTPException(status_code=400, detail="Hora de envio invalida. Use HH:MM.")
    hour, minute, second = [int(part) for part in candidate.split(":")]
    if hour > 23 or minute > 59 or second > 59:
        raise HTTPException(status_code=400, detail="Hora de envio invalida. Use HH:MM.")
    return f"{hour:02d}:{minute:02d}:00"


def _normalize_email_template(value: Optional[str], default: str, *, max_length: int) -> str:
    template = str(value or "").strip() or default
    if len(template) > max_length:
        raise HTTPException(status_code=400, detail=f"La plantilla no puede exceder {max_length} caracteres.")
    return template


def _sanitize_missing_days_email_settings_row(
    row: Optional[Dict[str, Any]],
    mall_id: str,
    notification_type: str = MISSING_DAYS_NOTIFICATION_TYPE,
) -> Dict[str, Any]:
    notification_type = _normalize_missing_days_notification_type(notification_type)
    if not row:
        return _default_missing_days_email_settings(mall_id, notification_type)
    row_notification_type = _normalize_missing_days_notification_type(
        row.get("notification_type") or notification_type
    )
    subject_template, body_template = _missing_days_default_templates(row_notification_type)
    data = _default_missing_days_email_settings(mall_id, row_notification_type)
    try:
        lookback_days = int(row.get("lookback_days") or 7)
    except (TypeError, ValueError):
        lookback_days = 7
    lookback_days = max(1, min(90, lookback_days))
    data.update({
        "id": row.get("id"),
        "mall_id": row.get("mall_id") or mall_id,
        "notification_type": row_notification_type,
        "enabled": bool(row.get("enabled")),
        "weekdays": _normalize_weekdays(row.get("weekdays") or []),
        "send_time": str(row.get("send_time") or "08:00")[:5],
        "lookback_days": lookback_days,
        "send_only_with_gaps": row.get("send_only_with_gaps") is not False,
        "cc_emails": _normalize_email_list(row.get("cc_emails") or [], strict=False),
        "subject_template": _normalize_email_template(
            row.get("subject_template"),
            subject_template,
            max_length=160,
        ),
        "body_template": _normalize_email_template(
            row.get("body_template"),
            body_template,
            max_length=2000,
        ),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    })
    return data


def _is_missing_email_settings_table_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "email_notification_settings" in text and (
        "does not exist" in text or "schema cache" in text or "pgrst205" in text
    )


def _normalize_missing_days_sale_date(raw_value: Any) -> Optional[str]:
    if raw_value is None:
        return None
    if isinstance(raw_value, datetime):
        return raw_value.strftime("%Y-%m-%d")
    value = str(raw_value).strip()
    if not value:
        return None
    if len(value) >= 10 and value[4] == "-" and value[7] == "-":
        return value[:10]
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.strftime("%Y-%m-%d")
    except Exception:
        return None


def _load_missing_days_details_for_local(
    local_id: str,
    local_name: str,
    mall_id: str,
    fecha_inicio: str,
    fecha_fin: str,
) -> List[Dict[str, Any]]:
    start_date = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    end_date = datetime.strptime(fecha_fin, "%Y-%m-%d")
    total_days = (end_date - start_date).days + 1
    expected_dates = {
        (start_date + timedelta(days=x)).strftime("%Y-%m-%d")
        for x in range(total_days)
    }

    rows: List[Dict[str, Any]] = []
    page_size = 2000
    page = 0
    while True:
        chunk = (
            supabase.table("ventas")
            .select("id, fecha")
            .eq("local_id", local_id)
            .gte("fecha", fecha_inicio)
            .lte("fecha", fecha_fin)
            .order("id")
            .range(page * page_size, (page + 1) * page_size - 1)
            .execute()
        ).data or []
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        page += 1

    actual_dates = {
        normalized
        for normalized in (_normalize_missing_days_sale_date(row.get("fecha")) for row in rows)
        if normalized
    }
    missing_dates = sorted(list(expected_dates - actual_dates))
    if not missing_dates:
        return []

    try:
        logs_resp = (
            supabase.table("logs_carga")
            .select("*")
            .eq("local_id", local_id)
            .gte("fecha_hora", f"{fecha_inicio}T00:00:00")
            .lte("fecha_hora", f"{fecha_fin}T23:59:59")
            .order("fecha_hora", desc=True)
            .execute()
        )
    except Exception:
        logs_resp = type("Tmp", (), {"data": []})()

    if not logs_resp.data and local_name:
        legacy_q = (
            supabase.table("logs_carga")
            .select("*")
            .eq("local_nombre", local_name)
            .gte("fecha_hora", f"{fecha_inicio}T00:00:00")
            .lte("fecha_hora", f"{fecha_fin}T23:59:59")
            .order("fecha_hora", desc=True)
        )
        if mall_id:
            legacy_q = legacy_q.eq("mall_id", mall_id)
        logs_resp = legacy_q.execute()

    logs_df = pd.DataFrame(logs_resp.data or [])
    if not logs_df.empty:
        logs_df["fecha_log"] = logs_df["fecha_hora"].apply(lambda x: str(x).split("T")[0] if x else None)

    details: List[Dict[str, Any]] = []
    for missing_date in missing_dates:
        cause = "Proceso no ejecutado / Sin conexión"
        log_id = None
        if not logs_df.empty:
            day_logs = logs_df[logs_df["fecha_log"] == missing_date]
            if not day_logs.empty:
                last_log = day_logs.iloc[0]
                log_id = last_log.get("id")
                status_text = str(last_log.get("estado") or "").strip().lower()
                if status_text == "error":
                    cause = "Fallo Técnico / Error de Lectura"
                elif status_text in {"no_encontrado", "no encontrado"}:
                    cause = "Archivo no disponible en FTP"
                elif status_text in {"exito", "éxito", "success", "parcial"}:
                    cause = "Procesado con Éxito (Posible archivo vacío)"
        details.append({"fecha": missing_date, "causa": cause, "log_id": log_id})
    return details


def _missing_days_report_url(mall_id: str, local_id: str, fecha_inicio: str, fecha_fin: str) -> Optional[str]:
    app_url = (os.getenv("APP_BASE_URL") or os.getenv("FRONTEND_URL") or "").strip().rstrip("/")
    if not app_url:
        return None
    return (
        f"{app_url}/?view=reports"
        f"&mall_id={mall_id}"
        f"&local_id={local_id}"
        f"&start_date={fecha_inicio}"
        f"&end_date={fecha_fin}"
    )


@app.get("/api/v1/admin/messaging/resend")
async def get_resend_messaging_config(admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    return _resend_config_status()


@app.put("/api/v1/admin/messaging/resend/sender")
async def update_resend_sender_config(
    payload: ResendSenderUpdateRequest,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    from_email = _normalize_resend_sender_email(payload.from_email)
    from_name = _normalize_resend_sender_name(payload.from_name)
    try:
        _upsert_system_health_value_sync(RESEND_SENDER_EMAIL_KEY, from_email)
        _upsert_system_health_value_sync(RESEND_SENDER_NAME_KEY, from_name)
    except Exception as exc:
        logger.error("Error guardando remitente Resend: %s", exc)
        raise HTTPException(status_code=500, detail="No se pudo guardar el remitente de Resend.")
    return _resend_config_status()


@app.post("/api/v1/admin/messaging/resend/test")
async def send_resend_test_message(
    payload: ResendTestMessageRequest,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    to_email = (payload.to or "").strip()
    subject = (payload.subject or "Prueba de notificaciones MSMALL").strip()
    message = (payload.message or "Mensaje de prueba desde MSMALL usando Resend.").strip()

    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", to_email):
        raise HTTPException(status_code=400, detail="Destinatario invalido.")
    if not subject:
        raise HTTPException(status_code=400, detail="Asunto requerido.")
    if not message:
        raise HTTPException(status_code=400, detail="Mensaje requerido.")

    result = await asyncio.to_thread(_send_resend_email, to_email, subject, message)
    return {
        "status": "success",
        "id": result.get("id"),
        "to": to_email,
        "from_email": _resolve_resend_sender_config()["from_email"],
        "domain": RESEND_DOMAIN,
        "message": "Mensaje de prueba enviado correctamente.",
    }


@app.post("/api/v1/admin/messaging/missing-days/preview-html")
async def preview_missing_days_email_html(
    payload: MissingDaysEmailPreviewRequest,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    return {
        "html": build_missing_days_email_html(
            mall_name=payload.mall_name,
            local_name=payload.local_name,
            fecha_inicio=payload.fecha_inicio,
            fecha_fin=payload.fecha_fin,
            missing_details=payload.missing_details,
            report_url=payload.report_url,
        )
    }


@app.get("/api/v1/admin/messaging/missing-days/settings")
async def get_missing_days_email_settings(
    mall_id: str = Query(...),
    notification_type: str = Query(MISSING_DAYS_NOTIFICATION_TYPE),
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado.")
    _ensure_operator_can_access_mall(admin_ctx, mall_id)
    notification_type = _normalize_missing_days_notification_type(notification_type)
    try:
        res = (
            supabase.table("email_notification_settings")
            .select("*")
            .eq("mall_id", mall_id)
            .eq("notification_type", notification_type)
            .maybe_single()
            .execute()
        )
        return _sanitize_missing_days_email_settings_row(res.data, mall_id, notification_type)
    except Exception as exc:
        if _is_missing_email_settings_table_error(exc):
            logger.warning(
                "Tabla email_notification_settings no disponible al cargar programacion; usando defaults para mall %s: %s",
                mall_id,
                exc,
            )
            return _default_missing_days_email_settings(mall_id, notification_type)
        logger.warning(
            "Error cargando configuracion de emails de dias faltantes; usando defaults para mall %s: %s",
            mall_id,
            exc,
        )
        return _default_missing_days_email_settings(mall_id, notification_type)


@app.put("/api/v1/admin/messaging/missing-days/settings")
async def save_missing_days_email_settings(
    payload: MissingDaysEmailSettingsRequest,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado.")
    mall_id = (payload.mall_id or "").strip()
    if not mall_id:
        raise HTTPException(status_code=400, detail="mall_id es requerido.")
    _ensure_operator_can_access_mall(admin_ctx, mall_id)
    notification_type = _normalize_missing_days_notification_type(payload.notification_type)

    weekdays = _normalize_weekdays(payload.weekdays)
    if payload.enabled and not weekdays:
        raise HTTPException(status_code=400, detail="Seleccione al menos un dia de envio.")

    lookback_days = int(payload.lookback_days or 7)
    if lookback_days < 1 or lookback_days > 90:
        raise HTTPException(status_code=400, detail="La ventana de auditoria debe estar entre 1 y 90 dias.")

    cc_emails = _normalize_email_list(payload.cc_emails)
    if notification_type == MISSING_DAYS_CONSOLIDATED_NOTIFICATION_TYPE and payload.enabled and not cc_emails:
        raise HTTPException(
            status_code=400,
            detail="Agregue al menos un correo administrativo para activar el envio consolidado.",
        )

    subject_template, body_template = _missing_days_default_templates(notification_type)
    row = {
        "mall_id": mall_id,
        "notification_type": notification_type,
        "enabled": bool(payload.enabled),
        "weekdays": weekdays,
        "send_time": _normalize_send_time(payload.send_time),
        "lookback_days": lookback_days,
        "send_only_with_gaps": bool(payload.send_only_with_gaps),
        "cc_emails": cc_emails,
        "subject_template": _normalize_email_template(
            payload.subject_template,
            subject_template,
            max_length=160,
        ),
        "body_template": _normalize_email_template(
            payload.body_template,
            body_template,
            max_length=2000,
        ),
        "updated_by": admin_ctx.get("user_id"),
        "updated_at": datetime.utcnow().isoformat(),
    }

    try:
        res = (
            supabase.table("email_notification_settings")
            .upsert(row, on_conflict="mall_id,notification_type")
            .execute()
        )
        saved = (res.data or [row])[0]
        return _sanitize_missing_days_email_settings_row(saved, mall_id, notification_type)
    except Exception as exc:
        if _is_missing_email_settings_table_error(exc):
            raise HTTPException(
                status_code=503,
                detail="La base de datos no está actualizada: ejecute el script 20260511_email_notification_settings.sql.",
            )
        logger.error("Error guardando configuracion de emails de dias faltantes: %s", exc)
        raise HTTPException(status_code=500, detail="No se pudo guardar la programacion de envio.")


@app.post("/api/v1/admin/messaging/missing-days/send-now")
async def send_missing_days_email_now(
    payload: MissingDaysSendNowRequest,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado.")
    if not os.getenv(RESEND_API_KEY_ENV):
        raise HTTPException(status_code=503, detail=f"Resend no esta configurado. Falta {RESEND_API_KEY_ENV}.")

    mall_id = (payload.mall_id or "").strip()
    if not mall_id:
        raise HTTPException(status_code=400, detail="mall_id es requerido.")
    _ensure_operator_can_access_mall(admin_ctx, mall_id)
    notification_type = _normalize_missing_days_notification_type(payload.notification_type)

    try:
        settings_res = (
            supabase.table("email_notification_settings")
            .select("*")
            .eq("mall_id", mall_id)
            .eq("notification_type", notification_type)
            .maybe_single()
            .execute()
        )
        settings = _sanitize_missing_days_email_settings_row(
            settings_res.data,
            mall_id,
            notification_type,
        )
    except Exception as exc:
        if _is_missing_email_settings_table_error(exc):
            logger.warning(
                "Tabla email_notification_settings no disponible para envio inmediato; usando defaults para mall %s: %s",
                mall_id,
                exc,
            )
            settings = _default_missing_days_email_settings(mall_id, notification_type)
        else:
            logger.warning(
                "Error cargando configuracion para envio inmediato; usando defaults para mall %s: %s",
                mall_id,
                exc,
            )
            settings = _default_missing_days_email_settings(mall_id, notification_type)

    if notification_type == MISSING_DAYS_CONSOLIDATED_NOTIFICATION_TYPE and not settings.get("cc_emails"):
        raise HTTPException(
            status_code=400,
            detail="Agregue al menos un correo administrativo antes de enviar el consolidado.",
        )

    result = await asyncio.to_thread(
        send_missing_days_emails_for_mall,
        supabase,
        settings,
        logger=logger,
    )
    return {
        **result,
        "message": (
            f"Envio inmediato completado: {result['sent']} enviados, "
            f"{result['skipped']} omitidos, {result['failed']} fallidos."
        ),
    }


@app.delete("/api/v1/admin/reset-sales")
async def reset_sales(admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    """Wipes all sales data to reset testing environment."""
    if not supabase:
         raise HTTPException(status_code=500, detail="Supabase client not initialized")
    try:
        # Delete all records where ID is NOT a dummy value (effectively all UUIDs)
        # Using a dummy UUID known not to exist: 0000...0000
        res = supabase.table("ventas").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        count = len(res.data) if res.data else 0
        invalidated = _cache_delete_prefix("analytics:dashboard:")
        logger.info(f"Reset sales requested by admin. Deleted {count} records.")
        logger.info("Dashboard BI cache cleared after sales reset (keys=%s)", invalidated)
        return {"status": "success", "message": f"Se han eliminado {count} registros de ventas."}
    except Exception as e:
        logger.error(f"Error resetting sales: {e}")
        raise HTTPException(status_code=500, detail=f"Error borrando ventas: {str(e)}")

@app.get("/api/v1/analytics/dashboard")
async def get_dashboard_data(start_date: str, end_date: str, mall_id: str = Depends(get_current_mall)):
    """
    Returns aggregated KPI data for the dashboard.
    Bypasses RLS by using the backend Service Role key.
    """
    if not supabase:
         raise HTTPException(status_code=500, detail="Supabase client not initialized")
    try:
        parsed_start = date.fromisoformat(start_date)
        parsed_end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Las fechas deben usar el formato YYYY-MM-DD.")
    if parsed_start > parsed_end:
        raise HTTPException(status_code=400, detail="La fecha inicial no puede ser posterior a la fecha final.")

    mode = _dashboard_mode()
    cache_key = f"analytics:dashboard:{mode}:{mall_id}:{start_date}:{end_date}"
    started_at = time.perf_counter()
    cached = _cache_get(cache_key)
    if cached is not _CACHE_MISS:
        logger.info(
            "Dashboard BI served (mall=%s, mode=%s, source=cache, duration_ms=%.1f)",
            mall_id,
            mode,
            (time.perf_counter() - started_at) * 1000,
        )
        return cached

    load_lock = _dashboard_load_lock(cache_key)
    async with load_lock:
        cached = _cache_get(cache_key)
        if cached is not _CACHE_MISS:
            logger.info(
                "Dashboard BI served after single-flight wait (mall=%s, mode=%s, duration_ms=%.1f)",
                mall_id,
                mode,
                (time.perf_counter() - started_at) * 1000,
            )
            return cached

        try:
            result, source = await asyncio.to_thread(
                DashboardAnalyticsService(supabase, logger).load,
                mall_id,
                start_date,
                end_date,
                mode=mode,
            )
            _cache_set(cache_key, result, TTL_DASHBOARD)
            logger.info(
                "Dashboard BI served (mall=%s, mode=%s, source=%s, duration_ms=%.1f)",
                mall_id,
                mode,
                source,
                (time.perf_counter() - started_at) * 1000,
            )
            return result
        except HTTPException:
            raise
        except Exception as exc:
            logger.error("Error fetching dashboard data: %s", exc)
            raise HTTPException(status_code=500, detail=str(exc))

# --- EXPORT ENDPOINTS ---
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from services.export_service import ExportService
from typing import Optional

# Initialize service
# Ensure supabase client is available. It is global in this file.
export_service = ExportService(supabase)
router_export = APIRouter(prefix="/api/v1/export", tags=["export"])



@app.get("/api/v1/users/me/malls")
async def get_my_malls(user_id: str = Depends(get_current_user_id)):
    """
    Returns the list of malls assigned to the current user.
    """
    try:
        # Join usuarios_malls with malls
        # Supabase-py doesn't support easy joins in one go unless defined in View or specific syntax.
        # We'll do two queries or use a raw query if enabled (RPC).
        # Standard way: Fetch user-malls, then fetch malls.
        
        # 1. Get Mall IDs for user
        um_res = supabase.table("usuarios_malls").select("mall_id, rol").eq("usuario_id", user_id).execute()
        malls_links = um_res.data
        
        if not malls_links:
            return []
            
        mall_ids = [m['mall_id'] for m in malls_links]
        
        # 2. Get Mall Details
        malls_res = supabase.table("malls").select("*").in_("id", mall_ids).execute()
        malls_details = {m['id']: m for m in malls_res.data}
        
        # 3. Merge
        result = []
        for link in malls_links:
            mid = link['mall_id']
            if mid in malls_details:
                result.append({
                    "id": mid,
                    "nombre": malls_details[mid]['nombre'],
                    "rol": link['rol']
                })
                

                
        return result
    except Exception as e:
        logger.error(f"Error fetching user malls: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/users/me/access")
async def get_my_access_context(user_id: str = Depends(get_current_user_id)):
    """
    Returns the effective access context for the current user.
    Useful for frontend UI gating (role-based menus/actions).
    """
    try:
        return await _get_access_context(user_id)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching access context: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Mall Management Endpoints (Admin) ---

class MallCreate(BaseModel):
    nombre: str
    conf_locale: Optional[str] = 'es-CL'
    conf_moneda: Optional[str] = 'CLP'
    metadata: Optional[Dict] = {}

class MallUpdate(BaseModel):
    nombre: Optional[str] = None
    conf_locale: Optional[str] = None
    conf_moneda: Optional[str] = None
    metadata: Optional[Dict] = None

BIG_DATA_FEATURE_FLAGS = {
    "BIG_DATA_CORE",
    "BIG_DATA_BENCHMARK",
    "BIG_DATA_FORECAST",
    "BIG_DATA_OPERATIONS",
    "BIG_DATA_COPILOT",
}

class MallFeatureFlagUpdate(BaseModel):
    enabled: bool

@app.get("/api/v1/malls/all")
async def get_all_malls(admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    """
    Get all malls. Restricted to Admins/SuperAdmins.
    (RLS will filter if user is not admin, but good to check role here too if needed)
    """
    try:
        # Simplified role check - in prod rely on RLS or specific logic
        res = supabase.table("malls").select("*").execute()
        return res.data
    except Exception as e:
        logger.error(f"Error fetching all malls: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/malls/{mall_id}/feature-flags")
async def get_mall_feature_flags(
    mall_id: str,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    """Return the feature entitlement state for one mall; admin only."""
    try:
        mall = supabase.table("malls").select("id").eq("id", mall_id).maybe_single().execute()
        if not mall.data:
            raise HTTPException(status_code=404, detail="Mall no encontrado")
        response = (
            supabase.table("mall_feature_flags")
            .select("feature_key,enabled,updated_at,updated_by")
            .eq("mall_id", mall_id)
            .execute()
        )
        saved = {row.get("feature_key"): row for row in (response.data or [])}
        return [
            {
                "feature_key": key,
                "enabled": bool(saved.get(key, {}).get("enabled", False)),
                "updated_at": saved.get(key, {}).get("updated_at"),
                "updated_by": saved.get(key, {}).get("updated_by"),
            }
            for key in sorted(BIG_DATA_FEATURE_FLAGS)
        ]
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error consultando feature flags de mall {mall_id}: {e}")
        raise HTTPException(status_code=500, detail="No se pudieron cargar los módulos del mall")

@app.put("/api/v1/malls/{mall_id}/feature-flags/{feature_key}")
async def update_mall_feature_flag(
    mall_id: str,
    feature_key: str,
    payload: MallFeatureFlagUpdate,
    admin_ctx: Dict[str, Any] = Depends(require_admin_access),
):
    """Enable or disable a mall entitlement from the administrator UX."""
    normalized_key = (feature_key or "").strip().upper()
    if normalized_key not in BIG_DATA_FEATURE_FLAGS:
        raise HTTPException(status_code=400, detail="Feature flag no soportado")
    try:
        mall = supabase.table("malls").select("id").eq("id", mall_id).maybe_single().execute()
        if not mall.data:
            raise HTTPException(status_code=404, detail="Mall no encontrado")
        result = supabase.table("mall_feature_flags").upsert(
            {
                "mall_id": mall_id,
                "feature_key": normalized_key,
                "enabled": payload.enabled,
                "updated_at": datetime.utcnow().isoformat(),
                "updated_by": admin_ctx["user_id"],
            },
            on_conflict="mall_id,feature_key",
        ).execute()
        return (result.data or [{"feature_key": normalized_key, "enabled": payload.enabled}])[0]
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error actualizando feature flag {normalized_key} para mall {mall_id}: {e}")
        raise HTTPException(status_code=500, detail="No se pudo guardar el módulo del mall")

@app.post("/api/v1/malls")
async def create_mall(mall: MallCreate, admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    """
    Create a new mall.
    """
    try:
        res = supabase.table("malls").insert({
            "nombre": mall.nombre,
            "conf_locale": mall.conf_locale or 'es-CL',
            "conf_moneda": mall.conf_moneda or 'CLP',
            "api_secret_key": str(uuid4()) # Auto-generate key
        }).execute()
        
        if not res.data:
             raise HTTPException(status_code=400, detail="Failed to create mall")
             
        return res.data[0]
    except Exception as e:
        logger.error(f"Error creating mall: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/v1/malls/{mall_id}")
async def update_mall(mall_id: str, mall: MallUpdate, admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    """
    Update a mall.
    """
    try:
        update_data = {k: v for k, v in mall.dict().items() if v is not None and k != 'metadata'}
        if not update_data:
            return {"message": "No data to update"}
            
        res = supabase.table("malls").update(update_data).eq("id", mall_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Mall not found or permission denied")
            
        return res.data[0]
    except Exception as e:
        logger.error(f"Error updating mall: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/v1/malls/{mall_id}")
async def delete_mall(mall_id: str, admin_ctx: Dict[str, Any] = Depends(require_admin_access)):
    """
    Delete a mall.
    """
    try:
        # Check for dependencies first? Or let DB constraint fail?
        # Assuming cascade delete or relying on error if FK exists.
        res = supabase.table("malls").delete().eq("id", mall_id).execute()
        if not res.data: # Note: delete returns deleted data usually
             # If no data returned, it might mean it didn't exist or RLS blocked it.
             # Supabase-py delete behavior can vary on response if empty.
             pass 

        return {"message": "Mall deleted successfully"}
    except Exception as e:
        logger.error(f"Error deleting mall: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- Admin User Management ---

class UserMallAssignment(BaseModel):
    mall_ids: List[str]
    rol: str = 'auditor'
    role_id: Optional[str] = None

class AdminCreateUserRequest(BaseModel):
    email: str
    password: str
    rol: str = 'auditor'
    role_id: Optional[str] = None
    mall_ids: List[str] = []

class AdminUpdateUserRequest(BaseModel):
    email: Optional[str] = None
    nombre: Optional[str] = None
    rol: Optional[str] = None
    role_id: Optional[str] = None
    mall_ids: Optional[List[str]] = None

class RolePermissionRequest(BaseModel):
    module_key: str
    can_view: bool = False
    can_create: bool = False
    can_update: bool = False
    can_delete: bool = False

class RoleRequest(BaseModel):
    key: str
    nombre: str
    descripcion: Optional[str] = None
    permissions: List[RolePermissionRequest] = []

def _validate_role_key(value: str) -> str:
    key = _normalize_role(value)
    if not re.fullmatch(r"[a-z][a-z0-9_]{1,63}", key):
        raise HTTPException(status_code=400, detail="El identificador del rol sólo admite letras, números y guion bajo.")
    return key

def _role_permissions_payload(role_id: str, permissions: List[RolePermissionRequest]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen = set()
    for permission in permissions:
        module_key = _validate_role_key(permission.module_key)
        if module_key in seen:
            raise HTTPException(status_code=400, detail=f"El módulo {module_key} está repetido.")
        seen.add(module_key)
        can_create = bool(permission.can_create)
        can_update = bool(permission.can_update)
        can_delete = bool(permission.can_delete)
        rows.append({
            "role_id": role_id,
            "module_key": module_key,
            "can_view": bool(permission.can_view or can_create or can_update or can_delete),
            "can_create": can_create,
            "can_update": can_update,
            "can_delete": can_delete,
        })
    return rows

async def _list_roles_with_permissions() -> List[Dict[str, Any]]:
    roles = supabase.table("app_roles").select("id,key,nombre,descripcion,is_factory,created_at,updated_at").order("nombre").execute().data or []
    permissions = supabase.table("app_role_permissions").select("role_id,module_key,can_view,can_create,can_update,can_delete").execute().data or []
    by_role: Dict[str, List[Dict[str, Any]]] = {}
    for permission in permissions:
        by_role.setdefault(permission["role_id"], []).append(permission)
    return [{**role, "permissions": by_role.get(role["id"], [])} for role in roles]

@app.get("/api/v1/admin/roles")
async def admin_get_roles(access_ctx: Dict[str, Any] = Depends(require_module_permission("roles", "view"))):
    return await _list_roles_with_permissions()

@app.post("/api/v1/admin/roles")
async def admin_create_role(payload: RoleRequest, access_ctx: Dict[str, Any] = Depends(require_module_permission("roles", "create"))):
    key = _validate_role_key(payload.key)
    if key in {"admin", "it", "auditor", "visualizador"}:
        raise HTTPException(status_code=400, detail="Los roles de fábrica ya existen y no se pueden duplicar.")
    created = supabase.table("app_roles").insert({
        "key": key, "nombre": payload.nombre.strip(), "descripcion": payload.descripcion, "is_factory": False,
    }).execute().data or []
    if not created:
        raise HTTPException(status_code=400, detail="No se pudo crear el rol.")
    rows = _role_permissions_payload(created[0]["id"], payload.permissions)
    if rows:
        supabase.table("app_role_permissions").insert(rows).execute()
    return {"id": created[0]["id"], "message": "Rol creado correctamente."}

@app.put("/api/v1/admin/roles/{role_id}")
async def admin_update_role(role_id: str, payload: RoleRequest, access_ctx: Dict[str, Any] = Depends(require_module_permission("roles", "update"))):
    role = supabase.table("app_roles").select("id,key,is_factory").eq("id", role_id).maybe_single().execute().data
    if not role:
        raise HTTPException(status_code=404, detail="Rol no encontrado.")
    key = _validate_role_key(payload.key)
    if role.get("is_factory") and key != role.get("key"):
        raise HTTPException(status_code=400, detail="No se puede cambiar el identificador de un rol de fábrica.")
    supabase.table("app_roles").update({
        "key": key, "nombre": payload.nombre.strip(), "descripcion": payload.descripcion, "updated_at": datetime.utcnow().isoformat(),
    }).eq("id", role_id).execute()
    supabase.table("app_role_permissions").delete().eq("role_id", role_id).execute()
    rows = _role_permissions_payload(role_id, payload.permissions)
    if rows:
        supabase.table("app_role_permissions").insert(rows).execute()
    return {"message": "Permisos del rol actualizados."}

@app.delete("/api/v1/admin/roles/{role_id}")
async def admin_delete_role(role_id: str, access_ctx: Dict[str, Any] = Depends(require_module_permission("roles", "delete"))):
    role = supabase.table("app_roles").select("is_factory").eq("id", role_id).maybe_single().execute().data
    if not role:
        raise HTTPException(status_code=404, detail="Rol no encontrado.")
    if role.get("is_factory"):
        raise HTTPException(status_code=400, detail="Los roles de fábrica no se eliminan; puedes ajustar o restaurar sus permisos.")
    assigned = supabase.table("profile_role_assignments").select("user_id").eq("role_id", role_id).limit(1).execute().data or []
    if assigned:
        raise HTTPException(status_code=400, detail="No puedes eliminar un rol que tiene usuarios asignados.")
    supabase.table("app_roles").delete().eq("id", role_id).execute()
    return {"message": "Rol eliminado correctamente."}

@app.post("/api/v1/admin/roles/{role_id}/restore-factory")
async def admin_restore_factory_role(role_id: str, access_ctx: Dict[str, Any] = Depends(require_module_permission("roles", "update"))):
    role = supabase.table("app_roles").select("id,key,is_factory").eq("id", role_id).maybe_single().execute().data
    if not role or not role.get("is_factory") or role.get("key") not in FACTORY_ROLE_PERMISSIONS:
        raise HTTPException(status_code=400, detail="Sólo los roles de fábrica pueden restaurarse.")
    permissions = FACTORY_ROLE_PERMISSIONS[role["key"]]
    supabase.table("app_role_permissions").delete().eq("role_id", role_id).execute()
    rows = [
        {"role_id": role_id, "module_key": module, "can_view": bool(actions.get("view")),
         "can_create": bool(actions.get("create")), "can_update": bool(actions.get("update")), "can_delete": bool(actions.get("delete"))}
        for module, actions in permissions.items()
    ]
    if rows:
        supabase.table("app_role_permissions").insert(rows).execute()
    return {"message": "Permisos de fábrica restaurados."}

def _resolve_role_assignment(role_id: Optional[str], legacy_role: Optional[str]) -> Tuple[Optional[str], str]:
    """Returns the RBAC role id plus the compatible legacy role stored in existing tables."""
    role = None
    if role_id:
        role = supabase.table("app_roles").select("id,key").eq("id", role_id).maybe_single().execute().data
    else:
        canonical = _canonical_admin_role(legacy_role) or "auditor"
        role = supabase.table("app_roles").select("id,key").eq("key", canonical).maybe_single().execute().data
    if not role:
        raise HTTPException(status_code=400, detail="El rol seleccionado no existe. Aplica primero la migración de roles.")
    # profiles.role and usuarios_malls.rol are legacy fields; custom roles keep the
    # safe auditor fallback there while profile_role_assignments is authoritative.
    legacy_candidate = _canonical_admin_role(role.get("key"))
    legacy = legacy_candidate if legacy_candidate in {"admin", "it", "auditor"} else "auditor"
    return role["id"], legacy

def _assign_rbac_role(user_id: str, role_id: str, assigned_by: Optional[str]) -> None:
    supabase.table("profile_role_assignments").upsert({
        "user_id": user_id, "role_id": role_id, "assigned_by": assigned_by, "assigned_at": datetime.utcnow().isoformat(),
    }, on_conflict="user_id").execute()

@app.get("/api/v1/admin/users")
async def admin_get_users(admin_ctx: Dict[str, Any] = Depends(require_module_permission("users", "view"))):
    """
    List all users and their assigned malls. Requires ADMIN role.
    """
    try:
        auth_users = _list_all_auth_users()

        users_list = []
        for u in auth_users:
            uid = _user_field(u, "id")
            email = _user_field(u, "email")
            metadata = _user_field(u, "user_metadata", {}) or {}
            metadata_rol = metadata.get("rol") if isinstance(metadata, dict) else None
            metadata_role = metadata.get("role") if isinstance(metadata, dict) else None
            metadata_name = (
                metadata.get("nombre") or metadata.get("full_name")
                if isinstance(metadata, dict)
                else None
            )
            users_list.append({
                "id": uid,
                "email": email,
                "nombre": metadata_name,
                "metadata": metadata,
                "last_sign_in_at": _user_field(u, "last_sign_in_at"),
                "created_at": _user_field(u, "created_at"),
                "_role_candidates": [metadata_rol, metadata_role]
            })
            
        # 2. Get Assignments
        assignments = supabase.table("usuarios_malls").select("*").execute().data or []
        assign_map = {}
        for a in assignments:
            uid = a['usuario_id']
            if uid not in assign_map: assign_map[uid] = []
            assign_map[uid].append(a)

        # 3. Get profile roles in one query
        user_ids = [u["id"] for u in users_list if u.get("id")]
        profile_role_map: Dict[str, str] = {}
        if user_ids:
            profiles = supabase.table("profiles").select("id, role").in_("id", user_ids).execute().data or []
            for p in profiles:
                profile_role_map[p["id"]] = p.get("role")
            
        # 4. Resolve configurable role assignments in one query.
        rbac_role_map: Dict[str, Dict[str, Any]] = {}
        if user_ids:
            role_links = supabase.table("profile_role_assignments").select("user_id,role_id").in_("user_id", user_ids).execute().data or []
            role_ids = [link["role_id"] for link in role_links if link.get("role_id")]
            roles_by_id = {}
            if role_ids:
                role_rows = supabase.table("app_roles").select("id,key,nombre").in_("id", role_ids).execute().data or []
                roles_by_id = {row["id"]: row for row in role_rows}
            rbac_role_map = {link["user_id"]: roles_by_id[link["role_id"]] for link in role_links if link.get("role_id") in roles_by_id}

        # 5. Merge
        result = []
        for u in users_list:
            u['malls'] = assign_map.get(u['id'], [])
            effective_role = _resolve_effective_role(
                u.get("email"),
                [profile_role_map.get(u["id"]), *(u.get("_role_candidates") or [])]
            )
            assigned_role = rbac_role_map.get(u["id"])
            u['rol'] = assigned_role.get("key") if assigned_role else effective_role
            u['role_id'] = assigned_role.get("id") if assigned_role else None
            u['role_name'] = assigned_role.get("nombre") if assigned_role else effective_role.title()
            u.pop("_role_candidates", None)
            result.append(u)
            
        return result
    except Exception as e:
        logger.error(f"Error listing users: {e}")
        # If admin API fails (unsupported), returns empty or error
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/admin/users")
async def admin_create_user(payload: AdminCreateUserRequest, admin_ctx: Dict[str, Any] = Depends(require_module_permission("users", "create"))):
    """
    Create a new auth user and optionally assign malls.
    Requires ADMIN role.
    """
    try:
        role_id, role = _resolve_role_assignment(payload.role_id, payload.rol)

        email = (payload.email or "").strip().lower()
        if not email:
            raise HTTPException(status_code=400, detail="Email requerido.")

        existing_user = _find_auth_user_by_email(email)
        created_user_id = None
        user_previously_existed = existing_user is not None

        if existing_user:
            created_user_id = _user_field(existing_user, "id")
            existing_meta = _user_field(existing_user, "user_metadata", {}) or {}
            if not isinstance(existing_meta, dict):
                existing_meta = {}
            updated_meta = {**existing_meta, "rol": role, "role": role}
            try:
                supabase.auth.admin.update_user_by_id(created_user_id, {
                    "user_metadata": updated_meta
                })
            except Exception as update_err:
                logger.warning(f"No se pudo actualizar metadata de usuario existente {created_user_id}: {update_err}")
        else:
            if len(payload.password or "") < 8:
                raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres.")

            created = supabase.auth.admin.create_user({
                "email": email,
                "password": payload.password,
                "email_confirm": True,
                "user_metadata": {"rol": role, "role": role}
            })
            created_user = _extract_auth_user(created)
            created_user_id = _user_field(created_user, "id")
            if not created_user_id:
                raise HTTPException(status_code=500, detail="No se pudo obtener el ID del usuario creado.")

        if not created_user_id:
            raise HTTPException(status_code=500, detail="No se pudo resolver el ID del usuario.")

        # Keep global role in profiles for frontend role checks.
        try:
            supabase.table("profiles").upsert({"id": created_user_id, "role": role}, on_conflict="id").execute()
        except Exception as p_err:
            logger.warning(f"No se pudo upsert profiles para {created_user_id}: {p_err}")

        _assign_rbac_role(created_user_id, role_id, admin_ctx.get("user_id"))

        # Assign malls if requested.
        supabase.table("usuarios_malls").delete().eq("usuario_id", created_user_id).execute()
        if payload.mall_ids:
            inserts = [{"usuario_id": created_user_id, "mall_id": mid, "rol": role} for mid in payload.mall_ids]
            supabase.table("usuarios_malls").insert(inserts).execute()

        return {
            "id": created_user_id,
            "email": email,
            "rol": role,
            "role_id": role_id,
            "mall_ids": payload.mall_ids,
            "message": "Usuario ya existía; rol y asignaciones actualizados" if user_previously_existed else "Usuario creado correctamente"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating admin user: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/admin/users/{target_user_id}/malls")
async def admin_assign_malls(target_user_id: str, payload: UserMallAssignment, admin_ctx: Dict[str, Any] = Depends(require_module_permission("users", "update"))):
    """
    Assign a list of malls to a user.
    """
    try:
        role_id, role = _resolve_role_assignment(payload.role_id, payload.rol)
        _assign_rbac_role(target_user_id, role_id, admin_ctx.get("user_id"))

        # Transaction? (Not supported natively in HTTP API, do sequentially)
        
        # 1. Delete existing assignments
        supabase.table("usuarios_malls").delete().eq("usuario_id", target_user_id).execute()
        
        # 2. Insert new
        if payload.mall_ids:
            inserts = [{"usuario_id": target_user_id, "mall_id": mid, "rol": role} for mid in payload.mall_ids]
            res = supabase.table("usuarios_malls").insert(inserts).execute()
            
        # Update profile/global role for consistent UI gating.
        try:
            supabase.table("profiles").upsert({"id": target_user_id, "role": role}, on_conflict="id").execute()
        except Exception as p_err:
            logger.warning(f"No se pudo upsert profiles al asignar malls: {p_err}")

        return {"message": "Assignments updated"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error assigning malls: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/v1/admin/users/{target_user_id}")
async def admin_update_user(
    target_user_id: str,
    payload: AdminUpdateUserRequest,
    admin_ctx: Dict[str, Any] = Depends(require_module_permission("users", "update"))
):
    """
    Updates editable user profile fields (email/nombre), role and optional mall assignments.
    """
    try:
        auth_result = supabase.auth.admin.get_user_by_id(target_user_id)
        auth_user = _extract_auth_user(auth_result)
        if not auth_user or not _user_field(auth_user, "id"):
            raise HTTPException(status_code=404, detail="Usuario no encontrado.")

        current_email = (_user_field(auth_user, "email") or "").strip().lower()
        current_metadata = _user_field(auth_user, "user_metadata", {}) or {}
        if not isinstance(current_metadata, dict):
            current_metadata = {}

        role_to_set: Optional[str] = None
        role_assignment_id: Optional[str] = None
        if payload.role_id is not None or payload.rol is not None:
            role_assignment_id, role_to_set = _resolve_role_assignment(payload.role_id, payload.rol)

        email_to_set: Optional[str] = None
        if payload.email is not None:
            email_candidate = (payload.email or "").strip().lower()
            if not email_candidate:
                raise HTTPException(status_code=400, detail="Email requerido.")
            email_owner = _find_auth_user_by_email(email_candidate)
            if email_owner and _user_field(email_owner, "id") != target_user_id:
                raise HTTPException(status_code=409, detail="Ya existe otro usuario con ese email.")
            if email_candidate != current_email:
                email_to_set = email_candidate

        updated_metadata = dict(current_metadata)
        if payload.nombre is not None:
            clean_name = (payload.nombre or "").strip()
            if clean_name:
                updated_metadata["nombre"] = clean_name
                updated_metadata["full_name"] = clean_name
            else:
                updated_metadata.pop("nombre", None)
                updated_metadata.pop("full_name", None)

        if role_to_set:
            updated_metadata["rol"] = role_to_set
            updated_metadata["role"] = role_to_set

        update_payload: Dict[str, Any] = {"user_metadata": updated_metadata}
        if email_to_set:
            update_payload["email"] = email_to_set

        supabase.auth.admin.update_user_by_id(target_user_id, update_payload)

        effective_email = email_to_set or current_email
        if role_to_set:
            try:
                supabase.table("profiles").upsert({"id": target_user_id, "role": role_to_set}, on_conflict="id").execute()
            except Exception as p_err:
                logger.warning(f"No se pudo upsert profiles al actualizar usuario {target_user_id}: {p_err}")

            _assign_rbac_role(target_user_id, role_assignment_id, admin_ctx.get("user_id"))

            # Keep existing assignments synchronized to the new role when malls are not being replaced.
            if payload.mall_ids is None:
                try:
                    supabase.table("usuarios_malls").update({"rol": role_to_set}).eq("usuario_id", target_user_id).execute()
                except Exception as um_err:
                    logger.warning(f"No se pudo actualizar rol en usuarios_malls para {target_user_id}: {um_err}")

        if payload.mall_ids is not None:
            # If role not provided in this request, infer current effective role for mall assignments.
            role_for_malls = role_to_set
            if not role_for_malls:
                existing_profile_role = None
                try:
                    profile = supabase.table("profiles").select("role").eq("id", target_user_id).maybe_single().execute()
                    if profile and profile.data:
                        existing_profile_role = profile.data.get("role")
                except Exception:
                    existing_profile_role = None

                existing_meta_role = (
                    updated_metadata.get("rol")
                    or updated_metadata.get("role")
                    or current_metadata.get("rol")
                    or current_metadata.get("role")
                )
                role_for_malls = _resolve_effective_role(effective_email, [existing_profile_role, existing_meta_role])

            role_for_malls = _canonical_admin_role(role_for_malls) or "auditor"

            supabase.table("usuarios_malls").delete().eq("usuario_id", target_user_id).execute()
            if payload.mall_ids:
                inserts = [
                    {"usuario_id": target_user_id, "mall_id": mall_id, "rol": role_for_malls}
                    for mall_id in payload.mall_ids
                ]
                supabase.table("usuarios_malls").insert(inserts).execute()

        return {
            "id": target_user_id,
            "email": email_to_set or current_email,
            "nombre": updated_metadata.get("nombre") or updated_metadata.get("full_name"),
            "rol": role_to_set or _resolve_effective_role(
                effective_email,
                [updated_metadata.get("rol"), updated_metadata.get("role")]
            ),
            "role_id": role_assignment_id,
            "message": "Usuario actualizado correctamente."
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating user {target_user_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _security_token_service():
    return build_token_auth_service()


def _security_allowed_mall_ids(operator_ctx: Dict[str, Any]) -> Optional[set]:
    if operator_ctx.get("role") == "admin":
        return None
    return set(_get_user_mall_ids(operator_ctx.get("user_id")))


def _security_filter_rows_by_mall_access(rows: List[Dict[str, Any]], operator_ctx: Dict[str, Any]) -> List[Dict[str, Any]]:
    allowed = _security_allowed_mall_ids(operator_ctx)
    if allowed is None:
        return rows
    return [row for row in rows if row.get("mall_id") in allowed]


def _security_ensure_row_access(operator_ctx: Dict[str, Any], row: Optional[Dict[str, Any]], not_found_detail: str):
    if not row:
        raise HTTPException(status_code=404, detail=not_found_detail)
    _ensure_operator_can_access_mall(operator_ctx, row.get("mall_id"))
    return row


def _security_validate_local_alignment(local_id: Optional[str], mall_id: str, operator_ctx: Dict[str, Any]) -> None:
    if not local_id:
        return
    local_cfg = _load_local_config_with_access(local_id, operator_ctx)
    local_mall_id = str(local_cfg.get("mall_id") or "")
    if local_mall_id and str(mall_id) != local_mall_id:
        raise HTTPException(status_code=400, detail="local_id no pertenece al mall_id indicado")


def _security_text_search(rows: List[Dict[str, Any]], q: Optional[str], fields: List[str]) -> List[Dict[str, Any]]:
    term = (q or "").strip().lower()
    if not term:
        return rows
    out: List[Dict[str, Any]] = []
    for row in rows:
        haystack = " ".join(str(row.get(field) or "") for field in fields).lower()
        if term in haystack:
            out.append(row)
    return out


@app.get("/api/v1/security/service-accounts")
async def security_list_service_accounts(
    mall_id: Optional[str] = None,
    local_id: Optional[str] = None,
    token_type: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    q: Optional[str] = None,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    rows = svc.store.list_service_accounts({
        "mall_id": mall_id,
        "local_id": local_id,
        "token_type": token_type,
        "status": status_filter,
    })
    rows = _security_filter_rows_by_mall_access(rows, operator_ctx)
    rows = _security_text_search(rows, q, ["name", "client_id", "created_by", "mall_id", "local_id"])

    tokens = svc.store.list_tokens({
        "mall_id": mall_id,
        "local_id": local_id,
        "token_type": TOKEN_TYPE_EXPORTER,
    })
    tokens = _security_filter_rows_by_mall_access(tokens, operator_ctx)
    usage_by_sa: Dict[str, Dict[str, Any]] = {}
    for token in tokens:
        sa_id = token.get("service_account_id")
        if not sa_id:
            continue
        usage = usage_by_sa.setdefault(sa_id, {
            "last_used_at": None,
            "last_used_ip": None,
            "last_used_ua": None,
            "active_tokens": 0,
            "total_tokens": 0,
        })
        usage["total_tokens"] += 1
        if token.get("status") == TOKEN_ACTIVE:
            usage["active_tokens"] += 1
        last_used_at = token.get("last_used_at")
        if last_used_at and (not usage["last_used_at"] or str(last_used_at) > str(usage["last_used_at"])):
            usage["last_used_at"] = last_used_at
            usage["last_used_ip"] = token.get("last_used_ip")
            usage["last_used_ua"] = token.get("last_used_ua")

    out: List[Dict[str, Any]] = []
    for row in rows:
        safe = sanitize_token_service_account_row(row) or {}
        usage = usage_by_sa.get(safe.get("id"), {})
        safe.update({
            "last_used_at": usage.get("last_used_at"),
            "last_used_ip": usage.get("last_used_ip"),
            "last_used_ua": usage.get("last_used_ua"),
            "active_tokens": usage.get("active_tokens", 0),
            "total_tokens": usage.get("total_tokens", 0),
        })
        out.append(safe)
    return out


@app.post("/api/v1/security/service-accounts")
async def security_create_service_account(
    payload: TokenCreateServiceAccountRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if payload.token_type == TOKEN_TYPE_EXPORTER and not payload.local_id:
        raise HTTPException(status_code=400, detail="local_id requerido para exporter")
    if not (payload.name or "").strip():
        raise HTTPException(status_code=400, detail="name es requerido")
    _ensure_operator_can_access_mall(operator_ctx, payload.mall_id)
    _security_validate_local_alignment(payload.local_id, payload.mall_id, operator_ctx)

    svc = _security_token_service()
    client_id = f"msa_{secrets.token_hex(8)}"
    client_secret = secrets.token_urlsafe(32)
    row = svc.store.create_service_account({
        "name": payload.name.strip(),
        "mall_id": payload.mall_id,
        "local_id": payload.local_id,
        "token_type": payload.token_type,
        "client_id": client_id,
        "client_secret_hash": token_auth_hash_token(client_secret),
        "scopes": token_auth_parse_scopes(payload.scopes),
        "status": TOKEN_ACTIVE,
        "created_by": operator_ctx.get("user_id"),
        "created_at": token_auth_utcnow().isoformat(),
        "updated_at": token_auth_utcnow().isoformat(),
    })
    safe = sanitize_token_service_account_row(row) or {}
    safe["client_secret"] = client_secret
    safe["warning"] = "Este secreto no volverá a mostrarse completo."
    return safe


@app.patch("/api/v1/security/service-accounts/{service_account_id}/status")
async def security_patch_service_account_status(
    service_account_id: str,
    payload: TokenPatchServiceAccountStatusRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    base = _security_ensure_row_access(
        operator_ctx,
        svc.store.get_service_account(service_account_id),
        "Service account no encontrado",
    )
    row = svc.store.update_service_account(base["id"], {"status": payload.status, "updated_at": token_auth_utcnow().isoformat()})
    return sanitize_token_service_account_row(row)


@app.post("/api/v1/security/service-accounts/{service_account_id}/regenerate")
async def security_regenerate_service_account(
    service_account_id: str,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    base = _security_ensure_row_access(
        operator_ctx,
        svc.store.get_service_account(service_account_id),
        "Service account no encontrado",
    )
    client_secret = secrets.token_urlsafe(32)
    row = svc.store.update_service_account(base["id"], {
        "client_secret_hash": token_auth_hash_token(client_secret),
        "status": TOKEN_ACTIVE,
        "updated_at": token_auth_utcnow().isoformat(),
    })
    svc.store.revoke_tokens_by_service_account(base["id"], revoked_by=operator_ctx.get("user_id"), reason="service_account_secret_regenerated")
    safe = sanitize_token_service_account_row(row) or {}
    safe["client_secret"] = client_secret
    safe["warning"] = "Este secreto no volverá a mostrarse completo."
    return safe


@app.post("/api/v1/security/service-accounts/{service_account_id}/revoke-tokens")
async def security_revoke_service_account_tokens(
    service_account_id: str,
    payload: TokenRevokeServiceAccountTokensRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    base = _security_ensure_row_access(
        operator_ctx,
        svc.store.get_service_account(service_account_id),
        "Service account no encontrado",
    )
    count = svc.store.revoke_tokens_by_service_account(base["id"], revoked_by=operator_ctx.get("user_id"), reason=payload.reason)
    return {"revoked_count": count, "service_account_id": base["id"]}


@app.get("/api/v1/security/tokens")
async def security_list_tokens(
    mall_id: Optional[str] = None,
    local_id: Optional[str] = None,
    token_type: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    q: Optional[str] = None,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    rows = svc.store.list_tokens({
        "mall_id": mall_id,
        "local_id": local_id,
        "token_type": token_type,
        "status": status_filter,
    })
    rows = _security_filter_rows_by_mall_access(rows, operator_ctx)
    rows = _security_text_search(rows, q, ["jti", "created_by", "mall_id", "local_id", "token_type", "status"])
    return [sanitize_token_auth_row(row) for row in rows]


@app.post("/api/v1/security/tokens")
async def security_create_token(
    payload: TokenCreateTokenRequest,
    request: Request,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if payload.token_type == TOKEN_TYPE_EXPORTER and not payload.local_id:
        raise HTTPException(status_code=400, detail="local_id requerido para exporter")
    _ensure_operator_can_access_mall(operator_ctx, payload.mall_id)
    _security_validate_local_alignment(payload.local_id, payload.mall_id, operator_ctx)
    scopes = token_auth_parse_scopes(payload.scopes)
    if not scopes:
        raise HTTPException(status_code=400, detail="scopes requeridos")

    svc = _security_token_service()
    if payload.service_account_id:
        sa = _security_ensure_row_access(operator_ctx, svc.store.get_service_account(payload.service_account_id), "Service account no encontrado")
        if sa.get("mall_id") != payload.mall_id or sa.get("local_id") != payload.local_id:
            raise HTTPException(status_code=400, detail="service_account_id no coincide con mall_id/local_id")

    return svc._issue_pair(
        mall_id=payload.mall_id,
        local_id=payload.local_id,
        token_type=payload.token_type,
        scopes=scopes,
        created_by=operator_ctx.get("user_id"),
        service_account_id=payload.service_account_id,
        request=request,
        access_ttl_seconds=payload.expires_in,
        access_never_expires=token_auth_request_explicit_never_expires(payload),
    )


@app.patch("/api/v1/security/tokens/{token_id}/status")
async def security_patch_token_status(
    token_id: str,
    payload: TokenPatchTokenStatusRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    base = _security_ensure_row_access(operator_ctx, svc.store.get_token_by_id(token_id), "Token no encontrado")
    row = svc.store.update_api_token(base["id"], {"status": payload.status, "updated_at": token_auth_utcnow().isoformat()})
    return sanitize_token_auth_row(row)


@app.post("/api/v1/security/tokens/{token_id}/regenerate")
async def security_regenerate_token(
    token_id: str,
    request: Request,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    base = _security_ensure_row_access(operator_ctx, svc.store.get_token_by_id(token_id), "Token no encontrado")
    svc.store.update_api_token(base["id"], {
        "status": TOKEN_REVOKED,
        "revoked_at": token_auth_utcnow().isoformat(),
        "revoked_by": operator_ctx.get("user_id"),
        "revoke_reason": "regenerated",
        "updated_at": token_auth_utcnow().isoformat(),
    })
    return svc._issue_pair(
        mall_id=base["mall_id"],
        local_id=base.get("local_id"),
        token_type=base["token_type"],
        scopes=token_auth_parse_scopes(base.get("scopes")),
        created_by=operator_ctx.get("user_id"),
        service_account_id=base.get("service_account_id"),
        request=request,
    )


@app.post("/api/v1/security/tokens/revoke")
async def security_revoke_token(
    payload: TokenRevokeRequest,
    request: Request,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    svc = _security_token_service()
    target = None
    if payload.token_id:
        target = svc.store.get_token_by_id(payload.token_id)
    elif payload.jti:
        target = svc.store.get_token_by_jti(payload.jti)
    _security_ensure_row_access(operator_ctx, target, "Token no encontrado")
    return svc.revoke(
        token_id=payload.token_id,
        jti=payload.jti,
        actor=operator_ctx.get("user_id"),
        reason=payload.reason,
        current_ctx=None,
        request=request,
    )


@app.post("/api/v1/security/tokens/revoke/local")
async def security_revoke_tokens_by_local(
    payload: TokenRevokeLocalRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    _ensure_operator_can_access_mall(operator_ctx, payload.mall_id)
    _security_validate_local_alignment(payload.local_id, payload.mall_id, operator_ctx)
    svc = _security_token_service()
    count = svc.store.revoke_tokens_by_scope(
        mall_id=payload.mall_id,
        local_id=payload.local_id,
        revoked_by=operator_ctx.get("user_id"),
        reason=payload.reason,
    )
    return {"revoked_count": count}


@app.post("/api/v1/security/tokens/revoke/mall")
async def security_revoke_tokens_by_mall(
    payload: TokenRevokeMallRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    _ensure_operator_can_access_mall(operator_ctx, payload.mall_id)
    svc = _security_token_service()
    count = svc.store.revoke_tokens_by_scope(
        mall_id=payload.mall_id,
        revoked_by=operator_ctx.get("user_id"),
        reason=payload.reason,
    )
    return {"revoked_count": count}


@app.get("/api/v1/security/token-audit")
async def security_list_token_audit(
    mall_id: Optional[str] = None,
    local_id: Optional[str] = None,
    event_type: Optional[str] = None,
    token_id: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = Query(200, ge=1, le=1000),
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if mall_id:
        _ensure_operator_can_access_mall(operator_ctx, mall_id)
    elif operator_ctx.get("role") != "admin":
        # For IT users without explicit mall filter, restrict to their first allowed mall set in-memory after query.
        pass

    svc = _security_token_service()
    rows = svc.store.list_audit_logs({
        "mall_id": mall_id,
        "local_id": local_id,
        "event_type": event_type,
        "token_id": token_id,
    }, limit=limit)
    rows = _security_filter_rows_by_mall_access(rows, operator_ctx)
    rows = _security_text_search(rows, q, ["event_type", "mall_id", "local_id", "ip", "ua", "token_id"])
    return rows

@app.get("/api/v1/security/exporter/configs")
async def security_list_exporter_webservice_configs(
    mall_id: Optional[str] = None,
    local_id: Optional[str] = None,
    enabled: Optional[bool] = None,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    if mall_id:
        _ensure_operator_can_access_mall(operator_ctx, mall_id)
        if local_id:
            _security_validate_local_alignment(local_id, mall_id, operator_ctx)
    elif local_id:
        local_cfg = _load_local_config_with_access(local_id, operator_ctx)
        mall_id = str(local_cfg.get("mall_id") or "") or None

    svc = _security_token_service()
    lister = getattr(svc.store, "list_exporter_webservice_configs", None)
    if not callable(lister):
        raise HTTPException(status_code=500, detail="Store no soporta configuracion exporter webservice")

    rows = lister({"mall_id": mall_id, "local_id": local_id, "enabled": enabled})
    rows = _security_filter_rows_by_mall_access(rows, operator_ctx)
    return [sanitize_token_exporter_webservice_config_row(row) for row in rows]


@app.get("/api/v1/security/exporter/configs/{local_id}")
async def security_get_exporter_webservice_config(
    local_id: str,
    mall_id: str,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    _ensure_operator_can_access_mall(operator_ctx, mall_id)
    _security_validate_local_alignment(local_id, mall_id, operator_ctx)

    svc = _security_token_service()
    getter = getattr(svc.store, "get_exporter_webservice_config", None)
    if not callable(getter):
        raise HTTPException(status_code=500, detail="Store no soporta configuracion exporter webservice")

    row = getter(mall_id, local_id)
    if not row:
        raise HTTPException(status_code=404, detail="Configuracion exporter webservice no encontrada")
    return sanitize_token_exporter_webservice_config_row(row)


@app.put("/api/v1/security/exporter/configs/{local_id}")
async def security_put_exporter_webservice_config(
    local_id: str,
    payload: TokenUpsertExporterWebserviceConfigRequest,
    operator_ctx: Dict[str, Any] = Depends(require_it_or_admin_access),
):
    _ensure_operator_can_access_mall(operator_ctx, payload.mall_id)
    _security_validate_local_alignment(local_id, payload.mall_id, operator_ctx)

    svc = _security_token_service()
    upserter = getattr(svc.store, "upsert_exporter_webservice_config", None)
    if not callable(upserter):
        raise HTTPException(status_code=500, detail="Store no soporta configuracion exporter webservice")

    granularity = str(payload.default_granularity or "transaction").strip().lower()
    if granularity == "daily_summary":
        granularity = "daily"

    try:
        row = upserter({
            "mall_id": payload.mall_id,
            "local_id": local_id,
            "enabled": payload.enabled,
            "contract_type": payload.contract_type,
            "default_granularity": granularity,
            "allow_transaction": payload.allow_transaction,
            "allow_daily": payload.allow_daily,
            "strict_validation": payload.strict_validation,
            "notes": payload.notes.strip() if payload.notes else None,
            "updated_by": operator_ctx.get("user_id"),
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            "Error guardando exporter webservice config mall=%s local=%s user=%s: %s",
            payload.mall_id,
            local_id,
            operator_ctx.get("user_id"),
            e,
        )
        detail = str(e).strip() or "No se pudo guardar la configuracion exporter webservice"
        raise HTTPException(status_code=500, detail=detail)
    if not row:
        raise HTTPException(status_code=500, detail="No se pudo guardar la configuracion exporter webservice")
    return sanitize_token_exporter_webservice_config_row(row)


@router_export.get("/sales-report/excel")
async def export_sales_report_excel(
    fecha_inicio: str, 
    fecha_fin: str, 
    local_id: Optional[str] = None, 
    type: str = 'detailed',
    current_mall: str = Depends(get_current_mall)
):
    try:
        if type not in ['detailed', 'summary', 'missing_days']:
            type = 'detailed'
        if type == 'missing_days':
            data = await export_service.generate_missing_days_report_excel(
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                mall_id=current_mall,
                local_id=local_id,
            )
        else:
            data = await export_service.generate_sales_report_excel(
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                local_id=local_id,
                report_type=type,
                mall_id=current_mall,
            )
        return StreamingResponse(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_ventas_{type}_{fecha_inicio}_{fecha_fin}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error exporting excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router_export.get("/sales-report/pdf")
async def export_sales_report_pdf(
    fecha_inicio: str, 
    fecha_fin: str, 
    local_id: Optional[str] = None, 
    type: str = 'detailed',
    current_mall: str = Depends(get_current_mall)
):
    try:
        if type not in ['detailed', 'summary', 'missing_days']:
            type = 'detailed'
        
        # Fetch Mall Name
        mall_name = "MS MALL"
        try:
             m_res = supabase.table("malls").select("nombre").eq("id", current_mall).single().execute()
             if m_res.data:
                 mall_name = m_res.data['nombre']
        except Exception:
            logger.warning(f"Could not fetch mall name for {current_mall}, using default.")

        logger.info(f"Exporting PDF for Mall: {mall_name} ({current_mall}) type={type}")
        if type == 'missing_days':
            data = await export_service.generate_missing_days_report_pdf(
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                mall_id=current_mall,
                local_id=local_id,
                mall_name=mall_name,
            )
        else:
            data = await export_service.generate_sales_report_pdf(
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                local_id=local_id,
                report_type=type,
                mall_name=mall_name,
                mall_id=current_mall,
            )
        return StreamingResponse(
            data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=reporte_ventas_{type}_{fecha_inicio}_{fecha_fin}.pdf"}
        )
    except Exception as e:
        logger.error(f"Error exporting pdf: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router_export.get("/sales-cube/excel")
async def export_sales_cube_excel(
    fecha_inicio: str,
    fecha_fin: str,
    agrupacion: str = "dia",
    metrica: str = "total_neto",
    local_id: Optional[str] = None,
    custom_dimension_key: Optional[str] = None,
    custom_filters: Optional[str] = None,
    current_mall: str = Depends(get_current_mall)
):
    try:
        parsed_custom_filters = json.loads(custom_filters) if custom_filters else None
        data = await export_service.generate_sales_cube_excel(
            fecha_inicio,
            fecha_fin,
            agrupacion,
            metrica,
            mall_id=current_mall,
            local_id=local_id,
            custom_dimension_key=custom_dimension_key,
            custom_filters=parsed_custom_filters,
        )
        return StreamingResponse(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=matriz_ventas_{fecha_inicio}_{fecha_fin}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error exporting cube excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router_export.get("/financial-dashboard/excel")
async def export_financial_dashboard_excel(fecha_inicio: str, fecha_fin: str):
    try:
        data = await export_service.generate_financial_dashboard_excel(fecha_inicio, fecha_fin)
        return StreamingResponse(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=salud_cartera_{fecha_inicio}_{fecha_fin}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error exporting financial excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router_export.get("/financial-dashboard/pdf")
async def export_financial_dashboard_pdf(fecha_inicio: str, fecha_fin: str):
    try:
        data = await export_service.generate_financial_dashboard_pdf(fecha_inicio, fecha_fin)
        return StreamingResponse(
            data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=salud_cartera_{fecha_inicio}_{fecha_fin}.pdf"}
        )
    except Exception as e:
        logger.error(f"Error exporting financial pdf: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/auditoria/brechas-ventas")
async def get_sales_gaps(
    local_id: Optional[str], 
    fecha_inicio: str, 
    fecha_fin: str,
    current_mall: str = Depends(get_current_mall)
):
    try:
        def _normalize_sales_date(raw_value: Any) -> Optional[str]:
            """Normalize DB date/timestamp values to YYYY-MM-DD for day-level comparisons."""
            if raw_value is None:
                return None
            if isinstance(raw_value, datetime):
                return raw_value.strftime('%Y-%m-%d')

            value = str(raw_value).strip()
            if not value:
                return None

            # Fast path for ISO-like values: 2026-02-03 or 2026-02-03T...
            if len(value) >= 10 and value[4] == '-' and value[7] == '-':
                return value[:10]

            try:
                parsed = pd.to_datetime(value, errors='coerce')
                if pd.isna(parsed):
                    return None
                return parsed.strftime('%Y-%m-%d')
            except Exception:
                return None

        def _load_actual_dates_for_local(target_local_id: str) -> Set[str]:
            rows: List[Dict[str, Any]] = []
            page_size = 2000
            page = 0
            while True:
                chunk = (
                    supabase.table('ventas')
                    .select('id, fecha')
                    .eq('local_id', target_local_id)
                    .gte('fecha', fecha_inicio)
                    .lte('fecha', fecha_fin)
                    .order('id')
                    .range(page * page_size, (page + 1) * page_size - 1)
                    .execute()
                ).data or []
                if not chunk:
                    break
                rows.extend(chunk)
                if len(chunk) < page_size:
                    break
                page += 1

            actual_dates: Set[str] = set()
            for row in rows:
                normalized_date = _normalize_sales_date(row.get('fecha'))
                if normalized_date:
                    actual_dates.add(normalized_date)
            return actual_dates

        # 1. Calendario Ideal
        start_date = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        end_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
        total_days = (end_date - start_date).days + 1
        expected_dates = { (start_date + timedelta(days=x)).strftime('%Y-%m-%d') for x in range(total_days) }
        
        # --- MODO GLOBAL (Matrix View) ---
        if not local_id or local_id == 'null' or local_id == 'ALL':
            logger.info(f"Auditing Global Gaps for Mall: {current_mall}")
            
            # Obtener TODOS los locales (filtrar por mall si tuviéramos tabla usuarios_malls poblada y lógica RLS)
            # Como aún no tenemos RLS activo en 'locales' para filtrar por mall_id automágicamente,
            # DEBERÍAMOS filtrar aquí manualmente usando current_mall.
            # Pero la tabla 'locales' ya tiene 'mall_id'.
            
            # Obtener todos los locales DEL MALL ACTUAL
            # Para la Fase 1: asumo que el endpoint necesita ver SOLO los locales del mall.
            # Si el backend lee 'locales', supongo que debo filtrar.
            
            # stores_resp = supabase.table('locales').select('id, nombre, rubro').execute()
            # CHANGE: Filter by current_mall
            # But wait, current_mall comes from DB or Header.
            # If migrating, current_mall might be a Mall ID (UUID).
            
            # stores_resp = supabase.table('locales').select('id, nombre, rubro').eq('mall_id', current_mall).execute()
            # If current_mall is reliable.
            
            # Since I am "migrating", I should probably use the filter.
            # But `locales` table has `mall_id`.
            
            stores_resp = supabase.table('locales').select('id, nombre, rubro').eq('mall_id', current_mall).execute()
            stores = stores_resp.data or []
            
            global_summary = []
            
            for store in stores:
                sid = str(store['id'])
                s_actual = _load_actual_dates_for_local(sid)
                
                missing = sorted(list(expected_dates - s_actual))
                count_missing = len(missing)
                compliance = ((total_days - count_missing) / total_days) * 100
                
                # Definir estado
                status = 'Completo'
                if count_missing > 5: status = 'Crítico'
                elif count_missing > 0: status = 'Alerta'
                
                global_summary.append({
                    "local_id": sid,
                    "nombre": store['nombre'],
                    "rubro": store.get('rubro', 'General'),
                    "dias_faltantes_count": count_missing,
                    "dias_totales_periodo": total_days,
                    "porcentaje_cumplimiento": round(compliance, 1),
                    "estado": status,
                    "lista_dias": missing # Para visualización rápida (heatmap)
                })
            
            # Ordenar por criticidad (más faltantes primero)
            global_summary.sort(key=lambda x: x['dias_faltantes_count'], reverse=True)
            
            return {
                "modo": "global",
                "resumen": global_summary
            }

        # --- MODO INDIVIDUAL (Detailed View) ---
        # 2. Calendario Real (Individual)
        actual_dates = _load_actual_dates_for_local(local_id)
        
        # 3. Brechas
        missing_dates = sorted(list(expected_dates - actual_dates))
        
        # 4. Enriquecimiento con Logs (logs_carga)
        local_resp = supabase.table('locales').select('nombre, mall_id').eq('id', local_id).single().execute()
        local_name = local_resp.data['nombre'] if local_resp.data else None
        local_mall_id = local_resp.data.get('mall_id') if local_resp.data else None
        
        audit_details = []
        if missing_dates:
            # Optimización: Consultar logs por local_id (tenant-safe) y fallback legacy por nombre+mall.
            try:
                logs_resp = supabase.table('logs_carga').select('*')\
                    .eq('local_id', local_id)\
                    .gte('fecha_hora', f"{fecha_inicio}T00:00:00")\
                    .lte('fecha_hora', f"{fecha_fin}T23:59:59")\
                    .order('fecha_hora', desc=True)\
                    .execute()
            except Exception:
                logs_resp = type("Tmp", (), {"data": []})()

            if (not logs_resp.data) and local_name:
                legacy_q = supabase.table('logs_carga').select('*')\
                    .eq('local_nombre', local_name)\
                    .gte('fecha_hora', f"{fecha_inicio}T00:00:00")\
                    .lte('fecha_hora', f"{fecha_fin}T23:59:59")\
                    .order('fecha_hora', desc=True)
                if local_mall_id:
                    legacy_q = legacy_q.eq('mall_id', local_mall_id)
                logs_resp = legacy_q.execute()
            
            logs_df = pd.DataFrame(logs_resp.data)
            if not logs_df.empty:
                logs_df['fecha_log'] = logs_df['fecha_hora'].apply(lambda x: x.split('T')[0] if x else None)
            
            for m_date in missing_dates:
                cause = "Proceso no ejecutado / Sin conexión"
                log_id = None
                
                if not logs_df.empty:
                    day_logs = logs_df[logs_df['fecha_log'] == m_date]
                    if not day_logs.empty:
                        last_log = day_logs.iloc[0]
                        log_id = last_log.get('id')
                        status = str(last_log.get('estado') or '').strip().lower()
                        if status == 'error':
                            cause = "Fallo Técnico / Error de Lectura"
                        elif status in {'no_encontrado', 'no encontrado'}:
                            cause = "Archivo no disponible en FTP"
                        elif status in {'exito', 'éxito', 'success', 'parcial'}:
                            cause = "Procesado con Éxito (Posible archivo vacío)"
                            
                audit_details.append({
                    "fecha": m_date,
                    "causa": cause,
                    "log_id": log_id
                })
        else:
             for m_date in missing_dates:
                audit_details.append({
                    "fecha": m_date,
                    "causa": "Proceso no ejecutado / Sin logs disponibles",
                    "log_id": None
                })

        return {
            "modo": "individual",
            "total_dias_faltantes": len(missing_dates),
            "detalle": audit_details
        }
        
    except Exception as e:
        logger.error(f"Error auditing gaps: {e}")
        raise HTTPException(status_code=500, detail=str(e))


app.include_router(router_export)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
