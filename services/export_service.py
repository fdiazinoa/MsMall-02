"""
Servicio de exportación de reportes
Contiene toda la lógica de generación de archivos Excel y PDF
"""

import io
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from supabase import Client
from services.local_custom_fields_service import LocalCustomFieldsService
from services.sales_gap_service import (
    SALES_PAGE_SIZE,
    expected_sales_dates,
    normalize_sales_date,
)
from services.sales_query_service import fetch_sales_rows_keyset
from services.sales_cube_query_service import fetch_sales_cube_daily_aggregates

class ExportService:
    def __init__(self, supabase_client: Client):
        self.supabase = supabase_client
        self.local_custom_fields_service = LocalCustomFieldsService(supabase_client, logger=None)
    
    # --- UTILS ---
    def _get_header_style(self):
        return PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid"), Font(bold=True, color="FFFFFF", size=11)
    
    def _format_currency(self, value):
        if value is None: return "$0.00"
        return f"${value:,.2f}"

    def _normalize_sale_totals_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        bruto = float(row.get('total_bruto') or 0)
        impuestos = float(row.get('total_impuestos') or 0) if row.get('total_impuestos') is not None else 0.0
        neto = float(row.get('total_neto') or 0)

        eps = 0.05
        as_is_delta = abs(neto - (bruto + impuestos))
        swapped_delta = abs(bruto - (neto + impuestos))
        if swapped_delta + eps < as_is_delta:
            row['total_bruto'] = neto
            row['total_neto'] = bruto
        return row

    def _get_pdf_styles(self):
        styles = getSampleStyleSheet()
        title = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1F4788'), spaceAfter=12, alignment=TA_CENTER, fontName='Helvetica-Bold')
        subtitle = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=12, textColor=colors.grey, spaceAfter=20, alignment=TA_CENTER)
        header = ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#1F4788'), spaceBefore=12, spaceAfter=6, fontName='Helvetica-Bold')
        return title, subtitle, header, styles['Normal']

    def _fetch_sales_rows_paginated(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        select_fields: str,
        local_id: Optional[str] = None,
        mall_id: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        return fetch_sales_rows_keyset(
            self.supabase,
            select_fields=select_fields,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            local_id=local_id,
            mall_id=mall_id,
        )

    def _build_missing_days_dataset(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        mall_id: str,
        local_id: Optional[str] = None
    ) -> Dict[str, Any]:
        start_date = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        end_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
        total_days = (end_date - start_date).days + 1
        expected_dates = expected_sales_dates(fecha_inicio, fecha_fin)

        stores_query = self.supabase.table('locales').select('id, nombre, rubro, activo').eq('mall_id', mall_id)
        if local_id:
            stores_query = stores_query.eq('id', local_id)
        stores = [row for row in (stores_query.execute().data or []) if row.get('activo') is not False]
        store_ids = [str(s['id']) for s in stores if s.get('id')]

        sales_rows: List[Dict[str, Any]] = []
        if store_ids:
            sales_rows = fetch_sales_rows_keyset(
                self.supabase,
                select_fields='local_id,fecha',
                local_ids=store_ids,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                page_size=SALES_PAGE_SIZE,
            )

        sales_df = pd.DataFrame(sales_rows)
        if not sales_df.empty:
            sales_df['local_id_norm'] = sales_df['local_id'].astype(str)
            sales_df['fecha_norm'] = sales_df['fecha'].apply(normalize_sales_date)

        summary_rows: List[Dict[str, Any]] = []
        detail_rows: List[Dict[str, Any]] = []

        for store in stores:
            sid = str(store['id'])
            if not sales_df.empty:
                actual_dates = set(
                    sales_df[sales_df['local_id_norm'] == sid]['fecha_norm']
                    .dropna()
                    .unique()
                )
            else:
                actual_dates = set()

            missing_dates = sorted(list(expected_dates - actual_dates))
            missing_count = len(missing_dates)
            compliance = ((total_days - missing_count) / total_days) * 100 if total_days > 0 else 100.0

            if missing_count == 0:
                continue  # "solo dias faltantes"

            status = 'Completo'
            if missing_count > 5:
                status = 'Crítico'
            elif missing_count > 0:
                status = 'Alerta'

            summary_rows.append({
                'local_id': sid,
                'local_nombre': store.get('nombre') or sid,
                'rubro': store.get('rubro') or 'General',
                'dias_faltantes_count': missing_count,
                'dias_totales_periodo': total_days,
                'porcentaje_cumplimiento': round(compliance, 1),
                'estado': status,
                'lista_dias': missing_dates,
            })

            for missing_date in missing_dates:
                detail_rows.append({
                    'local_id': sid,
                    'local_nombre': store.get('nombre') or sid,
                    'fecha_faltante': missing_date,
                })

        summary_rows.sort(key=lambda x: (x['dias_faltantes_count'], x['local_nombre']), reverse=True)
        detail_rows.sort(key=lambda x: (x['local_nombre'], x['fecha_faltante']))

        return {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'mall_id': mall_id,
            'local_id': local_id,
            'total_days': total_days,
            'summary_rows': summary_rows,
            'detail_rows': detail_rows,
            'is_local_mode': bool(local_id),
            'selected_local_name': (stores[0].get('nombre') if local_id and stores else None),
        }

    # --- SALES REPORT ---

    async def generate_sales_report_excel(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        local_id: Optional[str] = None,
        report_type: str = 'detailed',
        mall_id: Optional[str] = None,
    ) -> io.BytesIO:
        data = self._fetch_sales_rows_paginated(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            select_fields='local_id, locales(nombre), fecha, hora, factura_no, total_neto, total_impuestos, total_bruto',
            local_id=local_id,
            mall_id=mall_id,
        )

        df = pd.DataFrame(data)
        # Flatten locales(nombre)
        if not df.empty:
            df['nombre_local'] = df['locales'].apply(lambda x: x.get('nombre') if x else 'Desc')
        else:
            df = pd.DataFrame(columns=['local_id', 'nombre_local', 'total_neto', 'total_impuestos', 'total_bruto', 'fecha', 'hora', 'factura_no'])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        ws['A1'] = 'REPORTE DE AUDITORÍA DE VENTAS'
        ws['A2'] = f'Período: {fecha_inicio} al {fecha_fin} ({report_type.capitalize()})'
        ws['A1'].font = Font(size=14, bold=True)
        
        fill, font = self._get_header_style()
        
        if report_type == 'detailed':
            # --- VISTA DETALLADA (Filas Individuales) ---
            headers = ['Local', 'Fecha', 'Hora', 'Factura #', 'Bruto', 'Impuestos', 'Neto']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.fill = fill
                cell.font = font
                ws.column_dimensions[chr(64+col)].width = 15
            
            if not df.empty:
                # Ordenar: Nombre Local, Fecha, Hora
                df_sorted = df.sort_values(by=['nombre_local', 'fecha', 'hora'])
                
                row_idx = 5
                for _, row in df_sorted.iterrows():
                    ws.cell(row=row_idx, column=1, value=row.get('nombre_local'))
                    ws.cell(row=row_idx, column=2, value=row.get('fecha'))
                    ws.cell(row=row_idx, column=3, value=row.get('hora'))
                    ws.cell(row=row_idx, column=4, value=row.get('factura_no'))
                    ws.cell(row=row_idx, column=5, value=row.get('total_neto')).number_format = '$#,##0.00;[Red]-$#,##0.00'  # Bruto UI
                    ws.cell(row=row_idx, column=6, value=row.get('total_impuestos')).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    ws.cell(row=row_idx, column=7, value=row.get('total_bruto')).number_format = '$#,##0.00;[Red]-$#,##0.00' # Neto UI
                    row_idx += 1
        
        else:
            # --- VISTA RESUMIDA (Agrupada) ---
            headers = ['Local', 'Ventas Brutas (Base)', 'Impuestos', 'Ventas Netas (Total)']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[chr(64+col)].width = 20

            if not df.empty:
                resumen = df.groupby(['local_id', 'nombre_local']).agg({
                    'total_neto': 'sum',
                    'total_impuestos': 'sum',
                    'total_bruto': 'sum'
                }).reset_index().sort_values(by=['nombre_local'])
                
                row_idx = 5
                for _, row in resumen.iterrows():
                    ws.cell(row=row_idx, column=1, value=row['nombre_local'])
                    ws.cell(row=row_idx, column=2, value=row['total_neto']).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    ws.cell(row=row_idx, column=3, value=row['total_impuestos']).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    ws.cell(row=row_idx, column=4, value=row['total_bruto']).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    row_idx += 1
                
                # Totales
                ws.cell(row=row_idx, column=1, value='TOTAL').font = Font(bold=True)
                ws.cell(row=row_idx, column=2, value=resumen['total_neto'].sum()).number_format = '$#,##0.00;[Red]-$#,##0.00'
                ws.cell(row=row_idx, column=2).font = Font(bold=True)
                ws.cell(row=row_idx, column=3, value=resumen['total_impuestos'].sum()).number_format = '$#,##0.00;[Red]-$#,##0.00'
                ws.cell(row=row_idx, column=3).font = Font(bold=True)
                ws.cell(row=row_idx, column=4, value=resumen['total_bruto'].sum()).number_format = '$#,##0.00;[Red]-$#,##0.00'
                ws.cell(row=row_idx, column=4).font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def generate_sales_report_pdf(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        local_id: Optional[str] = None,
        report_type: str = 'detailed',
        mall_name: str = "CENTRO COMERCIAL MS MALL",
        mall_id: Optional[str] = None,
    ) -> io.BytesIO:
        data = self._fetch_sales_rows_paginated(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            select_fields='local_id, locales(nombre), fecha, total_neto, total_impuestos, total_bruto',
            local_id=local_id,
            mall_id=mall_id,
        )
        df = pd.DataFrame(data or [])
        if not df.empty:
            df['nombre_local'] = df['locales'].apply(lambda x: x.get('nombre') if x else '?')
        else:
            df = pd.DataFrame(columns=['local_id', 'nombre_local', 'total_neto', 'total_impuestos', 'total_bruto'])

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        title_style, subtitle_style, header_style, normal_style = self._get_pdf_styles()
        
        elements.append(Paragraph(f"CENTRO COMERCIAL {mall_name.upper()}", title_style))
        elements.append(Paragraph("REPORTE DE AUDITORÍA DE VENTAS", title_style))
        elements.append(Paragraph(f"Período: {fecha_inicio} - {fecha_fin}", subtitle_style))
        elements.append(Spacer(1, 12))

        # Resumen
        elements.append(Paragraph("RESUMEN EJECUTIVO", header_style))
        
        total_base = df['total_neto'].sum() if not df.empty else 0
        total_tax = df['total_impuestos'].sum() if not df.empty else 0
        total_final = df['total_bruto'].sum() if not df.empty else 0
        tx_count = len(df)
        
        summary_data = [
            ["Total Ventas Brutas (Base):", self._format_currency(total_base)],
            ["Total Impuestos:", self._format_currency(total_tax)],
            ["Total Ventas Netas:", self._format_currency(total_final)],
            ["Transacciones:", str(tx_count)]
        ]
        t_summary = Table(summary_data, colWidths=[200, 150])
        t_summary.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 24))
        
        # Tabla Resumen por Local
        if not df.empty:
            section_title = "RESUMEN POR LOCAL" if report_type == 'summary' else "DETALLE POR LOCAL"
            elements.append(Paragraph(section_title, header_style))
            resumen = df.groupby(['local_id', 'nombre_local']).agg({'total_neto':'sum', 'total_impuestos':'sum', 'total_bruto':'sum'}).reset_index()
            
            table_data = [['Local', 'Ventas Brutas', 'Impuestos', 'Ventas Netas']]
            for _, row in resumen.iterrows():
                table_data.append([
                    row['nombre_local'],
                    self._format_currency(row['total_neto']),
                    self._format_currency(row['total_impuestos']),
                    self._format_currency(row['total_bruto'])
                ])
            # Footer
            table_data.append(['TOTAL', self._format_currency(total_base), self._format_currency(total_tax), self._format_currency(total_final)])
            
            t = Table(table_data, colWidths=[150, 100, 100, 100], repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4788')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 1, colors.grey),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#C5D9F1')),
                ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("RESUMEN POR LOCAL", header_style))
            elements.append(Paragraph("No hay ventas para el período seleccionado.", normal_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    async def generate_missing_days_report_excel(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        mall_id: str,
        local_id: Optional[str] = None
    ) -> io.BytesIO:
        dataset = self._build_missing_days_dataset(fecha_inicio, fecha_fin, mall_id, local_id)
        summary_rows = dataset['summary_rows']
        detail_rows = dataset['detail_rows']
        is_local_mode = dataset['is_local_mode']
        selected_local_name = dataset['selected_local_name'] or "Local"

        wb = Workbook()
        ws = wb.active
        ws.title = "Dias Faltantes"
        fill, font = self._get_header_style()

        ws['A1'] = 'REPORTE DE DÍAS FALTANTES (AUDITORÍA)'
        ws['A2'] = f"Período: {fecha_inicio} al {fecha_fin}"
        ws['A3'] = f"Alcance: {selected_local_name if is_local_mode else 'Todos los locales con brechas'}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'].font = Font(size=11, bold=False)
        ws['A3'].font = Font(size=11, bold=False)

        if is_local_mode:
            headers = ['Fecha Faltante']
            data_rows = [[r['fecha_faltante']] for r in detail_rows]
        else:
            headers = ['Local', 'Rubro', 'Días Faltantes', 'Días del Periodo', '% Cumplimiento', 'Estado']
            data_rows = [[
                r['local_nombre'],
                r['rubro'],
                r['dias_faltantes_count'],
                r['dias_totales_periodo'],
                r['porcentaje_cumplimiento'],
                r['estado'],
            ] for r in summary_rows]

        header_row = 5
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + min(col, 26))].width = 20

        row_idx = header_row + 1
        if not data_rows:
            ws.cell(row=row_idx, column=1, value='No se encontraron días faltantes en el período seleccionado.')
        else:
            for row in data_rows:
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1

        if not is_local_mode:
            ws2 = wb.create_sheet("Detalle Días")
            headers2 = ['Local', 'Fecha Faltante']
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.fill = fill
                cell.font = font
                ws2.column_dimensions[chr(64 + col)].width = 28 if col == 1 else 18
            r = 2
            if not detail_rows:
                ws2.cell(row=r, column=1, value='Sin brechas')
            else:
                for row in detail_rows:
                    ws2.cell(row=r, column=1, value=row['local_nombre'])
                    ws2.cell(row=r, column=2, value=row['fecha_faltante'])
                    r += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def generate_missing_days_report_pdf(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        mall_id: str,
        local_id: Optional[str] = None,
        mall_name: str = "CENTRO COMERCIAL MS MALL"
    ) -> io.BytesIO:
        dataset = self._build_missing_days_dataset(fecha_inicio, fecha_fin, mall_id, local_id)
        summary_rows = dataset['summary_rows']
        detail_rows = dataset['detail_rows']
        is_local_mode = dataset['is_local_mode']
        selected_local_name = dataset['selected_local_name'] or "Local"

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        title_style, subtitle_style, header_style, normal_style = self._get_pdf_styles()

        elements.append(Paragraph(f"CENTRO COMERCIAL {mall_name.upper()}", title_style))
        elements.append(Paragraph("REPORTE DE DÍAS FALTANTES", title_style))
        elements.append(Paragraph(f"Período: {fecha_inicio} - {fecha_fin}", subtitle_style))
        alcance = selected_local_name if is_local_mode else "Todos los locales con brechas"
        elements.append(Paragraph(f"Alcance: {alcance}", normal_style))
        elements.append(Spacer(1, 12))

        if is_local_mode:
            elements.append(Paragraph("DÍAS FALTANTES", header_style))
            table_data = [['Fecha Faltante']]
            if detail_rows:
                for row in detail_rows:
                    table_data.append([row['fecha_faltante']])
            else:
                table_data.append(['No se encontraron días faltantes en el período seleccionado.'])

            t = Table(table_data, colWidths=[400])
        else:
            elements.append(Paragraph("RESUMEN POR LOCAL", header_style))
            table_data = [['Local', 'Rubro', 'Días Falt.', '% Cumpl.']]
            if summary_rows:
                for row in summary_rows:
                    table_data.append([
                        row['local_nombre'],
                        row['rubro'],
                        str(row['dias_faltantes_count']),
                        f"{row['porcentaje_cumplimiento']}%",
                    ])
            else:
                table_data.append(['Sin brechas', '-', '0', '100%'])

            t = Table(table_data, colWidths=[170, 100, 70, 70])

        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(t)

        if (not is_local_mode) and detail_rows:
            elements.append(Spacer(1, 16))
            elements.append(Paragraph("DETALLE DE FECHAS FALTANTES (primeros 120 registros)", header_style))
            detail_table = [['Local', 'Fecha Faltante']]
            for row in detail_rows[:120]:
                detail_table.append([row['local_nombre'], row['fecha_faltante']])
            dt = Table(detail_table, colWidths=[230, 120])
            dt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C5D9F1')),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(dt)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    # --- SALES CUBE ---

    async def generate_sales_cube_excel(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        agrupacion: str,
        metrica: str,
        mall_id: str = None,
        local_id: str = None,
        custom_dimension_key: Optional[str] = None,
        custom_filters: Optional[Dict[str, Any]] = None,
    ) -> io.BytesIO:
        # Same logic as get_sales_cube in main.py but exporting
        stores_query = self.supabase.table("locales").select("id, nombre")
        if mall_id:
            stores_query = stores_query.eq("mall_id", mall_id)
        if local_id:
            stores_query = stores_query.eq("id", local_id)
        stores = stores_query.execute().data
        store_map = {str(s['id']): s['nombre'] for s in stores}
        local_ids = list(store_map.keys())
        snapshot = self.local_custom_fields_service.build_snapshot(mall_id, local_ids, include_inactive=False) if mall_id else self.local_custom_fields_service.build_snapshot("", [], include_inactive=False)
        if custom_filters:
            local_ids = self.local_custom_fields_service.filter_local_ids_by_custom_filters(local_ids, snapshot, custom_filters)

        sales = []
        if local_ids:
            sales_rows = (
                fetch_sales_cube_daily_aggregates(
                    self.supabase,
                    mall_id=mall_id,
                    local_ids=local_ids,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                )
                if mall_id
                else None
            )
            if sales_rows is None:
                sales_rows = fetch_sales_rows_keyset(
                    self.supabase,
                    select_fields=(
                        "id,local_id,fecha,total_bruto,total_neto,total_impuestos"
                    ),
                    local_id=local_id,
                    local_ids=None if local_id else local_ids,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                )
            sales.extend(self._normalize_sale_totals_row(dict(row)) for row in sales_rows)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz de Ventas"
        
        ws['A1'] = "MATRIZ DE VENTAS"
        ws['A2'] = f"Agrupación: {agrupacion} | Métrica: {metrica}"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not sales:
            ws['A4'] = "No data found"
        else:
            df = pd.DataFrame(sales)
            df['local_id'] = df['local_id'].astype(str)
            df['local_nombre'] = df['local_id'].map(store_map).fillna(df['local_id'])
            cube = self.local_custom_fields_service.build_cube_response(
                sales_df=df,
                grouping=agrupacion.upper(),
                metric=metrica,
                start_date=fecha_inicio,
                end_date=fecha_fin,
                snapshot=snapshot,
                custom_dimension_key=custom_dimension_key,
            )

            columns = list(cube.get('columns') or [])
            rows = list(cube.get('data') or [])
            grand_totals = dict(cube.get('grand_totals') or {})

            if not columns:
                columns = ['local_nombre', 'TOTAL_FILA']

            fill, font = self._get_header_style()

            def _display_col_label(col_name: str) -> str:
                if col_name in {'local_nombre', 'row_label'}:
                    return str(cube.get('row_label') or 'LOCAL').upper()
                if col_name == 'TOTAL_FILA':
                    return 'TOTAL'
                return str(col_name)

            def _coerce_excel_value(value: Any):
                if value is None:
                    return None
                if hasattr(value, "item"):
                    try:
                        return value.item()
                    except Exception:
                        pass
                return value

            # Write headers (same order shown by UI)
            for col_idx, col_name in enumerate(columns, start=1):
                c = ws.cell(row=4, column=col_idx, value=_display_col_label(col_name))
                c.fill = fill
                c.font = font
                c.alignment = Alignment(horizontal='center')

            # Write matrix data
            number_format = '$#,##0.00' if metrica != 'transacciones' else '0'
            row_idx = 5
            for row in rows:
                for col_idx, col_name in enumerate(columns, start=1):
                    value = _coerce_excel_value(row.get(col_name))
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1:
                        cell.font = Font(bold=True)
                    else:
                        if value in (None, ""):
                            cell.value = 0
                        cell.number_format = number_format
                        cell.alignment = Alignment(horizontal='right')
                        if col_name == 'TOTAL_FILA' or row.get('row_type') == 'group':
                            cell.font = Font(bold=True)
                row_idx += 1

            # Grand totals row (same semantics as UI footer)
            total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            for col_idx, col_name in enumerate(columns, start=1):
                if col_idx == 1:
                    cell = ws.cell(row=row_idx, column=col_idx, value='TOTAL GENERAL')
                    cell.font = Font(bold=True)
                else:
                    total_value = _coerce_excel_value(grand_totals.get(col_name, 0))
                    cell = ws.cell(row=row_idx, column=col_idx, value=total_value)
                    cell.number_format = number_format
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right')
                cell.fill = total_fill

            # Basic usability formatting
            ws.freeze_panes = "B5"
            ws.column_dimensions['A'].width = 34
            for col_idx in range(2, len(columns) + 1):
                col_letter = ws.cell(row=4, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 16

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # --- FINANCIAL DASHBOARD ---

    async def _get_financial_data(self, fecha_inicio, fecha_fin):
        # Logic from getKPIs to calculate OCR
        sales = fetch_sales_rows_keyset(
            self.supabase,
            select_fields="local_id,total_bruto,total_neto",
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        )
        stores = self.supabase.table("locales").select("*").execute().data
        store_map = {str(s['id']): s for s in stores}
        
        agg = {}
        for s in sales:
            lid = str(s['local_id'])
            if lid not in agg: agg[lid] = {'neto':0, 'bruto':0}
            agg[lid]['neto'] += (s['total_neto'] or 0)
            agg[lid]['bruto'] += (s['total_bruto'] or 0) # This is Total
        
        results = []
        for lid, store in store_map.items():
            vals = agg.get(lid, {'neto':0, 'bruto':0})
            venta_neta_total = vals['bruto'] # USER TERM: Netas = Total
            # Previous logic in frontend used 'venta' = salesMap value.
            # Usually salesMap in frontend was populated by sum of bruto? Check main.py dashboard logic.
            # main.py dashboard logic: sales_by_store[s_name] += bruto.
            # And frontend calls it "Venta Actual".
            # So "Venta Netas" (User Term) = Bruto (DB).
            
            renta_fija = float(store.get('renta_fija') or 0)
            mts = float(store.get('mts') or 1)
            breakpoint = float(store.get('breakpoint_venta') or 0)
            pct = float(store.get('porcentaje_variable') or store.get('porciento_renta') or 0)
            
            ocr = (renta_fija / venta_neta_total * 100) if venta_neta_total > 0 else 0
            
            renta_var = 0
            if venta_neta_total > breakpoint:
                renta_var = (venta_neta_total * pct / 100) - renta_fija
                if renta_var < 0: renta_var = 0
            
            status = "Saludable"
            if ocr > 20: status = "Riesgo"
            elif ocr > 15: status = "Atención"
            
            results.append({
                "nombre": store['nombre'],
                "ventas": venta_neta_total,
                "renta_fija": renta_fija,
                "breakpoint": breakpoint,
                "renta_var": renta_var,
                "ocr": ocr,
                "venta_m2": venta_neta_total / mts,
                "estado": status
            })
        return results

    async def generate_financial_dashboard_excel(self, fecha_inicio: str, fecha_fin: str) -> io.BytesIO:
        data = await self._get_financial_data(fecha_inicio, fecha_fin)
        wb = Workbook()
        ws = wb.active
        ws.title = "Salud de Cartera"
        
        ws['A1'] = "SALUD DE CARTERA - ANÁLISIS FINANCIERO"
        ws['A2'] = f"Período: {fecha_inicio} al {fecha_fin}"
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Local', 'Ventas Netas', 'Renta Fija', 'Break-point', 'Renta Variable', 'OCR (%)', 'Venta/m2', 'Estado']
        fill, font = self._get_header_style()
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = fill
            c.font = font
            ws.column_dimensions[chr(64+col)].width = 15
            
        row_idx = 5
        for item in data:
            ws.cell(row=row_idx, column=1, value=item['nombre'])
            ws.cell(row=row_idx, column=2, value=item['ventas']).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=3, value=item['renta_fija']).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=4, value=item['breakpoint']).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=5, value=item['renta_var']).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=6, value=item['ocr']/100).number_format = '0.00%'
            ws.cell(row=row_idx, column=7, value=item['venta_m2']).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=8, value=item['estado'])
            
            if item['ocr'] > 20:
                ws.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row_idx += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def generate_financial_dashboard_pdf(self, fecha_inicio: str, fecha_fin: str) -> io.BytesIO:
        data = await self._get_financial_data(fecha_inicio, fecha_fin)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        title_s, sub_s, head_s, norm_s = self._get_pdf_styles()
        
        elements.append(Paragraph("SALUD DE CARTERA", title_s))
        elements.append(Paragraph(f"{fecha_inicio} - {fecha_fin}", sub_s))
        
        table_data = [['Local', 'Ventas Netas', 'Renta Fija', 'Var', 'OCR %', 'Estado']]
        for d in data:
            table_data.append([
                d['nombre'],
                self._format_currency(d['ventas']),
                self._format_currency(d['renta_fija']),
                self._format_currency(d['renta_var']),
                f"{d['ocr']:.2f}%",
                d['estado']
            ])
            
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4788')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
