import asyncio

from openpyxl import load_workbook

import services.export_service as export_service_module
from services.export_service import ExportService


class _FakeResponse:
    def __init__(self, data=None):
        self.data = data


class _TableQuery:
    def __init__(self, supabase, table_name):
        self.supabase = supabase
        self.table_name = table_name
        self._filters = []
        self._order = None
        self._range = None
        self._limit = None

    def select(self, *_args, **_kwargs):
        return self

    def eq(self, key, value):
        self._filters.append(("eq", key, value))
        return self

    def gte(self, key, value):
        self._filters.append(("gte", key, value))
        return self

    def lte(self, key, value):
        self._filters.append(("lte", key, value))
        return self

    def gt(self, key, value):
        self._filters.append(("gt", key, value))
        return self

    def order(self, column, desc=False):
        self._order = (column, bool(desc))
        return self

    def range(self, start, end):
        self._range = (int(start), int(end))
        return self

    def limit(self, value):
        self._limit = int(value)
        return self

    def _apply_filters(self, rows):
        result = list(rows)
        for op, key, value in self._filters:
            if op == "eq":
                result = [r for r in result if r.get(key) == value]
            elif op == "gte":
                result = [r for r in result if r.get(key) is not None and r.get(key) >= value]
            elif op == "lte":
                result = [r for r in result if r.get(key) is not None and r.get(key) <= value]
            elif op == "gt":
                result = [r for r in result if r.get(key) is not None and r.get(key) > value]
        return result

    def execute(self):
        rows = [dict(r) for r in self.supabase.tables.get(self.table_name, [])]
        rows = self._apply_filters(rows)

        if self._order:
            col, desc = self._order
            rows = sorted(rows, key=lambda r: (r.get(col) is None, r.get(col)), reverse=desc)

        if self._range is not None:
            start, end = self._range
            rows = rows[start:end + 1]
        if self._limit is not None:
            rows = rows[:self._limit]
        if self.table_name == "ventas":
            # Simula límite duro por request de PostgREST/Supabase, incluso con range.
            rows = rows[:1000]

        return _FakeResponse(rows)


class _FakeSupabase:
    def __init__(self, tables=None):
        self.tables = tables or {}

    def table(self, table_name):
        return _TableQuery(self, table_name)


def _build_sales_rows():
    rows = []
    seq = 1

    def add_rows(*, count, mall_id, local_id, local_name):
        nonlocal seq
        for _ in range(count):
            rows.append({
                "id": f"{seq:06d}",
                "mall_id": mall_id,
                "local_id": local_id,
                "locales": {"nombre": local_name},
                "fecha": "2026-01-15",
                "hora": "12:00:00",
                "factura_no": f"F-{seq}",
                "total_neto": 100.0,
                "total_impuestos": 18.0,
                "total_bruto": 118.0,
            })
            seq += 1

    # Primeras 1000 filas: solo Alpha y Beta.
    add_rows(count=500, mall_id="mall-1", local_id="l-alpha", local_name="Alpha")
    add_rows(count=500, mall_id="mall-1", local_id="l-beta", local_name="Beta")

    # Estas filas quedan fuera si no hay paginación.
    add_rows(count=10, mall_id="mall-1", local_id="l-gamma", local_name="Gamma")
    add_rows(count=10, mall_id="mall-1", local_id="l-delta", local_name="Delta")

    # Otro mall para validar aislamiento.
    add_rows(count=5, mall_id="mall-2", local_id="l-omega", local_name="Omega")

    return rows


def _read_summary_locals_from_excel(workbook_bytes):
    wb = load_workbook(workbook_bytes)
    ws = wb["Reporte de Ventas"]
    locals_found = []

    row_idx = 5
    while row_idx <= ws.max_row:
        local_name = ws.cell(row=row_idx, column=1).value
        if local_name == "TOTAL":
            break
        if local_name:
            locals_found.append(str(local_name))
        row_idx += 1

    return locals_found


def test_generate_sales_report_excel_summary_includes_all_locals_with_pagination_and_mall_filter():
    fake_supabase = _FakeSupabase({"ventas": _build_sales_rows()})
    service = ExportService(fake_supabase)

    result = asyncio.run(
        service.generate_sales_report_excel(
            fecha_inicio="2026-01-01",
            fecha_fin="2026-01-31",
            report_type="summary",
            mall_id="mall-1",
        )
    )

    locals_found = _read_summary_locals_from_excel(result)

    assert set(locals_found) == {"Alpha", "Beta", "Gamma", "Delta"}
    assert "Omega" not in locals_found


def test_generate_sales_report_pdf_summary_includes_local_table(monkeypatch):
    captured_tables = []
    original_table = export_service_module.Table

    def _spy_table(data, *args, **kwargs):
        captured_tables.append(data)
        return original_table(data, *args, **kwargs)

    monkeypatch.setattr(export_service_module, "Table", _spy_table)

    fake_supabase = _FakeSupabase({"ventas": _build_sales_rows()})
    service = ExportService(fake_supabase)

    pdf_buffer = asyncio.run(
        service.generate_sales_report_pdf(
            fecha_inicio="2026-01-01",
            fecha_fin="2026-01-31",
            report_type="summary",
            mall_name="Mall Test",
            mall_id="mall-1",
        )
    )

    assert len(pdf_buffer.getvalue()) > 0

    detail_table = next(
        (
            table
            for table in captured_tables
            if table and table[0] == ["Local", "Ventas Brutas", "Impuestos", "Ventas Netas"]
        ),
        None,
    )
    assert detail_table is not None

    locals_in_table = {str(row[0]) for row in detail_table[1:] if row and row[0] != "TOTAL"}
    assert {"Alpha", "Beta", "Gamma", "Delta"}.issubset(locals_in_table)
    assert "Omega" not in locals_in_table
